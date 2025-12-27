"""Legacy docstring removed for English-only repo."""
import logging
import time
import sys
import os
import csv
import openpyxl
from openpyxl import Workbook
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, Set
import atexit
import threading
import signal
from contextlib import contextmanager
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from pathlib import Path
from filelock import FileLock
from datetime import datetime, timedelta, timezone

# database
from database.ingest_task_info import log_task_info_to_db
from database.mysql_reader import creator_has_reply

# NOTE: legacy comment removed for English-only repo.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models.email_code import GmailVerificationCode

# NOTE: legacy comment removed for English-only repo.
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # NOTE: legacy comment removed for English-only repo.
TASK_DIR = os.path.join(BASE_DIR, "task")

CATEGORY_NAME_TO_ID = {
    "Home Supplies": "600001",
    "Kitchenware": "600024",
    "Textiles & Soft Furnishings": "600154",
    "Household Appliances": "600942",
    "Womenswear & Underwear": "601152",
    "Shoes": "601352",
    "Beauty & Personal Care": "601450",
    "Phones & Electronics": "601739",
    "Computers & Office Equipment": "601755",
    "Pet Supplies": "602118",
    "Sports & Outdoor": "603014",
    "Toys": "604206",
    "Furniture": "604453",
    "Tools & Hardware": "604579",
    "Home Improvement": "604968",
    "Automotive & Motorcycle": "605196",
    "Fashion Accessories": "605248",
    "Food & Beverages": "700437",
    "Health": "700645",
    "Books, Magazines & Audio": "801928",
    "Kids' Fashion": "802184",
    "Menswear & Underwear": "824328",
    "Luggage & Bags": "824584",
    "Collections": "951432",
    "Jewellery Accessories & Derivatives": "953224",
}
CATEGORY_ID_TO_NAME = {v: k for k, v in CATEGORY_NAME_TO_ID.items()}
CATEGORY_NAME_TO_ID_LOWER = {k.lower(): v for k, v in CATEGORY_NAME_TO_ID.items()}


def update_task_info_row(task_id: str, updates: Dict[str, Any]) -> None:
    """Legacy docstring removed for English-only repo."""
    info_path = Path("data/task_info.xlsx")
    if not info_path.exists():
        return

    lock = FileLock(str(info_path) + ".lock")
    with lock:
        try:
            workbook = openpyxl.load_workbook(info_path)
            sheet = workbook.active
        except Exception:
            return

        header = [cell.value for cell in sheet[1]]
        if "task_id" not in header:
            workbook.close()
            return

        column_map = {name: idx for idx, name in enumerate(header)}
        task_col = column_map["task_id"] + 1
        target_row = None
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row_idx, task_col).value
            if cell_value is not None and str(cell_value) == str(task_id):
                target_row = row_idx
                break

        if target_row is None:
            workbook.close()
            return

        for key, value in updates.items():
            if key not in column_map:
                continue
            column = column_map[key] + 1
            if isinstance(value, datetime):
                sheet.cell(target_row, column).value = value.isoformat()
            elif isinstance(value, list):
                sheet.cell(target_row, column).value = ", ".join(str(v) for v in value)
            else:
                sheet.cell(target_row, column).value = value

        workbook.save(info_path)
        workbook.close()


class CrawlerCancelledError(Exception):
    """Legacy docstring removed for English-only repo."""


class CreatorFullCrawler:
    """Legacy docstring removed for English-only repo."""

    def __init__(self, search_strategy=None, max_creators_to_load=500,
                task_id=None, task_dir=None, shared_record_callback=None,
                account_info=None, cancel_event=None, task_metadata=None,
                shared_seen_creators: Optional[Set[str]] = None,
                shared_skipped_creators: Optional[Set[str]] = None):
        
        # NOTE: legacy comment removed for English-only repo.
        self.task_id = task_id or datetime.now().strftime("%Y%m%d_%H%M%S")
        self.task_dir = Path(task_dir) if task_dir else Path(f"data/tasks/{self.task_id}")
        self.task_dir.mkdir(parents=True, exist_ok=True)
        self.cancel_event = cancel_event
        self.was_cancelled = False
        self.console_logging_enabled = os.getenv("CRAWLER_STDOUT_LOG", "0") == "1"
        self.task_metadata = task_metadata or {}
        self._task_info_saved = False
        self.shared_record_callback = shared_record_callback
        self.latest_creator_name: Optional[str] = None
        self.setup_task_logging()

        # NOTE: legacy comment removed for English-only repo.
        self.brand_name = ""
        self.region = ""
        self.only_first = 0
        try:
            cfg_path = self.task_dir / "dify_out.txt"
            if os.path.exists(cfg_path):
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                self.brand_name = cfg.get("brand", {}).get("name", "") or ""
                self.region = cfg.get("region", "")
                self.only_first = int(cfg.get("brand", {}).get("only_first", 0))
                self.logger.info(f"配置 - 品牌名: {self.brand_name}, 地区: {self.region}, only_first: {self.only_first}")
        except Exception as e:
            self.logger.warning(f"读取配置失败: {e}")
        
        # NOTE: legacy comment removed for English-only repo.
        if account_info:
            self.login_email = account_info.get('login_email')
            self.login_password = account_info.get('login_password')
            self.gmail_username = account_info.get('gmail_username')
            self.gmail_app_password = account_info.get('gmail_app_password')
            self.account_name = account_info.get('name', 'Unknown')
            self.account_id = account_info.get('id', -1)
            
            self.logger.info(f"使用账号: {self.account_name} ({self.login_email})")
        else:
            # Use a default placeholder account
            self.login_email = "user@example.com"
            self.login_password = "<ACCOUNT_PASSWORD>"
            self.gmail_username = "user@example.com"
            self.gmail_app_password = "<GMAIL_APP_PASSWORD>"
            self.account_name = "Default account"
            self.account_id = -1
            
            self.logger.warning("No account info provided; using default account.")

        # NOTE: legacy comment removed for English-only repo.
        self.gmail_verifier = GmailVerificationCode(
            username=self.gmail_username,
            app_password=self.gmail_app_password
        )

        # NOTE: legacy comment removed for English-only repo.
        self.target_url = "https://partner.tiktokshop.com/affiliate-cmp/creator?market=19"
        
        # NOTE: legacy comment removed for English-only repo.
        if search_strategy:
            self.search_strategy = search_strategy
        else:
            self.search_strategy = {
                "search_keywords": "",
                "fans_age_range": [],
                "fans_gender": "",
                "content_type": [],
                "sales": [],
                "min_engagement_rate": 0,
                "min_fans": 800,
                "avg_views": 800,
                "gmv": []
            }
        
        # NOTE: legacy comment removed for English-only repo.
        self.parse_search_strategy()

        # NOTE: legacy comment removed for English-only repo.
        self.max_creators_to_load = max_creators_to_load
        
        # NOTE: legacy comment removed for English-only repo.
        self.shop_name = "GOKOCO.MX"

        # NOTE: legacy comment removed for English-only repo.
        self.task_id = task_id or datetime.now().strftime("%Y%m%d_%H%M%S")
        self.task_dir = Path(task_dir) if task_dir else Path(f"data/tasks/{self.task_id}")
        self.task_dir.mkdir(parents=True, exist_ok=True)
        
        
        # NOTE: legacy comment removed for English-only repo.
        self.csv_data_dir = self.task_dir / "xlsx_csv"
        self.screenshots_dir = self.task_dir / "screenshots"
        self.csv_data_dir.mkdir(exist_ok=True)
        self.screenshots_dir.mkdir(exist_ok=True)

        # NOTE: legacy comment removed for English-only repo.
        self.all_creators_data = []
        self.csv_filename = None
        self.xlsx_filename = None

        # NOTE: legacy comment removed for English-only repo.
        self.creator_counter = 0
        self.max_creators = int(os.getenv("MAX_CREATORS", "500"))
        
        # NOTE: legacy comment removed for English-only repo.
        self.existing_data = self.load_existing_excel()
        self.processed_creators = set()
        self.creator_reply_cache: Dict[str, bool] = {}

        # NOTE: legacy comment removed for English-only repo.
        self.blacklist_creators = self.load_blacklist()

        # NOTE: legacy comment removed for English-only repo.
        self.browser = None
        self.context = None
        self.main_page = None
        self.active_pages = set()
        self.cleanup_registered = False
        self.is_cleaned_up = False
        self.playwright_context_active = False
        self.runtime_limit_seconds = int(os.getenv("CRAWLER_RUNTIME_LIMIT_SECONDS", "9000"))
        self.internal_cancel_event = threading.Event()
        self._timeout_timer: Optional[threading.Timer] = None
        self.timeout_reason: Optional[str] = None
        self.run_started_at: Optional[float] = None
        self._cancel_monitor_thread: Optional[threading.Thread] = None
        self.abort_initiated = False
        self.browser_force_killed = False
        
        # NOTE: legacy comment removed for English-only repo.
        self.new_processed_count = 0
        self.target_new_count = 30
        self.refresh_after_new = int(os.getenv("CRAWLER_BROWSER_REFRESH_AFTER_NEW", "30"))
        self.skip_restart_threshold = int(os.getenv("CRAWLER_SKIP_RESTART_THRESHOLD", "60"))
        self.skipped_total = 0
        self.restart_requested = False
        self.restart_reason: str = ""
        self.shared_seen_creators = shared_seen_creators
        self.shared_skipped_creators = shared_skipped_creators

        # NOTE: legacy comment removed for English-only repo.
        self.register_cleanup()
    
    def setup_task_logging(self):
        """Legacy docstring removed for English-only repo."""
        log_file = self.task_dir / "crawler.log"
        
        # NOTE: legacy comment removed for English-only repo.
        self.logger = logging.getLogger(f'crawler_{self.task_id}')
        self.logger.setLevel(logging.INFO)
        
        # NOTE: legacy comment removed for English-only repo.
        self.logger.handlers.clear()
        
        # NOTE: legacy comment removed for English-only repo.
        file_handler = logging.FileHandler(log_file)
        file_handler.setFormatter(
            logging.Formatter(f'[{self.task_id}] %(asctime)s - %(levelname)s - %(message)s')
        )
        self.logger.addHandler(file_handler)
        
        if self.console_logging_enabled:
            console_handler = logging.StreamHandler()
            console_handler.setFormatter(
                logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            )
            self.logger.addHandler(console_handler)
        
        # NOTE: legacy comment removed for English-only repo.
        self.logger.propagate = False
        
        self.logger.info(f"任务日志已配置到: {log_file}")

    def _check_runtime_limit(self):
        """Legacy docstring removed for English-only repo."""
        if self.runtime_limit_seconds <= 0 or self.run_started_at is None:
            return
        elapsed = time.monotonic() - self.run_started_at
        if elapsed >= self.runtime_limit_seconds and not self.internal_cancel_event.is_set():
            minutes = self.runtime_limit_seconds // 60
            self._force_abort(f"运行超出 {minutes} 分钟上限，准备终止")

    def _trigger_runtime_timeout(self):
        """Legacy docstring removed for English-only repo."""
        minutes = self.runtime_limit_seconds // 60 if self.runtime_limit_seconds >= 60 else self.runtime_limit_seconds
        if not self.internal_cancel_event.is_set():
            self._force_abort(f"运行时间超过 {minutes} 分钟限制，触发超时取消")

    def _start_runtime_timer(self):
        """Legacy docstring removed for English-only repo."""
        if self.runtime_limit_seconds <= 0:
            self.logger.info("未配置运行时限，上限监控跳过")
            self.run_started_at = time.monotonic()
            return

        self.run_started_at = time.monotonic()
        self.timeout_reason = None
        self.internal_cancel_event.clear()

        if self._timeout_timer:
            self._timeout_timer.cancel()

        self._timeout_timer = threading.Timer(self.runtime_limit_seconds, self._trigger_runtime_timeout)
        self._timeout_timer.daemon = True
        self._timeout_timer.start()
        minutes = self.runtime_limit_seconds // 60 if self.runtime_limit_seconds >= 60 else self.runtime_limit_seconds
        self.logger.info(f"已启用运行时限监控：最长 {minutes} 分钟")

    def _start_cancel_watchdog(self):
        """Legacy docstring removed for English-only repo."""
        if not self.cancel_event:
            return
        if self._cancel_monitor_thread and self._cancel_monitor_thread.is_alive():
            return

        def _watch_cancel():
            try:
                self.cancel_event.wait()
            except Exception:
                return
            if not self.cancel_event.is_set():
                return
            self.logger.info("检测到外部取消请求，准备立即停止任务")
            self.was_cancelled = True
            if not self.timeout_reason:
                self.timeout_reason = "cancelled"
            self.internal_cancel_event.set()
            self._kill_browser_process("external cancel")

        self._cancel_monitor_thread = threading.Thread(
            target=_watch_cancel,
            name=f"cancel-watch-{self.task_id}",
            daemon=True,
        )
        self._cancel_monitor_thread.start()

    def _force_abort(self, reason: str):
        """Legacy docstring removed for English-only repo."""
        if self.abort_initiated:
            return
        self.abort_initiated = True
        self.timeout_reason = reason
        self.was_cancelled = True
        self.logger.warning(reason)
        self.internal_cancel_event.set()
        if self.cancel_event:
            try:
                self.cancel_event.set()
            except Exception:
                pass
        self._kill_browser_process(reason)

    def _record_skip(self, creator_name: str, reason: str) -> bool:
        """Legacy docstring removed for English-only repo."""
        if self.shared_skipped_creators is not None:
            if creator_name in self.shared_skipped_creators:
                self.logger.debug(f"{creator_name} 已在全局跳过列表中，跳过计数不再增加")
                return False
            self.shared_skipped_creators.add(creator_name)

        self.skipped_total += 1
        if self.skip_restart_threshold > 0 and self.skipped_total >= self.skip_restart_threshold:
            self.restart_requested = True
            self.restart_reason = f"累计跳过 {self.skipped_total} 个达人 ({reason})"
            self.logger.warning(
                f"跳过达人已达 {self.skipped_total} 个，触发浏览器重启请求：{reason}"
            )
            return True
        return False

    def _should_restart_after_new(self) -> bool:
        if self.refresh_after_new <= 0:
            return False
        if self.new_processed_count >= self.refresh_after_new:
            self.restart_requested = True
            self.restart_reason = f"单批新增 {self.new_processed_count} 个达人"
            self.logger.info(
                "单批新增达人达到 %s，上报浏览器重启请求",
                self.new_processed_count,
            )
            return True
        return False

    def _abort_due_to_cancel(self, detail: str = ""):
        """Legacy docstring removed for English-only repo."""
        reason = detail or self.timeout_reason or "cancelled"
        if not self.timeout_reason:
            self.timeout_reason = reason
        self.was_cancelled = True
        self.logger.info(f"检测到取消请求，准备终止任务 {reason}".strip())
        self._kill_browser_process(reason)
        if not self.is_cleaned_up:
            try:
                self.emergency_cleanup()
            except Exception as exc:  # NOTE: legacy comment removed for English-only repo.
                self.logger.warning(f"取消后的清理失败: {exc}")
        raise CrawlerCancelledError(reason)

    def raise_if_cancelled(self, stage: str = ""):
        """Legacy docstring removed for English-only repo."""
        self._check_runtime_limit()
        if self.internal_cancel_event.is_set():
            detail = self.timeout_reason or (f"阶段: {stage}" if stage else "")
            self._abort_due_to_cancel(detail)

        if self.cancel_event and self.cancel_event.is_set():
            detail = f"阶段: {stage}" if stage else "cancelled"
            self._abort_due_to_cancel(detail)

    def ensure_task_info_saved(self):
        if not self._task_info_saved:
            self.save_task_info()
            self._task_info_saved = True

    # NOTE: legacy comment removed for English-only repo.
    def save_task_info(self):
        try:
            info_path = Path("data/task_info.xlsx")
            info_path.parent.mkdir(parents=True, exist_ok=True)
            lock = FileLock(str(info_path) + ".lock")
            metadata = self.task_metadata or {}
            task_id = str(metadata.get("task_id") or self.task_id)

            headers = [
                "task_id",
                "task_name",
                "campaign_id",
                "campaign_name",
                "product_id",
                "product_name",
                "region",
                "brand",
                "only_first",
                "task_type",
                "search_keywords",
                "product_category",
                "fans_age_range",
                "fans_gender",
                "min_fans",
                "content_type",
                "gmv",
                "sales",
                "min_GMV",
                "max_GMV",
                "avg_views",
                "min_engagement_rate",
                "email_first_subject",
                "email_first_body",
                "email_later_subject",
                "email_later_body",
                "target_new_creators",
                "max_creators",
                "run_at_time",
                "run_end_time",
                "run_time",
                "task_directory",
                "created_at",
            ]

            row_data = {
                "task_id": task_id,
                "task_name": metadata.get("task_name"),
                "campaign_id": metadata.get("campaign_id"),
                "campaign_name": metadata.get("campaign_name"),
                "product_id": metadata.get("product_id"),
                "product_name": metadata.get("product_name"),
                "region": metadata.get("region", self.region),
                "brand": metadata.get("brand", {}).get("name", self.brand_name),
                "only_first": metadata.get("brand", {}).get("only_first", self.only_first),
                "task_type": metadata.get("task_type", "Connect"),
                "search_keywords": metadata.get("search_strategy", {}).get("search_keywords", ""),
                "product_category": metadata.get("search_strategy", {}).get("product_category", ""),
                "fans_age_range": metadata.get("search_strategy", {}).get("fans_age_range", []),
                "fans_gender": metadata.get("search_strategy", {}).get("fans_gender", ""),
                "min_fans": metadata.get("search_strategy", {}).get("min_fans", ""),
                "content_type": metadata.get("search_strategy", {}).get("content_type", []),
                "gmv": metadata.get("search_strategy", {}).get("gmv", self.search_strategy.get("gmv", [])),
                "sales": metadata.get("search_strategy", {}).get("sales", self.search_strategy.get("sales", [])),
                "min_GMV": metadata.get("search_strategy", {}).get("min_GMV", ""),
                "max_GMV": metadata.get("search_strategy", {}).get("max_GMV", ""),
                "avg_views": metadata.get("search_strategy", {}).get("avg_views", ""),
                "min_engagement_rate": metadata.get("search_strategy", {}).get("min_engagement_rate", ""),
                "email_first_subject": metadata.get("email_first", {}).get("subject", ""),
                "email_first_body": metadata.get("email_first", {}).get("email_body", ""),
                "email_later_subject": metadata.get("email_later", {}).get("subject", ""),
                "email_later_body": metadata.get("email_later", {}).get("email_body", ""),
                "target_new_creators": metadata.get("target_new_creators", self.target_new_count),
                "max_creators": metadata.get("max_creators", self.max_creators),
                "run_at_time": metadata.get("run_at_time"),
                "run_end_time": metadata.get("run_end_time"),
                "connect_creator": "",
                "new_creators": 0,
                "run_time": "00h00min00s",
                "task_directory": str(self.task_dir),
                "created_at": datetime.now(timezone(timedelta(hours=8))).isoformat(),
            }
            # NOTE: legacy comment removed for English-only repo.
            from database.ingest_task_info import log_task_info_to_db
            try:
                log_task_info_to_db(row_data)
            except Exception as e:
                self.logger.warning(f"数据库埋点失败: {e}")


            def _normalize(value):
                if isinstance(value, list):
                    return ", ".join(str(v) for v in value)
                if isinstance(value, datetime):
                    return value.isoformat()
                return value

            with lock:
                if info_path.exists():
                    wb = openpyxl.load_workbook(info_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "tasks"
                    ws.append(headers)

                header = [cell.value for cell in ws[1]]
                missing_columns = [col for col in headers if col not in header]
                if missing_columns:
                    next_col = len(header) + 1
                    for col in missing_columns:
                        ws.cell(row=1, column=next_col, value=col)
                        next_col += 1
                    header = [cell.value for cell in ws[1]]
                if "task_id" in header:
                    id_index = header.index("task_id")
                    existing_ids = {
                        str(row[id_index])
                        for row in ws.iter_rows(min_row=2, values_only=True)
                        if row[id_index] is not None
                    }
                    if task_id in existing_ids:
                        wb.close()
                        return
                row = [_normalize(row_data.get(col)) for col in header]
                ws.append(row)
                wb.save(info_path)
                wb.close()
        except Exception as exc:  # NOTE: legacy comment removed for English-only repo.
            self.logger.warning(f"保存任务信息失败: {exc}")

    def load_existing_excel(self) -> Dict[str, List[Dict]]:
        """Legacy docstring removed for English-only repo."""
        xlsx_path = Path(f"data/creator_{self.shop_name}.xlsx")
        if not xlsx_path.exists():
            self.logger.info(f"Excel文件不存在: {xlsx_path}")
            return {}
        
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            sheet = wb.active
            
            # NOTE: legacy comment removed for English-only repo.
            headers = [cell.value for cell in sheet[1]]
            
            # NOTE: legacy comment removed for English-only repo.
            try:
                name_idx = headers.index('creator_name')
                brand_idx = headers.index('brand_name')
                connect_idx = headers.index('connect')
            except ValueError as e:
                self.logger.warning(f"Excel表头缺少必要字段: {e}")
                return {}
            
            data = {}
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[name_idx]:
                    continue
                    
                creator_name = row[name_idx]
                brand_name = row[brand_idx] if brand_idx < len(row) else ""
                connect = row[connect_idx] if connect_idx < len(row) else False
                
                # NOTE: legacy comment removed for English-only repo.
                if isinstance(connect, str):
                    connect = connect.lower() in ('true', '1', 'yes')
                
                # NOTE: legacy comment removed for English-only repo.
                if creator_name not in data:
                    data[creator_name] = []
                
                data[creator_name].append({
                    'brand_name': brand_name or "",
                    'connect': bool(connect),
                    'row_index': row_num  # NOTE: legacy comment removed for English-only repo.
                })
            
            self.logger.info(f"从Excel加载了 {len(data)} 个达人的 {sum(len(v) for v in data.values())} 条记录")
            return data
            
        except Exception as e:
            self.logger.error(f"加载Excel失败: {e}")
            return {}

    def load_blacklist(self) -> set:
        """Legacy docstring removed for English-only repo."""
        # NOTE: legacy comment removed for English-only repo.
        blacklist_path = Path(f"data/TAP_nore_FR.xlsx")
        
        # NOTE: legacy comment removed for English-only repo.
        # blacklist_path = Path(BASE_DIR) / "data" / "TAP_nore_FR.xlsx"
        
        if not blacklist_path.exists():
            self.logger.info(f"黑名单文件不存在，跳过: {blacklist_path}")
            return set()
        
        try:
            self.logger.info(f"正在加载黑名单: {blacklist_path.absolute()}")
            wb = openpyxl.load_workbook(blacklist_path, read_only=True)
            ws = wb.active
            blacklist = set()
            
            # NOTE: legacy comment removed for English-only repo.
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # NOTE: legacy comment removed for English-only repo.
                    creator_name = str(row[0]).strip()
                    if creator_name:  # NOTE: legacy comment removed for English-only repo.
                        blacklist.add(creator_name)
                        row_count += 1
                        if row_count <= 3:  # NOTE: legacy comment removed for English-only repo.
                            self.logger.info(f"黑名单样例: {creator_name}")
            
            wb.close()
            self.logger.info(f"✓ 已加载黑名单达人: {len(blacklist)} 个")
            return blacklist
        except Exception as e:
            self.logger.error(f"✗ 加载黑名单失败: {e}")
            import traceback
            self.logger.error(traceback.format_exc())  # NOTE: legacy comment removed for English-only repo.
            return set()

    def should_send_message(self, creator_name: str) -> tuple:
        """Legacy docstring removed for English-only repo."""
        records = self.existing_data.get(creator_name, [])

        if self.only_first == 2:
            if not records:
                self.logger.info(f"{creator_name}: only_first=2 但为新达人 -> 跳过")
                return (None, None)

            has_reply_history = any(r.get('reply', False) for r in records) or self._has_reply_in_history(creator_name)
            if has_reply_history:
                self.logger.info(f"{creator_name}: only_first=2 但历史已回复 -> 跳过")
                return (None, None)

            self.logger.info(f"{creator_name}: only_first=2 且历史未回复 -> later")
            return ('later', self._load_email_later_msg())
        
        # NOTE: legacy comment removed for English-only repo.
        has_any_reply = any(r.get('reply', False) for r in records)
        if has_any_reply:
            self.logger.info(f"{creator_name}: 历史记录中存在回复 -> 不发送")
            return (None, None)
        
        # NOTE: legacy comment removed for English-only repo.
        if len(records) >= 5:
            all_no_reply = all(not r.get('reply', False) for r in records)
            if all_no_reply:
                self.logger.info(f"{creator_name}: 历史记录>=5条且全部未回复 -> 不发送")
                return (None, None)
        
        # NOTE: legacy comment removed for English-only repo.
        if not records:
            self.logger.info(f"{creator_name}: 新达人 -> first")
            return ('first', self._load_email_msg())
        
        # NOTE: legacy comment removed for English-only repo.
        all_not_connected = all(not r['connect'] for r in records)
        if all_not_connected:
            self.logger.info(f"{creator_name}: 所有记录都未建联 -> first")
            return ('first', self._load_email_msg())
        
        # NOTE: legacy comment removed for English-only repo.
        has_connected = any(r['connect'] for r in records)
        
        if has_connected:
            # NOTE: legacy comment removed for English-only repo.
            if self.only_first == 1:
                self.logger.info(f"{creator_name}: 已建联且only_first=1 -> 跳过")
                return (None, None)
            
            # NOTE: legacy comment removed for English-only repo.
            all_brands = {r['brand_name'] for r in records}
            if self.brand_name not in all_brands:
                self.logger.info(f"{creator_name}: 已建联但所有品牌({all_brands})都不同于当前品牌({self.brand_name}) -> later")
                return ('later', self._load_email_later_msg())
            else:
                self.logger.info(f"{creator_name}: 已建联且当前品牌({self.brand_name})已存在 -> 跳过")
                return (None, None)
        
        return (None, None)

    def _has_reply_in_history(self, creator_name: str) -> bool:
        if creator_name in self.creator_reply_cache:
            return self.creator_reply_cache[creator_name]
        exists = False
        try:
            exists = creator_has_reply(creator_name)
        except Exception as exc:
            self.logger.debug(f"查询 MySQL 回复记录失败({creator_name}): {exc}")
        self.creator_reply_cache[creator_name] = exists
        return exists

    def _load_email_msg(self) -> str:
        """Legacy docstring removed for English-only repo."""
        try:
            path = self.task_dir / "dify_out.txt"
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            subject = (data.get("email_first", {}).get("subject") or "").strip()
            body = (data.get("email_first", {}).get("email_body") or "").strip()
            
            if subject and body:
                # NOTE: legacy comment removed for English-only repo.
                subject_clean = subject.replace('"', '')
                # NOTE: legacy comment removed for English-only repo.
                body_clean = body.replace('\r\n', '\n')
                body_clean = re.sub(r'\n{3,}', '\n\n', body_clean).strip()
                msg = f"{subject_clean}\n\n{body_clean}"
                return msg
            return subject or body
        except Exception as e:
            self.logger.warning(f"读取/组合email_first文案失败: {e}")
            return ""

    def _load_email_later_msg(self) -> str:
        """Legacy docstring removed for English-only repo."""
        try:
            path = self.task_dir / "dify_out.txt"
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            subject = (data.get("email_later", {}).get("subject") or "").strip()
            body = (data.get("email_later", {}).get("email_body") or "").strip()
            
            if subject and body:
                subject_clean = subject.replace('"', '')
                body_clean = body.replace('\r\n', '\n')
                body_clean = re.sub(r'\n{3,}', '\n\n', body_clean).strip()
                return f"{subject_clean}\n\n{body_clean}"
            return subject or body
        except Exception as e:
            self.logger.warning(f"读取email_later失败: {e}")
            return ""

    def register_cleanup(self):
        """Legacy docstring removed for English-only repo."""
        if self.cleanup_registered:
            return

        atexit.register(self.emergency_cleanup)

        # NOTE: legacy comment removed for English-only repo.
        try:
            if threading.current_thread() is threading.main_thread():
                signal.signal(signal.SIGTERM, self.signal_handler)
                signal.signal(signal.SIGINT, self.signal_handler)
            else:
                self.logger.debug("子线程环境，跳过 signal handler 注册")
        except ValueError as exc:
            self.logger.warning(f"注册 signal handler 失败: {exc}")

        self.cleanup_registered = True
        self.logger.info("已注册清理函数")

    def signal_handler(self, signum, frame):
        """Legacy docstring removed for English-only repo."""
        self.logger.info(f"收到信号 {signum}，开始清理...")
        self.emergency_cleanup()
        exit(0)

    def emergency_cleanup(self):
        """Legacy docstring removed for English-only repo."""
        if self.is_cleaned_up:
            self.logger.info("资源已经清理过，跳过重复清理")
            return
            
        self.logger.info("执行紧急资源清理...")
        try:
            # NOTE: legacy comment removed for English-only repo.
            if self.playwright_context_active:
                # NOTE: legacy comment removed for English-only repo.
                for page in list(self.active_pages):
                    try:
                        if not page.is_closed():
                            page.close()
                            self.logger.info(f"已关闭页面")
                    except Exception as page_error:
                        self.logger.warning(f"关闭页面时出错: {page_error}")
                
                # NOTE: legacy comment removed for English-only repo.
                try:
                    if self.context:
                        self.context.close()
                        self.logger.info("已关闭浏览器上下文")
                except Exception as context_error:
                    self.logger.warning(f"关闭浏览器上下文时出错: {context_error}")
                
                try:
                    if self.browser:
                        self.browser.close()
                        self.logger.info("已关闭浏览器")
                except Exception as browser_error:
                    self.logger.warning(f"关闭浏览器时出错: {browser_error}")
            else:
                self.logger.info("Playwright上下文已关闭，跳过浏览器清理")
            
            # NOTE: legacy comment removed for English-only repo.
            self.active_pages.clear()
            self.browser = None
            self.context = None
            self.main_page = None
            self.playwright_context_active = False
            self.is_cleaned_up = True
            
            self.logger.info("紧急清理完成")
            
        except Exception as e:
            self.logger.error(f"紧急清理失败: {e}")
            # NOTE: legacy comment removed for English-only repo.
            self.is_cleaned_up = True
        finally:
            self.browser_force_killed = False

    def _kill_browser_process(self, reason: str = ""):
        """Legacy docstring removed for English-only repo."""
        if self.browser_force_killed:
            return
        browser = self.browser
        if not browser:
            return
        process = getattr(browser, "process", None)
        if not process:
            return
        pid = getattr(process, "pid", None)
        try:
            self.browser_force_killed = True
            self.logger.warning(
                f"强制终止浏览器进程(pid={pid})，原因: {reason or '强制取消'}"
            )
            process.kill()
            try:
                process.wait(timeout=5)
            except Exception:
                pass
        except Exception as exc:
            self.logger.warning(f"强制终止浏览器进程失败: {exc}")

    def safe_cleanup(self):
        """Legacy docstring removed for English-only repo."""
        if self.is_cleaned_up:
            return
            
        self.logger.info("执行安全资源清理...")
        try:
            # NOTE: legacy comment removed for English-only repo.
            for page in list(self.active_pages):
                try:
                    if page and not page.is_closed():
                        page.close()
                except:
                    pass
            
            self.active_pages.clear()
            self.logger.info("安全清理完成")
            
        except Exception as e:
            self.logger.warning(f"安全清理过程出错: {e}")

    @contextmanager
    def managed_page(self, description=""):
        """Legacy docstring removed for English-only repo."""
        page = None
        try:
            self.logger.info(f"等待新页面打开: {description}")
            page = self.context.wait_for_event("page", timeout=10000)
            self.active_pages.add(page)
            self.logger.info(f"新页面已添加到管理列表")
            yield page
        except Exception as e:
            self.logger.error(f"页面管理出错: {e}")
            raise
        finally:
            if page and not page.is_closed():
                try:
                    page.close()
                    self.logger.info(f"已关闭管理的页面: {description}")
                except Exception as e:
                    self.logger.warning(f"关闭页面失败: {e}")
                finally:
                    self.active_pages.discard(page)

    def cleanup_stale_pages(self):
        """Legacy docstring removed for English-only repo."""
        stale_pages = []
        for page in list(self.active_pages):
            try:
                if page.is_closed():
                    stale_pages.append(page)
            except:
                stale_pages.append(page)
        
        for page in stale_pages:
            self.active_pages.discard(page)
        
        if stale_pages:
            self.logger.info(f"清理了 {len(stale_pages)} 个过期页面引用")

    def delay(self, seconds: float):
        """Legacy docstring removed for English-only repo."""
        if seconds <= 0:
            return
        end_time = time.monotonic() + seconds
        while True:
            remaining = end_time - time.monotonic()
            if remaining <= 0:
                break
            sleep_span = min(0.5, remaining)
            time.sleep(sleep_span)
            if self.internal_cancel_event.is_set() or (self.cancel_event and self.cancel_event.is_set()):
                self.raise_if_cancelled("休眠等待")
    
    def click_blank_area(self, page):
        """Legacy docstring removed for English-only repo."""
        try:
            page.mouse.click(500, 200)
            self.logger.info("点击了空白区域")
            self.delay(1)
        except Exception as e:
            self.logger.warning(f"点击空白区域失败: {e}")

    def safe_extract_text(self, page, selector: str) -> str:
        """Legacy docstring removed for English-only repo."""
        try:
            element = page.locator(selector)
            # NOTE: legacy comment removed for English-only repo.
            if element.count() > 0:
                return element.first.text_content().strip()
        except PlaywrightTimeoutError:
            self.logger.warning(f"Timeout while waiting for selector: {selector}")
            return ""
        except Exception as e:
            self.logger.warning(f"Error extracting text from {selector}: {e}")
            return ""

    def close_verification_popup(self, page) -> bool:
        """Legacy docstring removed for English-only repo."""
        try:
            self.logger.info("尝试关闭验证码弹窗...")
            # NOTE: legacy comment removed for English-only repo.
            page.locator('button[aria-label="Close"]').click(timeout=5000)
            self.logger.info("使用关闭按钮关闭弹窗")
            return True
        except Exception:
            # NOTE: legacy comment removed for English-only repo.
            try:
                page.keyboard.press("Escape")
                self.logger.info("使用ESC键关闭弹窗")
                self.delay(1)
                return True
            except Exception as e:
                self.logger.error(f"关闭验证码弹窗失败: {e}")
                return False

    def login(self, page) -> bool:
        """Legacy docstring removed for English-only repo."""
        max_retries = 5

        # NOTE: legacy comment removed for English-only repo.
        # NOTE: legacy comment removed for English-only repo.
        # NOTE: legacy comment removed for English-only repo.
        # NOTE: legacy comment removed for English-only repo.
        # NOTE: legacy comment removed for English-only repo.
        for i in range(max_retries):
            self.logger.info(f"开始登录流程（第 {i + 1} 次尝试）...")
            try:
                # NOTE: legacy comment removed for English-only repo.
                page.goto("https://partner-sso.tiktok.com/account/login?from=ttspc_logout&redirectURL=%2F%2Fpartner.tiktokshop.com%2Fhome&lang=en", wait_until="networkidle")

                
                self.logger.info("进入邮件登录模式...")
                email_login_btn = page.get_by_text("Log in with code").first
                email_login_btn.click()

                self.logger.info(f"输入邮箱: {self.login_email}")
                page.fill('#email input', self.login_email)
                
                # NOTE: legacy comment removed for English-only repo.
                try:
                    send_code_btn = page.locator('div[starling-key="profile_edit_userinfo_send_code"]').first
                    send_code_btn.wait_for(state="visible", timeout=5000)
                    send_code_btn.click()
                    self.logger.info("已点击 Send code 按钮")
                except Exception as e:
                    self.logger.error(f"点击 Send code 失败: {e}")

                self.logger.info("正在从Gmail获取验证码...")
                verification_code = self.gmail_verifier.get_verification_code()
                if not verification_code:
                    self.logger.error("验证码获取失败")
                    # NOTE: legacy comment removed for English-only repo.
                    continue

                self.logger.info(f"成功获取验证码: {verification_code}")
                page.fill('#emailCode_input', verification_code)

                login_btn = page.locator('button[starling-key="account_login_btn_loginform_login_text"]').first
                login_btn.click()
                if "partner.tiktokshop.com" in page.url:
                    self.logger.info("登录成功")
                    self.delay(5)  # NOTE: legacy comment removed for English-only repo.
                    # NOTE: legacy comment removed for English-only repo.
                    # screenshot_path = f"data/screenshots/login_success_{int(time.time())}.png"
                    # page.screenshot(path=screenshot_path, full_page=True)
                    # NOTE: legacy comment removed for English-only repo.
                    return True
                else:
                    self.logger.error("登录失败，未跳转到正确页面")
                    # NOTE: legacy comment removed for English-only repo.
                    time.sleep(3)
                    continue
            
            except Exception as e:
                self.logger.error(f"登录过程出错：{e}")
                # NOTE: legacy comment removed for English-only repo.
                if i < max_retries - 1:
                    self.logger.warning(f"第 {i + 1} 次尝试失败，等待 3 秒后重试...")
                    time.sleep(3)
        
        self.logger.error(f"登录失败，已达到最大重试次数 {max_retries}。程序终止。")
        return False

    def navigate_to_creator_connection(self, page) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("导航到创作者连接页面...")
        try:
            self.logger.info("等待页面加载，查找欢迎文字...")
            try:
                # NOTE: legacy comment removed for English-only repo.
                try:
                    page.wait_for_selector('text=Welcome to TikTok Shop Partner Center', timeout=20000)
                except:
                    try:
                        page.wait_for_selector('text=Account GMV trend', timeout=20000)
                    except:
                        try:
                            page.wait_for_selector('text=View your data and facilitate seller authorizations', timeout=20000)
                        except:
                            page.wait_for_selector('text=Hi', timeout=20000)
            except:
                self.logger.warning("未找到欢迎文字、GMV趋势或Hi文字，尝试重新开始登录")
                # screenshot_path = f"data/screenshots/login_success_{int(time.time())}.png"
                # page.screenshot(path=screenshot_path, full_page=True)
                # NOTE: legacy comment removed for English-only repo.
                return False

            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            self.logger.info("查找并点击创作者市场按钮...")
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            try:
                creator_button = page.locator('[data-uid="firstmenuitem:div:0912e"]')
                creator_button.wait_for(state="visible", timeout=10000)
                creator_button.click()
                self.logger.info("成功点击创作者市场按钮（通过data-uid精确匹配）")
            except:
                # NOTE: legacy comment removed for English-only repo.
                try:
                    creator_button = page.locator('.index__firstMenuItem--msEti.index__menuClosed--ROED2').first
                    creator_button.wait_for(state="visible", timeout=10000)
                    creator_button.click()
                    self.logger.info("成功点击创作者市场按钮（通过class组合）")
                except:
                    # NOTE: legacy comment removed for English-only repo.
                    try:
                        creator_button = page.locator('svg.arco-icon-creator_marketplace_unselect').locator('..').locator('..')
                        creator_button.wait_for(state="visible", timeout=10000)
                        creator_button.click()
                        self.logger.info("成功点击创作者市场按钮（通过SVG父级）")
                    except:
                        # NOTE: legacy comment removed for English-only repo.
                        try:
                            creator_button = page.locator('.index__firstMenuItem--msEti:has(svg.arco-icon-creator_marketplace_unselect)')
                            creator_button.wait_for(state="visible", timeout=10000)
                            creator_button.click()
                            self.logger.info("成功点击创作者市场按钮（通过has选择器）")
                        except Exception as e:
                            self.logger.error(f"所有点击方法都失败了: {e}")
                            raise
            
            # NOTE: legacy comment removed for English-only repo.
            self.delay(8)  # NOTE: legacy comment removed for English-only repo.

            self.logger.info("检查页面是否已成功加载...")
            # NOTE: legacy comment removed for English-only repo.
            page.wait_for_selector('text=Find creators', timeout=60000)
            
            self.logger.info("创作者页面已成功加载。")
            
            return True
        except PlaywrightTimeoutError:
            self.logger.error("在60秒内未找到 'Find creators' 文本，可能是页面加载失败或结构变化。")
            return False
        except Exception as e:
            self.logger.error(f"导航到创作者连接页面失败: {e}")
            return False
    
    @classmethod
    def load_search_strategy_from_file(cls, filepath):
        """Legacy docstring removed for English-only repo."""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                search_strategy = data.get('search_strategy', {})
                return search_strategy
        except Exception as e:
            logging.error(f"加载搜索策略文件失败: {e}")
            return None

    def parse_search_strategy(self):
        """Legacy docstring removed for English-only repo."""
        strategy = self.search_strategy
        
        # NOTE: legacy comment removed for English-only repo.
        keywords = strategy.get("search_keywords", "")
        if isinstance(keywords, str):
            self.search_keywords_raw = keywords.strip()
            self.search_keyword = (keywords.split(",")[0]).strip() if keywords.strip() else ""
        else:
            self.search_keywords_raw = ", ".join(str(k).strip() for k in (keywords or []))
            self.search_keyword = (keywords[0].strip() if keywords else "")
        
        # NOTE: legacy comment removed for English-only repo.
        raw_product_cat = strategy.get("product_category", "")
        if isinstance(raw_product_cat, str):
            product_cat_list = [cat.strip() for cat in raw_product_cat.split(",") if cat.strip()]
        elif isinstance(raw_product_cat, list):
            product_cat_list = [str(c).strip() for c in raw_product_cat if str(c).strip()]
        else:
            product_cat_list = []

        normalized_categories: List[str] = []
        normalized_category_ids: List[str] = []

        for entry in product_cat_list:
            key = entry.strip()
            if not key:
                continue

            name = None
            cat_id = None

            # NOTE: legacy comment removed for English-only repo.
            if key in CATEGORY_NAME_TO_ID:
                name = key
                cat_id = CATEGORY_NAME_TO_ID[key]
            else:
                # NOTE: legacy comment removed for English-only repo.
                lower_key = key.lower()
                if lower_key in CATEGORY_NAME_TO_ID_LOWER:
                    name = next(
                        original_name
                        for original_name, code in CATEGORY_NAME_TO_ID.items()
                        if original_name.lower() == lower_key
                    )
                    cat_id = CATEGORY_NAME_TO_ID[name]
                elif key in CATEGORY_ID_TO_NAME:
                    cat_id = key
                    name = CATEGORY_ID_TO_NAME[key]
                else:
                    self.logger.warning(f"未知的产品类别：{key}")
                    continue

            normalized_categories.append(name)
            normalized_category_ids.append(cat_id)

        self.product_categories = normalized_categories
        self.product_category_ids = normalized_category_ids
        if normalized_categories:
            # NOTE: legacy comment removed for English-only repo.
            self.search_strategy["product_category"] = normalized_categories
            self.search_strategy["product_category_ids"] = normalized_category_ids
        else:
            self.product_categories = []
            self.product_category_ids = []
        
        # NOTE: legacy comment removed for English-only repo.
        age_range = strategy.get("fans_age_range", [])
        if isinstance(age_range, str):
            if age_range.strip() == "":
                self.age_ranges = []
            else:
                self.age_ranges = [
                    age.strip().replace("-", " - ")
                    for age in age_range.split(",")
                    if age.strip()  # NOTE: legacy comment removed for English-only repo.
                ]
                self.age_ranges = [
                    age.replace("55 +", "55+").replace("55 - ", "55+")
                    for age in self.age_ranges
                ]
        elif isinstance(age_range, list):
            self.age_ranges = [str(a).strip().replace("-", " - ") for a in age_range if str(a).strip()]
        else:
            self.age_ranges = []
        
        # NOTE: legacy comment removed for English-only repo.
        gender_info = strategy.get("fans_gender", "")
        if isinstance(gender_info, str):
            s = gender_info.strip()
            if not s:
                self.gender, self.gender_percentage = "", None
            else:
                m = re.search(r'(\d+)%', s)
                pct = int(m.group(1)) if m else 50
                if "female" in s.lower():
                    self.gender, self.gender_percentage = "Female", pct
                elif "male" in s.lower():
                    self.gender, self.gender_percentage = "Male", pct
                else:
                    self.gender, self.gender_percentage = "", None
        else:
            self.gender, self.gender_percentage = "", None
        
        # NOTE: legacy comment removed for English-only repo.
        self.min_fans = int(strategy.get("min_fans", 5000))
        
        # NOTE: legacy comment removed for English-only repo.
        content = strategy.get("content_type", [])
        if isinstance(content, str):
            self.content_types = [content.strip()] if content.strip() else []
        elif isinstance(content, list):
            self.content_types = [str(c).strip() for c in content if str(c).strip()]
        else:
            self.content_types = []
        
        # NOTE: legacy comment removed for English-only repo.
        def _parse_gmv_value(value, default=None):
            if value in (None, "", []):
                return default
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                cleaned = (
                    value.replace("$", "")
                    .replace("€", "")
                    .replace("£", "")
                    .replace(",", "")
                    .replace(" ", "")
                )
                if "." in cleaned:
                    parts = cleaned.split(".")
                    if len(parts) > 1 and all(part.isdigit() for part in parts) and all(len(part) == 3 for part in parts[1:]):
                        cleaned = "".join(parts)
                if cleaned.upper().endswith("K"):
                    try:
                        return float(cleaned[:-1]) * 1000
                    except ValueError:
                        pass
                try:
                    return float(cleaned)
                except ValueError:
                    match = re.search(r"(\d+)", cleaned)
                    if match:
                        return float(match.group(1))
            return default

        gmv_code_definitions = [
            {"code": "0-100", "lower": 0, "upper": 100, "label_fr": "0 €-100 €", "label_default": "$0-$100"},
            {"code": "100-1k", "lower": 100, "upper": 1000, "label_fr": "100 €-1.000 €", "label_default": "$100-$1K"},
            {"code": "1k-10k", "lower": 1000, "upper": 10000, "label_fr": "1.000 €-10K €", "label_default": "$1K-$10K"},
            {"code": "10k+", "lower": 10000, "upper": None, "label_fr": "10K €+", "label_default": "$10K+"},
        ]

        def _label_for_code(code: str) -> Optional[str]:
            for entry in gmv_code_definitions:
                if entry["code"] == code:
                    return entry["label_fr"] if self.region == "FR" else entry["label_default"]
            return None

        def _normalize_gmv_code(code) -> Optional[str]:
            if code in (None, "", []):
                return None
            s = str(code).strip().lower()
            s = (
                s.replace("$", "")
                .replace("€", "")
                .replace("usd", "")
                .replace("eur", "")
                .replace("–", "-")
                .replace("—", "-")
                .replace("_", "-")
                .replace("to", "-")
                .replace(" ", "")
            )
            s = s.replace(",", "").replace(".", "")
            if s.endswith("plus"):
                s = s[:-4] + "+"
            if s in {"0-100", "0-100+", "0-100k"}:
                return "0-100"
            if s in {"100-1k", "100-1000", "100-1k+", "100-1000+"}:
                return "100-1k"
            if s in {"1k-10k", "1000-10000", "1000-10k", "1k-10000"}:
                return "1k-10k"
            if s in {"10k+", "10000+", "10k", "10000"}:
                return "10k+"
            return None

        gmv_codes_input = strategy.get("gmv", [])
        if isinstance(gmv_codes_input, str):
            gmv_codes_input = [gmv_codes_input]

        normalized_codes: List[str] = []
        if isinstance(gmv_codes_input, (list, tuple)):
            for code in gmv_codes_input:
                norm = _normalize_gmv_code(code)
                if norm and norm not in normalized_codes:
                    normalized_codes.append(norm)

        if normalized_codes:
            self.gmv_ranges = [
                label for code in normalized_codes if (label := _label_for_code(code))
            ]
            self.search_strategy["gmv"] = normalized_codes
        else:
            # NOTE: legacy comment removed for English-only repo.
            min_gmv = _parse_gmv_value(strategy.get("min_GMV"), default=0)
            max_gmv = _parse_gmv_value(strategy.get("max_GMV"), default=None)

            if max_gmv is not None and max_gmv < min_gmv:
                max_gmv = min_gmv

            epsilon = 1e-6

            def _locate_index(value):
                if value is None:
                    return 0
                for idx, entry in enumerate(gmv_code_definitions):
                    lower = entry["lower"]
                    upper = entry["upper"]
                    if value + epsilon < lower:
                        continue
                    if upper is None:
                        return idx
                    if value < upper - epsilon:
                        return idx
                return len(gmv_code_definitions) - 1

            start_idx = _locate_index(min_gmv)
            end_idx = len(gmv_code_definitions) - 1 if max_gmv is None else _locate_index(max_gmv)
            if end_idx < start_idx:
                end_idx = start_idx

            selected_entries = gmv_code_definitions[start_idx:end_idx + 1]
            self.gmv_ranges = [
                entry["label_fr"] if self.region == "FR" else entry["label_default"]
                for entry in selected_entries
            ]
            normalized_codes = [entry["code"] for entry in selected_entries]
            self.search_strategy["gmv"] = normalized_codes

        self.gmv_ranges = self.gmv_ranges or []

        # NOTE: legacy comment removed for English-only repo.
        sales_code_definitions = [
            {"code": "0-10", "lower": 0, "upper": 10, "label": "0-10"},
            {"code": "10-100", "lower": 10, "upper": 100, "label": "10-100"},
            {"code": "100-1k", "lower": 100, "upper": 1000, "label": "100-1K"},
            {"code": "1k+", "lower": 1000, "upper": None, "label": "1K+"},
        ]
        label_map = {entry["code"]: entry["label"] for entry in sales_code_definitions}

        def _normalize_sales_code(code) -> Optional[str]:
            if code in (None, "", []):
                return None
            s = str(code).strip().lower()
            s = (
                s.replace(" ", "")
                .replace("–", "-")
                .replace("—", "-")
                .replace("to", "-")
                .replace("_", "-")
                .replace("plus", "+")
            )
            s = s.replace(",", "")
            if s in {"0-10", "0_10"}:
                return "0-10"
            if s in {"10-100", "10_100"}:
                return "10-100"
            if s in {"100-1k", "100-1000", "100-1k+", "100_1k"}:
                return "100-1k"
            if s in {"1k+", "1000+", "1k", "1000"}:
                return "1k+"
            return None

        def _extract_numeric_threshold(value) -> Optional[int]:
            if isinstance(value, (int, float)):
                return int(value)
            if isinstance(value, str):
                match = re.search(r"(\d+)", value)
                if match:
                    return int(match.group(1))
            return None

        def _codes_from_threshold(threshold: int) -> List[str]:
            if threshold <= 0:
                return [entry["code"] for entry in sales_code_definitions]
            if threshold <= 10:
                return [entry["code"] for entry in sales_code_definitions[1:]]
            if threshold <= 100:
                return [entry["code"] for entry in sales_code_definitions[2:]]
            if threshold <= 1000:
                return [sales_code_definitions[3]["code"]]
            return []

        def _normalize_sales_input(raw) -> List[str]:
            if raw in (None, "", []):
                return []
            if isinstance(raw, (list, tuple)):
                items = list(raw)
            else:
                items = [raw]

            codes: List[str] = []
            thresholds: List[int] = []

            for item in items:
                if item in (None, "", []):
                    continue
                normalized_code = _normalize_sales_code(item)
                if normalized_code:
                    if normalized_code not in codes:
                        codes.append(normalized_code)
                    continue
                threshold = _extract_numeric_threshold(item)
                if threshold is not None:
                    thresholds.append(threshold)

            if not codes and thresholds:
                threshold = min(thresholds)
                for code in _codes_from_threshold(threshold):
                    if code not in codes:
                        codes.append(code)
            return codes

        sales_field_specified = "sales" in strategy
        legacy_sales_specified = "min_sales" in strategy

        normalized_sales_codes = _normalize_sales_input(strategy.get("sales") if sales_field_specified else None)

        if not normalized_sales_codes and not sales_field_specified:
            normalized_sales_codes = _normalize_sales_input(
                strategy.get("min_sales") if legacy_sales_specified else None
            )
            if (
                not normalized_sales_codes
                and not legacy_sales_specified
            ):
                normalized_sales_codes = _codes_from_threshold(10)

        normalized_sales_codes = [
            code for code in normalized_sales_codes if code in label_map
        ]
        self.sales_ranges = [label_map[code] for code in normalized_sales_codes]
        self.search_strategy["sales"] = normalized_sales_codes
        self.sales_ranges = self.sales_ranges or []
        
        # NOTE: legacy comment removed for English-only repo.
        self.avg_views = int(strategy.get("avg_views", 5000))
        
        # NOTE: legacy comment removed for English-only repo.
        engagement = strategy.get("min_engagement_rate")
        if engagement in (None, "", []):
            self.min_engagement_rate = 0.0
        else:
            try:
                self.min_engagement_rate = float(engagement)
            except (TypeError, ValueError):
                try:
                    self.min_engagement_rate = float(str(engagement).replace("%", ""))
                except (TypeError, ValueError):
                    self.min_engagement_rate = 0.0
        
        # NOTE: legacy comment removed for English-only repo.
        self.logger.info(f"解析后的搜索参数:")
        self.logger.info(f"  关键词: {self.search_keyword}")
        self.logger.info(f"  产品类别: {self.product_categories}")
        self.logger.info(f"  年龄范围: {self.age_ranges}")
        self.logger.info(f"  性别: {self.gender} ({self.gender_percentage}%)")
        self.logger.info(f"  最小粉丝数: {self.min_fans}")
        self.logger.info(f"  内容类型: {self.content_types}")
        self.logger.info(f"  GMV范围: {self.gmv_ranges}")
        self.logger.info(f"  销售数量范围: {self.sales_ranges}")
        self.logger.info(f"  平均观看数: {self.avg_views}")
        self.logger.info(f"  最小参与率: {self.min_engagement_rate}%")
    
    def search_and_filter(self, page, search_keyword: str = None) -> bool:
        """Legacy docstring removed for English-only repo."""
        # NOTE: legacy comment removed for English-only repo.
        if search_keyword is None:
            search_keyword = self.search_keyword

        self.logger.info(f"开始搜索和筛选流程（关键词：{search_keyword}）")
        self.logger.info(f"使用的搜索策略：{self.search_strategy}")

        # NOTE: legacy comment removed for English-only repo.
        applied = {
            "keyword": None,
            "creators": {
                "product_category": [],
                "content_type": []
            },
            "followers": {
                "age": [],
                "follower_min": None,
                "gender": None,
                "gender_percentage": None
            },
            "performance": {
                "gmv": [],
                "items_sold": [],
                "avg_views_min": None,
                "engagement_min": None
            }
        }

        # NOTE: legacy comment removed for English-only repo.
        def click_first(locators: list, desc: str, timeout=15000) -> bool:
            """Legacy docstring removed for English-only repo."""
            for sel in locators:
                try:
                    el = page.locator(sel).first
                    if el.count() > 0 and el.is_enabled():
                        el.click(timeout=timeout, force=True)
                        self.logger.info(f"点击成功：{desc}（{sel}）")
                        return True
                except Exception as e:
                    self.logger.debug(f"点击失败尝试：{desc}（{sel}）：{e}")
            self.logger.warning(f"未找到/未点中：{desc}")
            return False

        def fill_first(locators: list, value: str, desc: str, timeout=8000) -> bool:
            """Legacy docstring removed for English-only repo."""
            for sel in locators:
                try:
                    ipt = page.locator(sel).first
                    if ipt.count() > 0 and ipt.is_enabled():
                        ipt.click(timeout=timeout)
                        ipt.fill(str(value))
                        self.logger.info(f"输入成功：{desc} = {value}（{sel}）")
                        return True
                except Exception as e:
                    self.logger.debug(f"输入失败尝试：{desc}（{sel}）：{e}")
            self.logger.warning(f"未找到/未填入：{desc}")
            return False

        def open_dropdown(locators: list, desc: str) -> bool:
            """Legacy docstring removed for English-only repo."""
            for sel in locators:
                try:
                    el = page.locator(sel).first
                    # NOTE: legacy comment removed for English-only repo.
                    el.wait_for(state="visible", timeout=5000)
                    if el.count() > 0 and el.is_enabled():
                        el.click(timeout=5000, force=True)
                        self.logger.info(f"点击成功：{desc}（{sel}）")
                        # NOTE: legacy comment removed for English-only repo.
                        page.locator(
                            'xpath=//div[starts-with(@id,"arco-select-popup") and not(contains(@style,"display: none"))]'
                        ).first.wait_for(state="visible", timeout=5000)
                        return True
                except Exception as e:
                    self.logger.debug(f"点击失败尝试：{desc}（{sel}）：{e}")
            self.logger.warning(f"未找到/未点中：{desc}")
            return False

        def select_options_in_visible_popup(option_texts: list) -> list:
            chosen = []
            for txt in option_texts:
                # NOTE: legacy comment removed for English-only repo.
                if txt is None or str(txt).strip() == "":
                    continue
                try:
                    opt = page.locator(
                        f'xpath=//li[@role="option" and contains(@class, "arco-select-option")]'
                        f'[.//text()[contains(normalize-space(.), "{str(txt).strip()}")]]'
                    ).first
                    if opt.count() > 0:
                        try:
                            opt.scroll_into_view_if_needed(timeout=800)
                        except Exception:
                            pass
                        opt.click(force=True, timeout=3000)
                        self.logger.info(f"选择选项：{txt}")
                        chosen.append(str(txt).strip())
                        self.delay(0.2)
                    else:
                        self.logger.warning(f"未在可见面板找到选项：{txt}")
                except Exception as e:
                    self.logger.warning(f"选择选项失败：{txt}，原因：{e}")
            return chosen


        def wait_spinner_done(max_ms=15000):
            try:
                page.locator('[class*="arco-spin"], [class*="arco-spin-loading"]').wait_for(
                    state="hidden", timeout=max_ms
                )
            except Exception:
                pass

        def ensure_panel_active(panel_name: str, panel_selectors: list) -> bool:
            """Legacy docstring removed for English-only repo."""
            self.logger.info(f"确保 {panel_name} 面板激活")
            return click_first(panel_selectors, f"{panel_name} 面板")

        # NOTE: legacy comment removed for English-only repo.
        try:
            self.click_blank_area(page)
            
            # NOTE: legacy comment removed for English-only repo.
            self.logger.info("步骤1：定位并输入关键词")
            search_inputs = [
                'input[data-tid="m4b_input_search"]',
                'input[placeholder*="Search"]',
                'input[placeholder*="names"]',
                'xpath=//*[@id="content-container"]//span/span/input'
            ]
            if not fill_first(search_inputs, search_keyword, "搜索关键词"):
                return False
            applied["keyword"] = search_keyword


            # NOTE: legacy comment removed for English-only repo.
            creators_panel_selectors = [
                'label:has(input[value="creator"]) button',
                'input[value="creator"] ~ button',
            ]
            self.click_blank_area(page)
            
            # NOTE: legacy comment removed for English-only repo.
            if self.product_categories:
                self.logger.info(f"步骤2.1：Product category 多选：{self.product_categories}")
                
                for idx, (category_name, category_id) in enumerate(
                    zip(self.product_categories, self.product_category_ids), start=1
                ):
                    category_clean = category_name.strip()
                    value = category_id or CATEGORY_NAME_TO_ID.get(category_clean)

                    if not value:
                        self.logger.warning(f"未知的产品类别：{category_clean}")
                        continue

                    self.logger.info(f"Product category 第 {idx} 项：{category_clean}")
                    success = False
                    
                    for attempt in range(3):  # NOTE: legacy comment removed for English-only repo.
                        ensure_panel_active("Creators", creators_panel_selectors)
                        self.delay(1)
                        
                        # NOTE: legacy comment removed for English-only repo.
                        try:
                            product_cat_btn = page.locator('button:has(div:has-text("Product category"))').first
                            if product_cat_btn.count() > 0:
                                product_cat_btn.wait_for(state="visible", timeout=5000)
                                product_cat_btn.click(timeout=5000, force=True)
                                self.logger.info(f"✓ 点击 Product category 按钮成功")
                                self.delay(1)
                                
                                # NOTE: legacy comment removed for English-only repo.
                                try:
                                    page.locator('ul.arco-cascader-list').first.wait_for(state="visible", timeout=5000)
                                    self.logger.info("✓ Cascader 下拉列表已出现")
                                except:
                                    self.logger.warning("Cascader 下拉列表未出现，重试...")
                                    self.click_blank_area(page)
                                    self.delay(1)
                                    continue
                                
                                # NOTE: legacy comment removed for English-only repo.
                                result = page.evaluate(f"""Legacy docstring removed for English-only repo.""")
                                
                                if result == 'clicked':
                                    self.logger.info(f"✓ 选择产品类别：{category_clean}")
                                    applied["creators"]["product_category"].append(category_clean)
                                    success = True
                                    self.delay(1)
                                    break
                                elif result == 'already_checked':
                                    self.logger.info(f"类别 {category_clean} 已选中，跳过")
                                    applied["creators"]["product_category"].append(category_clean)
                                    success = True
                                    break
                                else:
                                    self.logger.warning(f"未找到 checkbox：{category_clean}")
                                    
                        except Exception as e:
                            self.logger.warning(f"第 {attempt+1} 次失败：{e}")
                        
                        if not success:
                            self.click_blank_area(page)
                            self.delay(1)
                    
                    if not success:
                        self.logger.error(f"✗ 最终未能选择：{category_clean}")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    self.click_blank_area(page)
            else:
                self.logger.info("跳过 Product category 筛选（为空）")
            
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            if self.content_types:
                self.logger.info(f"步骤2.3：Content type 多选：{self.content_types}")
                for content_type in self.content_types:
                    self.logger.info(f"步骤2.3：Content type 选择 {content_type}")
                    ensure_panel_active("Creators", creators_panel_selectors)
                    
                    if click_first([
                        'button:has(div:has-text("Content"))',
                        'button:has-text("Content type")',
                        'xpath=//button[.//div[contains(@class,"arco-typography")][contains(text(),"Content")]]'
                    ], "Content type"):
                        self.delay(1)
                        
                        # NOTE: legacy comment removed for English-only repo.
                        try:
                            video_selected = False
                            video_selectors = [
                                f'li[role="option"]:has-text("{content_type}")',
                                f'xpath=//li[@role="option"][contains(., "{content_type}")]',
                                f'.arco-select-option:has-text("{content_type}")'
                            ]
                            for selector in video_selectors:
                                try:
                                    video_option = page.locator(selector).first
                                    if video_option.count() > 0 and video_option.is_visible():
                                        video_option.click(force=True, timeout=3000)
                                        self.logger.info(f"选择选项：{content_type}")
                                        applied["creators"]["content_type"].append(content_type)
                                        video_selected = True
                                        break
                                except:
                                    continue
                            if not video_selected:
                                self.logger.warning(f"未找到 {content_type} 选项")        
                        except Exception as e:
                            self.logger.warning(f"选择 {content_type} 失败：{e}")
                    self.click_blank_area(page) 
            else:
                self.logger.info("跳过 Content type 筛选（为空）")   
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            followers_panel_selectors = [
                'label:has(input[value="follower"]) button',  # NOTE: legacy comment removed for English-only repo.
                'input[value="follower"] ~ button',  # NOTE: legacy comment removed for English-only repo.
            ]
            self.click_blank_area(page)

            # 3.1) Follower count >= min_fans
            if self.min_fans and self.min_fans > 0:
                self.logger.info(f"步骤3.1：Minimum followers 设置为 {self.min_fans}")
                ensure_panel_active("Followers", followers_panel_selectors)
                click_first([
                    'button:has(div:has-text("Follower count"))',
                    'button:has(div:has-text("Follower size"))',
                    'button[data-e2e="9ed553d9-3ba8-d083"]:has(div:has-text("Follower count"))',
                    'button[data-e2e="9ed553d9-3ba8-d083"]:has(div:has-text("Follower size"))',
                    'xpath=//button[.//div[contains(text(),"Follower count")]]'
                    'xpath=//button[.//div[contains(text(),"Follower size")]]'
                ], "Follower count")

                if fill_first([
                    'xpath=//*[@id="followerSize"]//input[1]',
                    'xpath=//*[@id="followerSize"]//span/div//div[1]/input[1]',
                    'xpath=//*[@id="followerSize"]//input[position()=1]',
                    'xpath=//div[@id="followerSize"]//input[@type="text"][1]',
                ], str(self.min_fans), "Follower 下限"):
                    applied["followers"]["follower_min"] = self.min_fans
                    self.logger.info(f"已填写 Follower 下限: {self.min_fans}")
                else:
                    self.logger.warning("未能正确填写 Follower count 下限输入框")
                self.click_blank_area(page)
            else:
                self.logger.info("跳过 Follower count 筛选（min_fans <= 0）")

            # NOTE: legacy comment removed for English-only repo.
            if self.age_ranges:
                self.logger.info(f"步骤3.1：Follower age 多选：{self.age_ranges}")
                for idx, age_range in enumerate(self.age_ranges, start=1):
                    self.logger.info(f"Follower age 第 {idx} 项：{age_range}")
                    ensure_panel_active("Followers", followers_panel_selectors)
                    if open_dropdown([
                        'button:has(div:has-text("Follower age"))',
                        'button[data-e2e="9ed553d9-3ba8-d083"]:has(div:has-text("Follower age"))',
                        'xpath=//button[.//div[contains(text(),"Follower age")]]'
                    ], "Follower age"):
                        chosen = select_options_in_visible_popup([age_range])
                        applied["followers"]["age"].extend(chosen)
                        self.click_blank_area(page)
            else:
                self.logger.info("跳过 Follower age 筛选（为空）")
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            if self.gender:   
                self.logger.info(f"步骤3.2：Follower gender 选择 {self.gender} ({self.gender_percentage}%)")
                ensure_panel_active("Followers", followers_panel_selectors)
                if open_dropdown([
                    'button:has(div:has-text("Follower gend"))',
                    'button[data-e2e="9ed553d9-3ba8-d083"]:has(div:has-text("Follower gend"))',
                    'xpath=//button[.//div[contains(text(),"Follower gend")]]'
                ], "Follower gender"):
                    self.delay(1)
                    chosen = select_options_in_visible_popup([self.gender])
                    
                    if chosen:
                        applied["followers"]["gender"] = self.gender
                        
                        # NOTE: legacy comment removed for English-only repo.
                        try:
                            slider_button = page.locator(
                                '//div[@id="followerGender"]//div[@role="slider"]'
                            ).first
                            slider_track = page.locator(
                                '//div[@id="followerGender"]//div[contains(@class,"arco-slider-road")]'
                            ).first

                            button_box = slider_button.bounding_box()
                            track_box = slider_track.bounding_box()

                            if button_box and track_box:
                                start_x = track_box['x']
                                center_y = track_box['y'] + (track_box['height'] / 2)
                                width = track_box['width']
                                
                                # NOTE: legacy comment removed for English-only repo.
                                target_percentage = self.gender_percentage / 100.0
                                target_x = start_x + width * target_percentage

                                page.mouse.move(button_box['x'] + button_box['width']/2, center_y)
                                page.mouse.down()
                                page.mouse.move(target_x, center_y, steps=10)
                                page.mouse.up()

                                self.logger.info(f"滑块已调整到{self.gender_percentage}%")
                                applied["followers"]["gender_percentage"] = self.gender_percentage
                            else:
                                self.logger.warning("未能获取 slider 的 bounding_box")
                        except Exception as e:
                            self.logger.warning(f"调整滑块失败：{e}")
            else:
                self.logger.info("跳过 Follower gender 筛选（为空）")  
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            performance_panel_selectors = [
                'button:has-text("Performance")'
            ]

            # NOTE: legacy comment removed for English-only repo.
            if self.gmv_ranges:
                self.logger.info(f"步骤4.1：GMV 多选：{self.gmv_ranges}")
                for idx, opt in enumerate(self.gmv_ranges, start=1):
                    self.logger.info(f"GMV 第 {idx} 项：{opt}")
                    success = False
                    for attempt in range(3):  # NOTE: legacy comment removed for English-only repo.
                        ensure_panel_active("Performance", performance_panel_selectors)
                        self.delay(1)
                        if open_dropdown([
                            '#gmv button',
                            'button:has(div:has-text("GMV"))',
                            'button:has(div.arco-typography:has-text("G"))',
                            'xpath=//button[.//div[contains(text(),"GMV")]]'
                        ], "GMV"):
                            self.delay(1)
                            chosen = select_options_in_visible_popup([opt])
                            if chosen:
                                applied["performance"]["gmv"].extend(chosen)
                                success = True
                                break
                        else:
                            self.logger.warning(f"第 {attempt+1} 次未找到 GMV，重置后重试...")
                            self.click_blank_area(page)
                            self.delay(1)
                    if not success:
                        self.logger.error(f"最终未能选择 GMV 选项：{opt}")
                    self.click_blank_area(page)
            else:
                self.logger.info("跳过 GMV 筛选（无设置）")

            # NOTE: legacy comment removed for English-only repo.
            if self.sales_ranges:
                self.logger.info(f"步骤4.2：Items sold 多选：{self.sales_ranges}")
                for idx, opt in enumerate(self.sales_ranges, start=1):
                    self.logger.info(f"Items sold 第 {idx} 项：{opt}")
                    success = False
                    for attempt in range(3):  # NOTE: legacy comment removed for English-only repo.

                        ensure_panel_active("Performance", performance_panel_selectors)

                        self.delay(1)
                        if open_dropdown([
                            '#unitsSold button',
                            'button:has(div:has-text("Items so"))',
                            'button:has(div:has-text("Items sold"))',
                            'button:has(div.arco-typography:has-text("Items so"))',
                            'xpath=//button[.//div[contains(text(),"Items so")]]',
                            # NOTE: legacy comment removed for English-only repo.
                            '#unitsSold button',
                            'button[data-e2e="9ed553d9-3ba8-d083"]'
                        ], "Items sold"):
                            self.delay(1)
                            chosen = select_options_in_visible_popup([opt])
 
                            if chosen:
                                applied["performance"]["items_sold"].extend(chosen)
                                success = True
                                break
                        else:
                            self.logger.warning(f"第 {attempt+1} 次未找到 Items sold，重置后重试...")
                            self.click_blank_area(page)
                            self.delay(1)
                    if not success:
                        self.logger.error(f"最终未能选择 Items sold 选项：{opt}")
                    self.click_blank_area(page)
            else:
                self.logger.info("跳过 Items sold 筛选（无设置）")

            # NOTE: legacy comment removed for English-only repo.
            if self.avg_views and self.avg_views > 0:
                self.logger.info(f"步骤4.3：Average views per video 设置下限 {self.avg_views}")
                ensure_panel_active("Performance", performance_panel_selectors)
                click_first([
                    'button:has(div:has-text("Average views per video"))',
                    'xpath=//button[.//div[contains(@class,"arco-typography")][normalize-space(text())="Average views per video"]]',
                    'xpath=(//*[@id="content-container"]//div[contains(@class,"index-module__button")]/button)[3]'
                ], "Average views per video")
                if fill_first([
                    'xpath=//*[@id="filter-container"]/div[2]/span/div/div[1]/input',
                    'input[data-tid="m4b_input"]:visible',
                    'input[type="text"]:visible',
                    'xpath=//div[@role="dialog"]//input[1]'
                ], str(self.avg_views), "Average views 下限"):
                    applied["performance"]["avg_views_min"] = self.avg_views
            else:
                self.logger.info("跳过 Average views per video 筛选（min_engagement_rate <= 0）")
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            if self.min_engagement_rate and self.min_engagement_rate > 0:
                self.logger.info(f"步骤4.4：Engagement rate 设置下限 {self.min_engagement_rate}")
                ensure_panel_active("Performance", performance_panel_selectors)
                click_first([
                    'button[data-e2e="9ed553d9-3ba8-d083"]',
                    'button:has(div:has-text("Engagement rate"))',
                    'xpath=//button[.//div[contains(@class,"arco-typography")][normalize-space(text())="Engagement rate"]]',
                    'xpath=(//*[@id="content-container"]//div[contains(@class,"index-module__button")]/button)[4]'
                ], "Engagement rate")
                if fill_first([
                    'xpath=//*[@id="filter-container"]/div[2]/span/div/div[1]/div/span/span/input',
                    'input[data-tid="m4b_input"]:visible',
                    'input[data-e2e="7f6a7b3f-260b-00c0"]',
                    'xpath=//div[@role="dialog"]//input[1]',
                    'input[type="text"]:visible'
                ], str(self.min_engagement_rate/10), "Engagement rate 下限"):
                    # =================BUG===================
                    # NOTE: legacy comment removed for English-only repo.
                    applied["performance"]["engagement_min"] = self.min_engagement_rate
            else:
                self.logger.info("跳过 Engagement rate 筛选（min_engagement_rate <= 0）")
            self.click_blank_area(page)

            # NOTE: legacy comment removed for English-only repo.
            self.logger.info("——— 本次筛选条件汇总（点击搜索前） ———")
            self.logger.info(f"关键词：{applied['keyword']}")
            self.logger.info(f"Creators：Product category = {', '.join(applied['creators']['product_category']) or '无'}; "
                        f"Follower >= {applied['followers']['follower_min']}; "
                        f"Content type = {', '.join(applied['creators']['content_type']) or '无'}")
            self.logger.info(f"Followers：Age = {', '.join(applied['followers']['age']) or '无'}; "
                        f"Gender = {applied['followers']['gender']} "
                        f"({'at ' + str(applied['followers']['gender_percentage']) + '%' if applied['followers']['gender_percentage'] else ''})")
            self.logger.info(f"Performance：GMV = {', '.join(applied['performance']['gmv']) or '无'}; "
                        f"Items sold = {', '.join(applied['performance']['items_sold']) or '无'}; "
                        f"Avg views >= {applied['performance']['avg_views_min']}; "
                        f"Engagement rate >= {applied['performance']['engagement_min']}")
            
            # NOTE: legacy comment removed for English-only repo.
            self.logger.info("点击搜索按钮")
            if not click_first([
                'button:has(svg.arco-icon-search)',
                'svg.arco-icon-search',
                'button:has([class*="search"])',
                'xpath=//button[.//svg[contains(@class,"search")]]'
            ], "搜索按钮"):
                self.logger.error("未找到搜索按钮")
                return False

            wait_spinner_done()
            self.logger.info("搜索和筛选完成")
            return True

        except Exception as e:
            self.logger.error(f"搜索和筛选过程出错：{e}")
            return False

    def get_creators_in_current_page(self, page) -> List[str]:
        """Legacy docstring removed for English-only repo."""
        
        creators = page.evaluate("""Legacy docstring removed for English-only repo.""")
        
        self.logger.info(f"当前DOM中有 {len(creators)} 个达人元素")  
        return creators

    def click_creator(self, page, creator_name: str):
        # NOTE: legacy comment removed for English-only repo.
        row = page.locator(
            f'tr:has(span[data-e2e="fbc99397-6043-1b37"]:has-text("{creator_name}"))'
        ).first
        if row.count() > 0:
            row.scroll_into_view_if_needed()
            row.wait_for(state="visible", timeout=5000)
            row.click()  # NOTE: legacy comment removed for English-only repo.
            self.logger.info(f"点击了达人(表格行) {creator_name}")
            return True

        # NOTE: legacy comment removed for English-only repo.
        for sel in [
            f'div[class*="creator-card"]:has(span:has-text("{creator_name}"))',
            f'div[class*="item"]:has(span:has-text("{creator_name}"))',
            f'a:has(span:has-text("{creator_name}"))',
            f'span:has-text("{creator_name}")'
        ]:
            el = page.locator(sel).first
            if el.count() > 0:
                el.scroll_into_view_if_needed()
                el.click()
                self.logger.info(f"点击了达人(卡片) {creator_name}")
                return True

        self.logger.error(f"未找到达人 {creator_name} 的可点击行/卡片")
        return False

    def _wait_for_new_page(self, context, base_pages, timeout_ms=10000):
        """Legacy docstring removed for English-only repo."""
        start = time.time()
        interval = 0.1
        while (time.time() - start) * 1000 < timeout_ms:
            self.raise_if_cancelled("等待新页面打开")
            for ctx_page in context.pages:
                if ctx_page not in base_pages:
                    return ctx_page
            time.sleep(interval)
        return None

    def _open_detail_and_get_page(self, page, creator_name: str):
        base_pages = list(page.context.pages)
        clicked = self.click_creator(page, creator_name)

        if not clicked:
            raise Exception(f"点击达人 {creator_name} 失败")

        new_page = self._wait_for_new_page(page.context, base_pages, timeout_ms=10000)

        if new_page:
            new_page.wait_for_load_state("domcontentloaded", timeout=15000)
            return new_page, True  # NOTE: legacy comment removed for English-only repo.

        # NOTE: legacy comment removed for English-only repo.
        try:
            page.wait_for_url(re.compile(r'/creator/detail'), timeout=15000)
        except PlaywrightTimeoutError:
            # NOTE: legacy comment removed for English-only repo.
            page.wait_for_selector('text=Partnered brands', timeout=8000)
        return page, False

    def extract_creator_details(self, page, creator_name: str) -> Dict[str, Any]:
        """Legacy docstring removed for English-only repo."""
        self.logger.info(f"提取达人 {creator_name} 的详情信息...")
        
        # NOTE: legacy comment removed for English-only repo.
        try:
            is_detail_page = page.evaluate("""
                () => {
                    const url = window.location.href;
                    return url.includes('/creator/detail') || 
                        document.querySelector('div.text-head-l') !== null ||
                        document.querySelector('text=Partnered brands') !== null;
                }
            """)
            if not is_detail_page:
                self.logger.error(f"未成功进入 {creator_name} 的详情页，返回None触发重试")
                return None  # NOTE: legacy comment removed for English-only repo.
        except:
            self.logger.error(f"无法验证是否在详情页")
            return None

        # NOTE: legacy comment removed for English-only repo.
        max_wait = 30
        wait_count = 0
        
        # NOTE: legacy comment removed for English-only repo.
        key_elements = [
            ('标题', 'div.text-head-l:has-text("Creator details"), div.text-head-l:has-text("达人详情")'),
            ('分类', '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[1]/span[1]/span[2]/span/span'),
            ('粉丝数', '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[1]/span[2]/span[2]/span/span'),
            ('简介', '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[3]/div/span'),
            ('销售数据', '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[1]'),
        ]
        
        loaded_elements = set()
        
        while wait_count < max_wait:
            try:
                # NOTE: legacy comment removed for English-only repo.
                for elem_name, selector in key_elements:
                    if elem_name not in loaded_elements:
                        try:
                            if selector.startswith('//') or selector.startswith('(//'):
                                # NOTE: legacy comment removed for English-only repo.
                                element = page.locator(f'xpath={selector}')
                            else:
                                # NOTE: legacy comment removed for English-only repo.
                                element = page.locator(selector)
                            
                            if element.count() > 0:
                                loaded_elements.add(elem_name)
                                self.logger.debug(f"✓ {elem_name} 已加载")
                        except:
                            pass
                
                # NOTE: legacy comment removed for English-only repo.
                if len(loaded_elements) == len(key_elements):
                    self.logger.info(f"所有关键元素已加载完成 ({len(loaded_elements)}/{len(key_elements)})，开始提取数据")
                    break
                    
            except Exception as e:
                self.logger.debug(f"检查页面元素时出错: {e}")
            
            wait_count += 1
            self.delay(1)
            
            if wait_count % 5 == 0:
                self.logger.info(f"已等待 {wait_count} 秒，已加载 {len(loaded_elements)}/{len(key_elements)} 个元素...")
                missing = set([name for name, _ in key_elements]) - loaded_elements
                if missing:
                    self.logger.info(f"  仍在等待: {', '.join(missing)}")
        
        # NOTE: legacy comment removed for English-only repo.
        if wait_count >= max_wait:
            self.logger.error(f"等待超时({max_wait}秒)，已加载 {len(loaded_elements)}/{len(key_elements)} 个元素")
            missing = set([name for name, _ in key_elements]) - loaded_elements
            if missing:
                self.logger.error(f"  未加载的元素: {', '.join(missing)}")
            
            # NOTE: legacy comment removed for English-only repo.
            self.logger.error(f"详情页加载失败，返回None触发重试")
            return None  # NOTE: legacy comment removed for English-only repo.
        
        # NOTE: legacy comment removed for English-only repo.
        self.delay(2)

        creator_data = {
            'region': self.region,
            'brand_name': self.brand_name,
            'search_keywords': self.search_keywords_raw,
            'creator_name': '',
            'categories': '',
            'followers': '',
            'intro': '',
            'sales_revenue': '',
            'sales_units_sold': '',
            'sales_gpm': '',
            'sales_revenue_per_buyer': '',
            'gmv_per_sales_channel': '',
            'gmv_by_product_category': '',
            'avg_commission_rate': '',
            'collab_products': '',
            'partnered_brands': '',
            'product_price': '',
            'video_gpm': '',
            'videos': '',
            'avg_video_views': '',
            'avg_video_engagement_rate': '',
            'avg_video_likes': '',
            'avg_video_comments': '',
            'avg_video_shares': '',
            'live_gpm': '',
            'live_streams': '',
            'avg_live_views': '',
            'avg_live_engagement_rate': '',
            'avg_live_likes': '',
            'avg_live_comments': '',
            'avg_live_shares': '',
            'followers_male': '',
            'followers_female': '',
            'followers_18_24': '',
            'followers_25_34': '',
            'followers_35_44': '',
            'followers_45_54': '',
            'followers_55_more': '',
            'creator_chaturl': '',
            'creator_id': '',
            'partner_id': '',
            'connect': False,
            'reply': False,
            'send': False,
            'send_time': '',
            'top_brands':'',
            'whatsapp': '',
            'email': '',
            
        }
        
        # NOTE: legacy comment removed for English-only repo.
        creator_data['creator_name'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[1]/div/span[1]/span[1]')
        if not creator_data['creator_name']:  # NOTE: legacy comment removed for English-only repo.
            creator_data['creator_name'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[1]/div/span[1]/span')
            
        if creator_data['creator_name']:
            self.logger.info(f"提取到creator_name: {creator_data['creator_name']}")
        else:
            self.logger.warning(f"未能提取creator_name")

        # NOTE: legacy comment removed for English-only repo.
        current_url = page.url
        self.logger.info(f"当前详情页URL: {current_url}")
        
        # NOTE: legacy comment removed for English-only repo.
        partner_match = re.search(r'partner_id=(\d+)', current_url)
        if partner_match:
            creator_data['partner_id'] = partner_match.group(1)
            self.logger.info(f"提取到partner_id: {creator_data['partner_id']}")
        
        # NOTE: legacy comment removed for English-only repo.
        cid_match = re.search(r'cid=(\d+)', current_url)
        if cid_match:
            creator_data['creator_id'] = cid_match.group(1)
            self.logger.info(f"提取到creator_id (cid): {creator_data['creator_id']}")
    

        try:
            # NOTE: legacy comment removed for English-only repo.
            creator_data['categories'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[1]/span[1]/span[2]/span/span')
            creator_data['followers'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[1]/span[2]/span[2]/span/span')            
            creator_data['intro'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[2]/div/div/div[1]/div[2]/div[2]/div[3]/div/span')
            
            # NOTE: legacy comment removed for English-only repo.
            creator_data['sales_revenue'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[1]/div/div[1]/div[2]/span/span')
            creator_data['sales_units_sold'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[1]/div/div[2]/div[2]/span/span')
            creator_data['sales_gpm'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[1]/div/div[3]/div[2]/span/span')
            creator_data['sales_revenue_per_buyer'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[1]/div/div[4]/div[2]/span/span')
   
            # NOTE: legacy comment removed for English-only repo.
            gmv_channel_parts = []
            gmv_channel_xpaths = [
                '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[2]/div[2]/div/div/div/div/div',
                '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[2]/div/div/div/div/div/div[1]/div/div[2]/div[2]/div[3]/div/div/div/div/div'
            ]
            for xpath in gmv_channel_xpaths:
                text = self.safe_extract_text(page, xpath)
                if text:
                    gmv_channel_parts.append(text)
            creator_data['gmv_per_sales_channel'] = ' | '.join(gmv_channel_parts)
            
            # NOTE: legacy comment removed for English-only repo.
            gmv_category_parts = []
            gmv_category_xpaths = [
                '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div/div/div',
                '//*[@id="submodule_layout_container_id"]/div[5]/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div/div'
            ]
            for xpath in gmv_category_xpaths:
                text = self.safe_extract_text(page, xpath)
                if text:
                    gmv_category_parts.append(text)
            creator_data['gmv_by_product_category'] = ' | '.join(gmv_category_parts)
            
            # NOTE: legacy comment removed for English-only repo.
            creator_data['avg_commission_rate'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[1]/div[2]/span/span')
            creator_data['collab_products'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[2]/div[2]/span/span')
            creator_data['partnered_brands'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[3]/div[2]/span/span')
            # NOTE: legacy comment removed for English-only repo.
            self.logger.info("爬取 top brands...")
            top_brands_list = []
            try:
                # NOTE: legacy comment removed for English-only repo.
                partnered_brands_count = 0
                if creator_data['partnered_brands']:
                    match = re.search(r'(\d+)', creator_data['partnered_brands'])
                    if match:
                        partnered_brands_count = int(match.group(1))
                        self.logger.info(f"检测到 partnered_brands 数量: {partnered_brands_count}")
                if partnered_brands_count > 0:
                    # NOTE: legacy comment removed for English-only repo.
                    view_brands_button = page.locator('xpath=//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[3]/div[2]/button')
                    if view_brands_button.count() > 0:
                        view_brands_button.first.click()
                        self.logger.info("已点击 'View top brands' 按钮")
                        self.delay(1.5)  # NOTE: legacy comment removed for English-only repo.
                        for i in range(2, partnered_brands_count + 2):
                            brand_xpath = f'//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[3]/div[2]/div/span/div[1]/div/div/div/div/div/div[{i}]'
                            brand_text = self.safe_extract_text(page, brand_xpath)
                            if brand_text and brand_text.strip():
                                top_brands_list.append(brand_text.strip())
                                self.logger.debug(f"找到品牌 [{i-1}]: {brand_text.strip()}")
                        self.logger.info(f"成功爬取 {len(top_brands_list)} 个 top brands (预期 {partnered_brands_count} 个，注意这里上限是10个)")
                    else:
                        self.logger.warning("未找到 'View top brands' 按钮")
                else:
                    self.logger.info("partnered_brands 数量为 0，跳过爬取 top brands")
            except Exception as e:
                self.logger.warning(f"爬取 top brands 失败: {e}")

            creator_data['top_brands'] = ', '.join(top_brands_list) if top_brands_list else '' # NOTE: legacy comment removed for English-only repo.
            creator_data['product_price'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[7]/div[2]/div/div/div/div[4]/div[2]/span/span')

            
            # NOTE: legacy comment removed for English-only repo.
            creator_data['video_gpm'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[1]/div[1]/div[2]/span/div/div/div/span')
            creator_data['videos'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[1]/div[2]/div[2]/span/span')
            creator_data['avg_video_views'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[1]/div[3]/div[2]/span/span')
            creator_data['avg_video_engagement_rate'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[1]/div[4]/div[2]/span/span')
            
            try:
                # NOTE: legacy comment removed for English-only repo.
                click_xpaths = [
                    '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[2]',
                    '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[3]',
                    '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[3]'
                ]
                for xpath in click_xpaths:
                    element = page.locator(f'xpath={xpath}')
                    if element.count() > 0:
                        element.first.click()
                        self.logger.info(f"点击了元素: {xpath}")
                        self.delay(0.1)

                # NOTE: legacy comment removed for English-only repo.
                creator_data['avg_video_likes'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[2]/div[2]/div[2]/span/span')
                creator_data['avg_video_comments'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[2]/div[3]/div[2]/span/span')
                creator_data['avg_video_shares'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[9]/div[2]/div[2]/div[4]/div[2]/span/span')
            except Exception as e:
                self.logger.warning(f"获取视频详细数据失败: {e}")
            
            # NOTE: legacy comment removed for English-only repo.
            creator_data['live_gpm'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[1]/div[1]/div[2]/span/div/div/div/span')
            creator_data['live_streams'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[1]/div[2]/div[2]/span/span')
            creator_data['avg_live_views'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[1]/div[3]/div[2]/span/span')
            creator_data['avg_live_engagement_rate'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[1]/div[4]/div[2]/span/span')
            

            try:
                # NOTE: legacy comment removed for English-only repo.
                click_xpaths = [
                    '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[2]',
                    '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[3]',
                    '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[3]'
                ]
                for xpath in click_xpaths:
                    element = page.locator(f'xpath={xpath}')
                    if element.count() > 0:
                        element.first.click()
                        self.logger.info(f"点击了元素: {xpath}")
                        self.delay(1)

                # NOTE: legacy comment removed for English-only repo.
                creator_data['avg_live_likes'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[2]/div[2]/div[2]/span/span')
                creator_data['avg_live_comments'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[2]/div[3]/div[2]/span/span')
                creator_data['avg_live_shares'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[11]/div[2]/div/div/div[2]/div[4]/div[2]/span/span')
            
            except Exception as e:
                self.logger.warning(f"获取直播详细数据失败: {e}")
            
            # NOTE: legacy comment removed for English-only repo.
            creator_data['followers_male'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[1]/div/div[2]/div[2]/div[3]/div[1]/div/div/div/div')
            creator_data['followers_female'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[1]/div/div[2]/div[2]/div[3]/div[2]/div/div/div/div')
            creator_data['followers_18_24'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]/div[2]/div[3]/div[1]/div/div/div/div')
            creator_data['followers_25_34'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]/div[2]/div[3]/div[2]/div/div/div/div')
            creator_data['followers_35_44'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]/div[2]/div[3]/div[3]/div/div/div/div')
            creator_data['followers_45_54'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]/div[2]/div[3]/div[4]/div/div/div/div')
            creator_data['followers_55_more'] = self.safe_extract_text(page, '//*[@id="submodule_layout_container_id"]/div[13]/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]/div[2]/div[3]/div[5]/div/div/div/div')
            
        except Exception as e:
            self.logger.error(f"提取达人详情失败: {e}")
        
        return creator_data

    def navigate_to_chat_page(self, page, partner_id: str, creator_id: str) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info(f"准备跳转到聊天页面 - partner_id: {partner_id}, creator_id: {creator_id}")
        
        if not creator_id:
            self.logger.error("缺少creator_id，无法跳转到聊天页面")
            return False
        
        market_mapping = {
            'MX': '19',  # NOTE: legacy comment removed for English-only repo.
            'FR': '17',  # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
            # NOTE: legacy comment removed for English-only repo.
        }

        # NOTE: legacy comment removed for English-only repo.
        market_id = market_mapping.get(self.region, '19')
        
        # NOTE: legacy comment removed for English-only repo.
        if self.region in ['FR']:
            base_domain = 'partner.eu.tiktokshop.com'
        else:
            base_domain = 'partner.tiktokshop.com'

        # NOTE: legacy comment removed for English-only repo.
        chat_url = f"https://{base_domain}/partner/im?creator_id={creator_id}&market={market_id}&enter_from=find_creator_detail"       
        # NOTE: legacy comment removed for English-only repo.
        MAX_PAGE_RETRIES = 3
        
        for page_attempt in range(1, MAX_PAGE_RETRIES + 1):
            try:
                self.logger.info(f"第 {page_attempt}/{MAX_PAGE_RETRIES} 次尝试加载聊天页面")
                
                # NOTE: legacy comment removed for English-only repo.
                if page_attempt > 1:
                    self.logger.info("准备重新加载页面...")
                    try:
                        # NOTE: legacy comment removed for English-only repo.
                        context = page.context
                        if context and hasattr(context, 'new_page'):
                            # NOTE: legacy comment removed for English-only repo.
                            self.logger.info("尝试在新标签页中打开...")
                            
                            # NOTE: legacy comment removed for English-only repo.
                            new_page = context.new_page()
                            
                            # NOTE: legacy comment removed for English-only repo.
                            try:
                                page.close()
                            except:
                                pass
                            
                            # NOTE: legacy comment removed for English-only repo.
                            page = new_page
                            self.delay(1)
                        else:
                            # NOTE: legacy comment removed for English-only repo.
                            self.logger.info("刷新当前页面...")
                            page.goto("about:blank", wait_until="load", timeout=5000)
                            self.delay(1)
                    except Exception as e:
                        self.logger.warning(f"无法创建新页面，将重新导航: {e}")
                        # NOTE: legacy comment removed for English-only repo.
                        pass
                
                # NOTE: legacy comment removed for English-only repo.
                self.logger.info(f"跳转到聊天URL: {chat_url}")
                try:
                    page.goto(chat_url, wait_until="networkidle", timeout=60000)

                    # NOTE: legacy comment removed for English-only repo.
                except Exception as e:
                    self.logger.error(f"页面导航失败: {e}")
                    if page_attempt < MAX_PAGE_RETRIES:
                        self.logger.info("等待3秒后重试...")
                        self.delay(3)
                        continue
                    else:
                        self.logger.error("页面导航多次失败，放弃")
                        return False
                
                # NOTE: legacy comment removed for English-only repo.
                self.delay(2)
                
                # NOTE: legacy comment removed for English-only repo.
                current_url = page.url
                self.logger.info(f"当前URL: {current_url}")
                
                if "login" in current_url.lower() or "signin" in current_url.lower():
                    self.logger.error("页面重定向到登录页，需要重新登录")
                    return False
                
                # NOTE: legacy comment removed for English-only repo.
                self.logger.info(f"等待聊天输入框出现（最多等待60秒）...")
                
                MAX_WAIT_SECONDS = 60  # NOTE: legacy comment removed for English-only repo.
                wait_interval = 2  # NOTE: legacy comment removed for English-only repo.
                MAX_RETRY_COUNT = MAX_WAIT_SECONDS // wait_interval
                retry_count = 0
                
                while retry_count < MAX_RETRY_COUNT:
                    try:
                        # NOTE: legacy comment removed for English-only repo.
                        iframe_found = False
                        try:
                            iframes = page.frames
                            if len(iframes) > 1:
                                self.logger.debug(f"页面包含 {len(iframes)} 个frame")
                                # NOTE: legacy comment removed for English-only repo.
                                for frame in iframes:
                                    try:
                                        if frame.locator('textarea').count() > 0:
                                            self.logger.info("在iframe中发现textarea")
                                            iframe_found = True
                                            break
                                    except:
                                        continue
                        except:
                            pass
                        
                        # NOTE: legacy comment removed for English-only repo.
                        textarea_selectors = [
                            'textarea[placeholder="Send a message"]',
                            'textarea[placeholder*="Send a message"]',
                            'textarea.index-module__textarea--qYh62',
                            'textarea[data-e2e="798845f5-2eb9-0980"]',
                            'textarea[placeholder="发送消息"]',
                            'textarea[placeholder*="发送消息"]',
                            'textarea[placeholder*="message" i]',
                            '#im_sdk_chat_input textarea',
                            'div[data-e2e="cda68c25-5112-89c2"] textarea',
                            'textarea'
                        ]
                        
                        found_textarea = False
                        for selector in textarea_selectors:
                            try:
                                textarea = page.locator(selector).first
                                if textarea.count() > 0:
                                    # NOTE: legacy comment removed for English-only repo.
                                    placeholder = None
                                    try:
                                        placeholder = textarea.get_attribute('placeholder', timeout=1000)
                                    except:
                                        pass
                                    
                                    # NOTE: legacy comment removed for English-only repo.
                                    if selector == 'textarea' and placeholder is None:
                                        # NOTE: legacy comment removed for English-only repo.
                                        continue
                                    
                                    if (placeholder is None or 
                                        placeholder == "" or
                                        'Send a message' in placeholder or 
                                        '发送消息' in placeholder or 
                                        'message' in placeholder.lower()):
                                        # NOTE: legacy comment removed for English-only repo.
                                        try:
                                            is_visible = textarea.is_visible(timeout=1000)
                                            if is_visible:
                                                self.logger.info(f"成功找到聊天输入框: {selector}")
                                                if placeholder:
                                                    self.logger.info(f"Placeholder内容: {placeholder}")
                                                found_textarea = True
                                                break
                                        except:
                                            # NOTE: legacy comment removed for English-only repo.
                                            continue
                            except:
                                continue
                        
                        if found_textarea:
                            self.logger.info("聊天页面已完全加载")
                            
                            # NOTE: legacy comment removed for English-only repo.
                            self.delay(1.5)
                            
                            # NOTE: legacy comment removed for English-only repo.
                            try:
                                final_check = page.locator('textarea').first
                                if final_check.count() > 0:
                                    self.logger.info("最终确认：聊天输入框存在")
                                    return True
                            except:
                                # NOTE: legacy comment removed for English-only repo.
                                self.logger.info("最终验证失败，但之前已找到输入框，认为成功")
                                return True
                        
                        # NOTE: legacy comment removed for English-only repo.
                        retry_count += 1
                        
                        # NOTE: legacy comment removed for English-only repo.
                        if retry_count % 5 == 0:
                            self.logger.info(f"等待聊天输入框... ({retry_count * wait_interval}/{MAX_WAIT_SECONDS}秒)")
                            
                            # NOTE: legacy comment removed for English-only repo.
                            current_url = page.url
                            if "partner/im" not in current_url:
                                self.logger.warning(f"页面已离开聊天页面: {current_url}")
                                break
                        
                        # NOTE: legacy comment removed for English-only repo.
                        self.delay(wait_interval)
                        
                    except Exception as e:
                        retry_count += 1
                        if retry_count == 1:
                            self.logger.debug(f"检查输入框时出现异常: {e}")
                        self.delay(wait_interval)
                
                # NOTE: legacy comment removed for English-only repo.
                self.logger.warning(f"第 {page_attempt} 次尝试超时（等待了{MAX_WAIT_SECONDS}秒）")
                
                if page_attempt < MAX_PAGE_RETRIES:
                    self.logger.info(f"准备第 {page_attempt + 1} 次重试...")
                    self.delay(2)
                else:
                    # NOTE: legacy comment removed for English-only repo.
                    self.logger.error("所有尝试都失败了")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    current_url = page.url
                    if "partner/im" in current_url and creator_id in current_url:
                        self.logger.warning("URL正确但未找到输入框，可能页面结构已变化")
                        # NOTE: legacy comment removed for English-only repo.
                        return False
                    else:
                        self.logger.error(f"URL不正确: {current_url}")
                        return False
                        
            except Exception as e:
                self.logger.error(f"第 {page_attempt} 次尝试出现异常: {e}")
                import traceback
                self.logger.debug(f"异常详情: {traceback.format_exc()}")
                
                if page_attempt < MAX_PAGE_RETRIES:
                    self.delay(3)
                    continue
                else:
                    return False
        self.logger.error("函数异常结束")
        return False


    def _fill_chat_input(self, page, text: str, auto_send: bool = False) -> bool:
        if not text:
            return False

        selectors = [
            'textarea[placeholder="Send a message"]',
            'textarea[placeholder*="Send a message"]'
        ]
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.count() > 0:
                    el.wait_for(state="visible", timeout=4000)
                    el.click()

                    el.fill(text)

                    self.logger.info("成功一次性填入完整文本（保留换行）")

                    if auto_send:
                        self.delay(0.5)  # NOTE: legacy comment removed for English-only repo.
                        el.press("Enter")
                        self.logger.info("消息已发送")
                    else:
                        self.logger.info("文本已填入，未发送")
                    return True
            except Exception as e:
                self.logger.debug(f"尝试选择器 {sel} 写入失败：{e}")

        # NOTE: legacy comment removed for English-only repo.
        try:
            page.evaluate(
                """Legacy docstring removed for English-only repo.""",
                text
            )
            if auto_send:
                page.keyboard.press("Enter")
                self.logger.info("消息已发送（兜底方案）")
            else:
                self.logger.info("文本已填入（兜底方案，未发送）")
            return True
        except Exception as e:
            self.logger.warning(f"JS 写入聊天框失败：{e}")
            return False

    def verify_message_sent(self, page, timeout: int = 5) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("验证消息是否发送成功...")
        
        try:
            # NOTE: legacy comment removed for English-only repo.
            for i in range(timeout):
                self.delay(1)
                
                # NOTE: legacy comment removed for English-only repo.
                has_message = page.evaluate("""
                    () => {
                        const container = document.querySelector('div.index-module__messageList--GBz6X') ||
                                        document.querySelector('div.messageList-k_OG24');
                        if (!container) return false;
                        
                        const messages = container.querySelectorAll('div.chatd-message--right');
                        return messages.length > 0;
                    }
                """)
                
                if has_message:
                    self.logger.info(f"✓ 检测到发送的消息（等待了 {i+1} 秒）")
                    return True
            
            self.logger.warning(f"等待 {timeout} 秒后仍未检测到消息")
            return False
            
        except Exception as e:
            self.logger.error(f"验证消息发送失败: {e}")
            return False

    def wait_for_chat_page(self, page, creator_name: str) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info(f"等待 {creator_name} 的聊天页面加载...")
        
        try:
            page.wait_for_selector(f'div:has-text("{creator_name}")', timeout=10000)
            self.logger.info(f"聊天页面已加载")
            return True
        except:
            self.logger.warning(f"聊天页面加载超时")
            return False
    
    def extract_chat_url_and_id(self, page) -> tuple:
        """Legacy docstring removed for English-only repo."""
        current_url = page.url
        self.logger.info(f"当前URL: {current_url}")
        
        creator_id = ""
        match = re.search(r'creator_id=(\d+)', current_url)
        if match:
            creator_id = match.group(1)
            self.logger.info(f"提取到creator_id: {creator_id}")
        
        return current_url, creator_id
    
    def check_connection_status(self, page) -> tuple:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("检查聊天连接状态与回复状态...")

        try:
            # NOTE: legacy comment removed for English-only repo.
            page.wait_for_load_state("networkidle")
            self.delay(3)
            
            # NOTE: legacy comment removed for English-only repo.
            chat_container = page.evaluate("""
                () => {
                    const selectors = [
                        'div.index-module__messageList--GBz6X',
                        'div.messageList-k_OG24',
                        'div.chatd-scrollView',
                        'div[class*="messageList"]',
                        'div[class*="chatd-scrollView"]'
                    ];
                    
                    for (const selector of selectors) {
                        const container = document.querySelector(selector);
                        if (container && container.offsetParent !== null) {
                            return selector;
                        }
                    }
                    return null;
                }
            """)
            
            if not chat_container:
                self.logger.warning("未找到聊天容器")
                return False, False
            
            # NOTE: legacy comment removed for English-only repo.
            messages = page.evaluate(f"""Legacy docstring removed for English-only repo.""")

            if not isinstance(messages, list):
                messages = []

            self.logger.info(f"找到 {len(messages)} 条历史消息")
            
            # NOTE: legacy comment removed for English-only repo.
            connect = len(messages) > 0
            
            # NOTE: legacy comment removed for English-only repo.
            reply = any(m.get('isFromCreator', False) for m in messages)
            
            self.logger.info(f"连接状态检查结果 - 历史消息数: {len(messages)}, connect={connect}, reply={reply}")
            return connect, reply

        except Exception as e:
            self.logger.error(f"检查连接/回复状态失败: {e}")
            return False, False
    
    def close_chat_and_details(self, page) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("关闭聊天窗口和达人详情页...")
        
        try:
            page.keyboard.press("Escape")
            self.delay(1)
            page.keyboard.press("Escape")
            self.delay(2)
            self.logger.info("已关闭窗口")
            return True
        except Exception as e:
            self.logger.error(f"关闭窗口失败: {e}")
            return False

    def scroll_page_for_more(self, page) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("滚动页面加载更多达人...")
        
        try:
            # NOTE: legacy comment removed for English-only repo.
            scrolled = page.evaluate("""Legacy docstring removed for English-only repo.""")
            
            if scrolled.get('scrolled'):
                self.logger.info(f"成功滚动容器: {scrolled.get('containerId') or scrolled.get('containerClass')}")
                self.logger.info(f"容器高度: {scrolled.get('scrollHeight')}, 可见高度: {scrolled.get('clientHeight')}, 当前位置: {scrolled.get('scrollTop')}")
            else:
                self.logger.warning(f"未找到可滚动容器: {scrolled.get('message')}")
                
            self.delay(1.5)
            return True
            
        except Exception as e:
            self.logger.error(f"滚动页面失败: {e}")
            return False

    def extract_contact_info(self, page) -> tuple:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("尝试获取达人联系方式...")
        whatsapp = ""
        email = ""
        
        import time
        start_time = time.time()
        MAX_TOTAL_TIME = 15
        
        try:
            # NOTE: legacy comment removed for English-only repo.
            contact_button_xpath = '//*[@id="arco-tabs-0-panel-0"]/div/div/div[1]/button'
            
            try:
                button = page.locator(f'xpath={contact_button_xpath}').first
                if button.count() > 0:
                    button.wait_for(state="visible", timeout=2000)
                    button.click()
                    self.logger.info("成功点击联系方式按钮")
                    self.delay(1)  # NOTE: legacy comment removed for English-only repo.
                else:
                    self.logger.info("未找到联系方式按钮，跳过")
                    return whatsapp, email
            except Exception as e:
                self.logger.info(f"点击联系方式按钮失败，跳过: {e}")
                return whatsapp, email

            if time.time() - start_time > MAX_TOTAL_TIME:
                self.logger.warning("获取联系方式超时，跳过")
                return whatsapp, email

            try:
                unavailable_locator = page.locator("text=This creator doesn't have contact information available.")
                unavailable_locator.wait_for(state="visible", timeout=2000)
                
                if unavailable_locator.count() > 0:
                    self.logger.info("检测到提示：该达人没有联系方式")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    got_it_xpath = '/html/body/div[4]/div[2]/div/div[2]/div[3]/div/button'
                    try:
                        got_it_btn = page.locator(f'xpath={got_it_xpath}').first
                        if got_it_btn.count() > 0:
                            got_it_btn.click()
                            self.logger.info("已点击 'Got it' 按钮")
                            self.delay(0.5)
                    except Exception as e:
                        self.logger.debug(f"点击 'Got it' 按钮失败: {e}")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    close_modal_xpath = '/html/body/div[3]/div[2]/div/div[2]/span'
                    try:
                        close_btn = page.locator(f'xpath={close_modal_xpath}').first
                        if close_btn.count() > 0:
                            close_btn.click()
                            self.logger.info("已关闭 'Contact info unavailable' 弹窗")
                            self.delay(0.5)
                    except Exception as e:
                        self.logger.debug(f"关闭 'Contact info unavailable' 弹窗失败: {e}")
                    
                    return whatsapp, email

            except Exception:
                # NOTE: legacy comment removed for English-only repo.
                pass

            if time.time() - start_time > MAX_TOTAL_TIME:
                self.logger.warning("获取联系方式超时，跳过")
                return whatsapp, email

            # NOTE: legacy comment removed for English-only repo.
            base_xpath = '/html/body/div[3]/div[2]/div/div[2]/div[2]/div'
            for i in range(2, 4): 
                if time.time() - start_time > MAX_TOTAL_TIME:
                    self.logger.warning("获取联系方式超时，使用已获取的结果")
                    break
                
                try:
                    label_xpath = f'xpath={base_xpath}/div[{i}]/span'
                    label_el = page.locator(label_xpath).first
                    
                    if label_el.count() == 0:
                        continue
                    
                    # NOTE: legacy comment removed for English-only repo.
                    label_text = label_el.text_content(timeout=1000)  # NOTE: legacy comment removed for English-only repo.
                    
                    if not label_text or not label_text.strip():
                        continue
                    
                    # NOTE: legacy comment removed for English-only repo.
                    value_xpath = f'xpath={base_xpath}/div[{i}]/div/div/span'
                    value_el = page.locator(value_xpath).first
                    
                    if value_el.count() == 0:
                        continue
                        
                    value_text = value_el.text_content(timeout=1000)  # NOTE: legacy comment removed for English-only repo.
                    
                    if not value_text or not value_text.strip():
                        continue
                    
                    # NOTE: legacy comment removed for English-only repo.
                    label_lower = label_text.lower()
                    if 'whatsapp' in label_lower:
                        whatsapp = value_text.strip()
                        self.logger.info(f"找到 WhatsApp: {whatsapp}")
                    elif 'email' in label_lower or 'e-mail' in label_lower:
                        email = value_text.strip()
                        self.logger.info(f"找到 Email: {email}")
                    
                except Exception as e:
                    self.logger.debug(f"检查位置 div[{i}] 时出错: {e}")
                    continue  # NOTE: legacy comment removed for English-only repo.
            
            # NOTE: legacy comment removed for English-only repo.
            try:
                close_button_xpath = 'xpath=/html/body/div[3]/div[2]/div/div[2]/span'
                close_btn = page.locator(close_button_xpath).first
                if close_btn.count() > 0:
                    close_btn.click(timeout=1000)  # NOTE: legacy comment removed for English-only repo.
                    self.logger.info("已关闭联系方式弹窗")
                    self.delay(0.3) 
            except Exception as e:
                self.logger.debug(f"关闭弹窗失败: {e}")
                # NOTE: legacy comment removed for English-only repo.
                try:
                    page.keyboard.press("Escape")
                    self.delay(0.3)
                except:
                    pass
            
        except Exception as e:
            self.logger.warning(f"获取联系方式失败: {e}")
        
        return whatsapp, email

    def process_single_creator(self, page, creator_name: str) -> Dict[str, Any]:
        """Legacy docstring removed for English-only repo."""
        self.raise_if_cancelled(f"处理达人 {creator_name}")
        
        # NOTE: legacy comment removed for English-only repo.
        msg_type, message = self.should_send_message(creator_name)

        if msg_type is None:
            self.logger.info(f"达人 {creator_name} 不符合发送条件，跳过")
            return None

        max_retries = 3
        attempt = 0
        creator_data = None
        detail_page = None
        is_new_tab = False

        while attempt < max_retries:
            self.raise_if_cancelled(f"处理达人 {creator_name} 第 {attempt+1} 次尝试")
            attempt += 1
            base_pages = set(page.context.pages)
            
            try:
                self.logger.info(f"[页面数量] 处理前: {len(page.context.pages)}")
                self.logger.info(f"[活跃页面数] 当前管理的页面: {len(self.active_pages)}")
                self.logger.info(f"处理达人 {creator_name}，第 {attempt} 次尝试")

                # NOTE: legacy comment removed for English-only repo.
                if attempt > 1 and detail_page and not detail_page.is_closed():
                    try:
                        detail_page.close()
                        self.logger.info("关闭上次失败的详情页")
                        self.active_pages.discard(detail_page)
                    except:
                        pass
                    self.delay(1)

                detail_page, is_new_tab = self._open_detail_and_get_page(page, creator_name)
                self.active_pages.add(detail_page)
                self.logger.info(f"[页面管理] 已添加详情页到管理列表")
                self.logger.info(f"[页面数量] 打开详情页后: {len(page.context.pages)}")

                creator_data = self.extract_creator_details(detail_page, creator_name)
                
                if creator_data is None:
                    self.logger.warning(f"第 {attempt} 次尝试：详情页加载失败，准备重试")
                    if is_new_tab and detail_page and not detail_page.is_closed():
                        detail_page.close()
                        self.active_pages.discard(detail_page)
                    if attempt < max_retries:
                        self.logger.info(f"等待1秒后重新点击达人 {creator_name}")
                        self.delay(1)
                        continue
                    else:
                        self.logger.error(f"达人 {creator_name} 详情页加载最终失败")
                        return None
                
                if not creator_data.get('partner_id') or not creator_data.get('creator_id'):
                    self.logger.warning(f"第 {attempt} 次尝试：未获取到partner_id或creator_id")
                    if attempt < max_retries:
                        if is_new_tab and detail_page and not detail_page.is_closed():
                            detail_page.close()
                            self.active_pages.discard(detail_page)
                        self.delay(1.5)
                        continue
                
                # NOTE: legacy comment removed for English-only repo.
                if creator_data.get('partner_id') and creator_data.get('creator_id'):
                    if self.navigate_to_chat_page(detail_page, creator_data['partner_id'], creator_data['creator_id']):
                        try:
                            creator_data['creator_chaturl'] = detail_page.url
                            
                            # NOTE: legacy comment removed for English-only repo.
                            actual_connect, reply = self.check_connection_status(detail_page)
                            creator_data['connect'] = actual_connect
                            creator_data['reply'] = reply if actual_connect else False
                            
                            # NOTE: legacy comment removed for English-only repo.
                            # NOTE: legacy comment removed for English-only repo.
                            if actual_connect and msg_type == 'first':
                                self.logger.info(f"{creator_name} 初判为新达人，但实际已建联 → 改用 later 消息")
                                msg_type = 'later'
                                message = self._load_email_later_msg()
                            
                            # NOTE: legacy comment removed for English-only repo.
                            should_attempt_send = False
                            if msg_type == 'first' and not actual_connect:
                                # NOTE: legacy comment removed for English-only repo.
                                should_attempt_send = True
                                self.logger.info(f"→ 发送 first 消息（真正的新达人）")
                            elif msg_type == 'later':
                                # NOTE: legacy comment removed for English-only repo.
                                should_attempt_send = True
                                self.logger.info(f"→ 发送 later 消息（已建联达人）")

                            if should_attempt_send and message:
                                self.logger.info(f"准备发送 {msg_type} 消息...")
                                
                                # NOTE: legacy comment removed for English-only repo.
                                MAX_SEND_RETRIES = 3
                                send_success = False
                                
                                for send_attempt in range(1, MAX_SEND_RETRIES + 1):
                                    self.logger.info(f"发送消息尝试 {send_attempt}/{MAX_SEND_RETRIES}")
                                    
                                    ok = self._fill_chat_input(detail_page, message, auto_send=True)
                                    
                                    if ok:
                                        self.logger.info("消息发送命令已执行，等待2秒后验证...")
                                        self.delay(2)
                                        
                                        # NOTE: legacy comment removed for English-only repo.
                                        new_connect, _ = self.check_connection_status(detail_page)
                                        
                                        if new_connect or msg_type == 'later':
                                            creator_data['send'] = True
                                            creator_data['send_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                            creator_data['connect'] = True
                                            self.logger.info(f"✓ {msg_type}消息发送成功，时间: {creator_data['send_time']}")
                                            send_success = True
                                            break
                                        else:
                                            self.logger.warning(f"第 {send_attempt} 次发送后未检测到消息...")
                                            self.delay(1.5)
                                    else:
                                        self.logger.warning(f"第 {send_attempt} 次发送命令执行失败")
                                        self.delay(1.5)
                                
                                if not send_success:
                                    self.logger.error(f"消息发送失败（尝试了 {MAX_SEND_RETRIES} 次）")
                                    creator_data['send'] = False
                                    creator_data['send_time'] = ''
                            else:
                                self.logger.info(f"跳过发送: msg_type={msg_type}, actual_connect={actual_connect}")
                                creator_data['send'] = False
                                creator_data['send_time'] = ''
                            
                            # NOTE: legacy comment removed for English-only repo.
                            whatsapp, email = self.extract_contact_info(detail_page)
                            creator_data['whatsapp'] = whatsapp
                            creator_data['email'] = email
                            self.logger.info(f"联系方式获取完成 - WhatsApp: {whatsapp or '无'}, Email: {email or '无'}")

                        except Exception as e:
                            self.logger.warning(f"处理聊天页面时出错: {e}")
                            creator_data['connect'] = False
                            creator_data['reply'] = False
                            creator_data['whatsapp'] = ''
                            creator_data['email'] = ''
                    else:
                        self.logger.warning(f"无法跳转到 {creator_name} 的聊天页面")
                        creator_data['connect'] = False
                        creator_data['reply'] = False
                        creator_data['whatsapp'] = ''
                        creator_data['email'] = ''
                else:
                    self.logger.warning(f"未能获取 {creator_name} 的partner_id或creator_id")
                    creator_data['connect'] = False
                    creator_data['reply'] = False
                    creator_data['whatsapp'] = ''
                    creator_data['email'] = ''

                # NOTE: legacy comment removed for English-only repo.
                try:
                    if is_new_tab and detail_page and not detail_page.is_closed():
                        detail_page.close()
                        self.logger.info(f"已关闭 {creator_name} 的页面")
                except Exception as close_error:
                    self.logger.warning(f"关闭页面时出错: {close_error}")
                finally:
                    self.active_pages.discard(detail_page)

                self.logger.info(f"[页面数量] 处理完成后: {len(page.context.pages)}")
                self.logger.info(f"[活跃页面数] 处理完成后管理的页面: {len(self.active_pages)}")
                self.cleanup_stale_pages()

                return creator_data

            except Exception as e:
                self.logger.error(f"处理达人 {creator_name} 失败（第 {attempt} 次尝试）: {e}")
                if attempt < max_retries:
                    self.logger.info("等待 1 秒后重试...")
                    self.delay(1)
                else:
                    self.logger.error(f"达人 {creator_name} 处理已处理/最终失败，跳过")
                    return None
                    
            finally:
                # NOTE: legacy comment removed for English-only repo.
                current_pages = set(page.context.pages)
                extra_pages = current_pages - base_pages
                for p in extra_pages:
                    try:
                        if not p.is_closed():
                            p.close()
                    except:
                        pass
                    finally:
                        self.active_pages.discard(p)

    def save_single_creator(self, creator_data: Dict[str, Any]):
        """Legacy docstring removed for English-only repo."""
        # NOTE: legacy comment removed for English-only repo.
        fieldnames = [
            'creator_name', 'categories', 'followers', 'intro',
            'sales_revenue', 'sales_units_sold', 'sales_gpm', 'sales_revenue_per_buyer',
            'gmv_per_sales_channel', 'gmv_by_product_category',
            'avg_commission_rate', 'collab_products', 'partnered_brands',
            'product_price', 'video_gpm', 'videos', 'avg_video_views', 'avg_video_engagement_rate',
            'avg_video_likes', 'avg_video_comments', 'avg_video_shares',
            'live_gpm', 'live_streams', 'avg_live_views', 'avg_live_engagement_rate',
            'avg_live_likes', 'avg_live_comments', 'avg_live_shares',
            'followers_male', 'followers_female', 'followers_18_24', 'followers_25_34',
            'followers_35_44', 'followers_45_54', 'followers_55_more',
            'creator_chaturl', 'creator_id', 'partner_id', 'connect', 'reply', 'send', 'send_time',
            'brand_name', 'search_keywords', 'top_brands', 'whatsapp', 'email', 'region'
        ]
        
        # NOTE: legacy comment removed for English-only repo.
        detail_fields = [
        #     'categories', 'followers', 'intro',
        #     'sales_revenue', 'sales_units_sold', 'sales_gpm', 'sales_revenue_per_buyer',
        #     'gmv_per_sales_channel', 'gmv_by_product_category',
        #     'avg_commission_rate', 'collab_products', 'partnered_brands',
        #     'product_price', 'video_gpm', 'videos', 'avg_video_views', 'avg_video_engagement_rate',
        #     'avg_video_likes', 'avg_video_comments', 'avg_video_shares',
        #     'live_gpm', 'live_streams', 'avg_live_views', 'avg_live_engagement_rate',
        #     'avg_live_likes', 'avg_live_comments', 'avg_live_shares',
        #     'followers_male', 'followers_female', 'followers_18_24', 'followers_25_34',
        #     'followers_35_44', 'followers_45_54', 'followers_55_more',
            'top_brands', 'whatsapp', 'email'
        ]

        # NOTE: legacy comment removed for English-only repo.
        creator_data['region'] = self.region
        
        if creator_data.get('send', False):
            creator_data['connect'] = True
        
        # NOTE: legacy comment removed for English-only repo.
        from database.ingest_creator_data import log_creator_snapshot_to_db
        try:
            log_creator_snapshot_to_db(
                creator_data,
                task_id=self.task_id,
                shop_name=self.shop_name,
            )
        except Exception as e:
            self.logger.warning(f"数据库埋点失败: {e}")

        if self.xlsx_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.xlsx_filename = os.path.join(self.csv_data_dir, f"creator_scan_{self.shop_name}_{timestamp}.xlsx")
        
        try:
            xlsx_file_exists = os.path.exists(self.xlsx_filename)
            if not xlsx_file_exists:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "达人数据"
                sheet.append(fieldnames)
            else:
                workbook = openpyxl.load_workbook(self.xlsx_filename)
                sheet = workbook.active
            
            row_data = [creator_data.get(field, '') for field in fieldnames]
            sheet.append(row_data)
            workbook.save(self.xlsx_filename)
            
            self.logger.info(f"已保存到任务文件: {self.xlsx_filename}")
            
        except Exception as e:
            self.logger.error(f"保存任务文件失败: {e}")

        # NOTE: legacy comment removed for English-only repo.
        try:
            unified_xlsx_path = Path("data") / f"creator_{self.shop_name}.xlsx"
            unified_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
            
            # NOTE: legacy comment removed for English-only repo.
            lock_path = unified_xlsx_path.with_suffix('.xlsx.lock')
            lock = FileLock(lock_path, timeout=5)  # NOTE: legacy comment removed for English-only repo.
            
            creator_name = creator_data.get('creator_name')
            current_brand = creator_data.get('brand_name')
            current_connect = creator_data.get('connect', False)
            
            # NOTE: legacy comment removed for English-only repo.
            with lock:
                self.logger.info(f"[{creator_name}] 获得Excel文件锁")
                
                if unified_xlsx_path.exists():
                    unified_workbook = openpyxl.load_workbook(unified_xlsx_path)
                    unified_sheet = unified_workbook.active
                    
                    headers = [cell.value for cell in unified_sheet[1]]
                    field_indices = {field: headers.index(field) for field in fieldnames if field in headers}
                    
                    name_col = field_indices['creator_name'] + 1
                    brand_col = field_indices['brand_name'] + 1
                    connect_col = field_indices['connect'] + 1
                    
                    # NOTE: legacy comment removed for English-only repo.
                    rows_to_delete = []
                    updated_rows = 0
                    
                    for row_idx in range(2, unified_sheet.max_row + 1):
                        row_creator_name = unified_sheet.cell(row_idx, name_col).value
                        
                        if row_creator_name == creator_name:
                            # NOTE: legacy comment removed for English-only repo.
                            for field in detail_fields:
                                if field in field_indices:
                                    col_idx = field_indices[field] + 1
                                    unified_sheet.cell(row_idx, col_idx).value = creator_data.get(field, '')
                            updated_rows += 1
                            
                            # NOTE: legacy comment removed for English-only repo.
                            row_brand_name = unified_sheet.cell(row_idx, brand_col).value
                            if current_connect and row_brand_name == current_brand:
                                rows_to_delete.append(row_idx)
                    
                    if updated_rows > 0:
                        self.logger.info(f"同步更新了 {creator_name} 的 {updated_rows} 条旧记录")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    if rows_to_delete:
                        for row_idx in sorted(rows_to_delete, reverse=True):
                            unified_sheet.delete_rows(row_idx)
                            self.logger.info(f"删除旧记录: {creator_name} + {current_brand}")
                    
                else:
                    unified_workbook = openpyxl.Workbook()
                    unified_sheet = unified_workbook.active
                    unified_sheet.title = "达人数据汇总"
                    unified_sheet.append(fieldnames)
                
                # NOTE: legacy comment removed for English-only repo.
                row_data = [creator_data.get(field, '') for field in fieldnames]
                unified_sheet.append(row_data)
                unified_workbook.save(unified_xlsx_path)
                
                self.logger.info(f"[{creator_name}] 已保存并释放Excel文件锁")
            
        except Exception as e:
            self.logger.error(f"处理统一文件失败: {e}")

        creator_name_for_summary = creator_data.get('creator_name')
        if creator_name_for_summary:
            self.latest_creator_name = creator_name_for_summary
            try:
                update_task_info_row(self.task_id, {"connect_creator": creator_name_for_summary})
            except Exception as exc:  # NOTE: legacy comment removed for English-only repo.
                self.logger.warning(f"更新任务概览的 connect_creator 失败: {exc}")
            if self.shared_record_callback:
                try:
                    self.shared_record_callback(
                        {
                            "latest_creator": creator_name_for_summary,
                            "new_creators": self.new_processed_count,
                        }
                    )
                except Exception as exc:  # pragma: no cover
                    self.logger.warning(f"任务进度回调失败: {exc}")

    def run(self, search_keyword: str = None, max_creators: Optional[int] = None) -> bool:
        """Legacy docstring removed for English-only repo."""
        self.logger.info("开始运行创作者建联爬虫...")
        self.restart_requested = False
        self.restart_reason = ""
        self.skipped_total = 0
        self.abort_initiated = False
        self.was_cancelled = False
        self.browser_force_killed = False

        # NOTE: legacy comment removed for English-only repo.
        self.register_cleanup()

        if search_keyword:
            self.search_keyword = search_keyword

        # NOTE: legacy comment removed for English-only repo.
        if max_creators is not None:
            try:
                self.max_creators = int(max_creators)
            except Exception:
                self.logger.warning(f"无效的 max_creators={max_creators}，使用默认 {self.max_creators}")

        self.logger.info(f"本次最大抓取达人数量上限：{self.max_creators}")
        self.ensure_task_info_saved()
        self.raise_if_cancelled("准备启动浏览器")
        self._start_runtime_timer()
        self._start_cancel_watchdog()

        with sync_playwright() as p:
            try:
                # NOTE: legacy comment removed for English-only repo.
                self.playwright_context_active = True
                
                self.browser = p.chromium.launch(
                    headless=True,
                    # channel="chrome",
                    # args=[
                    #     "--no-sandbox",
                    #     "--disable-dev-shm-usage",
                    #     "--disable-gpu",
                    #     "--disable-software-rasterizer",
                    #     "--disable-setuid-sandbox"
                    # ]
                )
                # NOTE: legacy comment removed for English-only repo.
                #     headless=True,
                #     timeout=60000
                # )
                # self.browser = p.webkit.launch(
                #     headless=True,
                #     timeout=60000
                # )
                self.context = self.browser.new_context(
                    viewport={'width': 1920, 'height': 1080}
                )
                self.context.set_default_timeout(60000)
                self.main_page = self.context.new_page()
                self.active_pages.add(self.main_page)
                
                # NOTE: legacy comment removed for English-only repo.
                max_retries = 3
                retry_count = 0

                while retry_count < max_retries:
                    self.raise_if_cancelled("登录流程")
                    retry_count += 1
                    self.logger.info("=" * 50)
                    self.logger.info(f"尝试第 {retry_count}/{max_retries} 次")
                    self.logger.info("=" * 50)
                    
                    # NOTE: legacy comment removed for English-only repo.
                    self.logger.info("步骤1: 登录TikTok")
                    if not self.login(self.main_page):
                        self.logger.error(f"第 {retry_count} 次登录失败")
                        if retry_count == max_retries:
                            self.logger.error("达到最大重试次数，终止流程")
                            return False
                        continue
                    
                    # NOTE: legacy comment removed for English-only repo.
                    self.logger.info("步骤2: 访问创作者连接页面")
                    if self.navigate_to_creator_connection(self.main_page):
                        self.logger.info("成功进入创作者连接页面")
                        break  # NOTE: legacy comment removed for English-only repo.
                    else:
                        self.logger.error(f"第 {retry_count} 次导航失败")
                        if retry_count < max_retries:
                            self.logger.info(f"将重新尝试登录和导航...")
                            self.delay(1)  # NOTE: legacy comment removed for English-only repo.
                        else:
                            self.logger.error("达到最大重试次数，终止流程")
                            return False
                    
                self.logger.info("=" * 50)
                self.logger.info("步骤3: 执行搜索和筛选")
                self.logger.info("=" * 50)
                if not self.search_and_filter(self.main_page):
                    self.logger.error("搜索失败，终止流程")
                    return False
                self.raise_if_cancelled("筛选完成")
                
                # NOTE: legacy comment removed for English-only repo.
                self.logger.info("步骤4: 确保页面稳定，点击空白处...")
                self.main_page.locator('body').click(position={'x': 10, 'y': 10}, force=True)
                self.delay(1)
                
                # NOTE: legacy comment removed for English-only repo.
                self.creator_counter = 0  # NOTE: legacy comment removed for English-only repo.
                all_loaded_creators = []  # NOTE: legacy comment removed for English-only repo.
                self.new_processed_count = 0  # NOTE: legacy comment removed for English-only repo.
                max_creators_to_load = 400  # NOTE: legacy comment removed for English-only repo.
                target_load_count = min(max_creators_to_load, self.max_creators * 2)  # NOTE: legacy comment removed for English-only repo.

                # NOTE: legacy comment removed for English-only repo.
                self.logger.info("=" * 50)
                self.logger.info(f"步骤5: 开始滚动加载达人，目标加载 {target_load_count} 个")
                self.logger.info("=" * 50)
                
                scroll_attempts = 0
                max_scroll_attempts = 50  # NOTE: legacy comment removed for English-only repo.
                last_count = 0
                no_new_content_count = 0  # NOTE: legacy comment removed for English-only repo.

                while len(all_loaded_creators) < target_load_count and scroll_attempts < max_scroll_attempts:
                    self.raise_if_cancelled("滚动加载达人")
                    scroll_attempts += 1
                    
                    # NOTE: legacy comment removed for English-only repo.
                    current_creators = self.get_creators_in_current_page(self.main_page)

                    # NOTE: legacy comment removed for English-only repo.
                    new_creators = []
                    for creator in current_creators:
                        if creator not in all_loaded_creators:
                            new_creators.append(creator)
                            all_loaded_creators.append(creator)

                    current_count = len(all_loaded_creators)
                    self.logger.info(f"滚动第 {scroll_attempts} 次，本次新增 {len(new_creators)} 个，累计已加载 {current_count} 个达人")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    if current_count >= target_load_count:
                        self.logger.info(f"已达到目标加载数量 {target_load_count}，停止滚动")
                        break
                    
                    # NOTE: legacy comment removed for English-only repo.
                    if current_count == last_count:
                        no_new_content_count += 1
                        self.logger.info(f"第 {no_new_content_count} 次无新增内容")
                        if no_new_content_count >= 3:
                            self.logger.info("连续3次滚动无新内容，确认已到达底部")
                            break
                    else:
                        no_new_content_count = 0
                        last_count = current_count
                    
                    # NOTE: legacy comment removed for English-only repo.
                    if not self.scroll_page_for_more(self.main_page):
                        self.logger.warning("滚动失败，停止加载")
                        break
                    
                    # NOTE: legacy comment removed for English-only repo.
                    self.delay(1)

                self.logger.info(f"滚动加载完成，共加载 {len(all_loaded_creators)} 个达人")

                # NOTE: legacy comment removed for English-only repo.
                self.logger.info("=" * 50)
                self.logger.info("步骤6: 滚动回页面顶部")
                self.logger.info("=" * 50)
                self.main_page.evaluate("window.scrollTo(0, 0)")
                self.delay(1)

                # NOTE: legacy comment removed for English-only repo.
                self.logger.info("=" * 50)
                self.logger.info(f"步骤7: 开始处理达人，目标新增 {self.target_new_count} 个")
                self.logger.info("=" * 50)

                total_creators = len(all_loaded_creators)
                scanned_creators = 0

                for idx, creator_name in enumerate(all_loaded_creators):
                    self.raise_if_cancelled("处理达人列表")
                    scanned_creators += 1
                    # NOTE: legacy comment removed for English-only repo.
                    if self.new_processed_count >= self.target_new_count:
                        self.logger.info(f"已新增 {self.new_processed_count} 个达人，达到目标")
                        break

                    if self.shared_seen_creators is not None:
                        if creator_name in self.shared_seen_creators:
                            self.logger.info(f"[{idx+1}/{len(all_loaded_creators)}] {creator_name} 在当前任务的先前批次已处理或跳过，直接跳过")
                            continue
                        self.shared_seen_creators.add(creator_name)

                    if creator_name in self.blacklist_creators:
                        self.logger.info(f"[{idx+1}/{len(all_loaded_creators)}] {creator_name} 在黑名单中，跳过")
                        if self._record_skip(creator_name, "黑名单"):
                            break
                        continue

                    # NOTE: legacy comment removed for English-only repo.
                    records = self.existing_data.get(creator_name, [])
                    if records:
                        # NOTE: legacy comment removed for English-only repo.
                        all_connected = all(r['connect'] for r in records)
                        if all_connected and self.brand_name in {r['brand_name'] for r in records}:
                            self.logger.info(f"[{idx+1}/{len(all_loaded_creators)}] {creator_name} 已完整处理（当前品牌已建联），跳过")
                            if self._record_skip(creator_name, "品牌已建联"):
                                break
                            continue

                    # NOTE: legacy comment removed for English-only repo.
                    if creator_name in self.processed_creators:
                        self.logger.info(f"达人 {creator_name} 已处理过，跳过")
                        if self._record_skip(creator_name, "重复达人"):
                            break
                        continue
                    
                    self.logger.info(f"处理第 {self.new_processed_count + 1}/{self.target_new_count} 个新达人: {creator_name} (列表中第 {idx + 1}/{len(all_loaded_creators)} 个)")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    try:
                        # NOTE: legacy comment removed for English-only repo.
                        self.main_page.evaluate(f"""
                            () => {{
                                const elements = document.querySelectorAll('span[data-e2e="fbc99397-6043-1b37"], span.text-body-m-medium.text-neutral-text1');
                                for (const el of elements) {{
                                    if (el.textContent?.trim() === '{creator_name}') {{
                                        el.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
                                        return true;
                                    }}
                                }}
                                return false;
                            }}
                        """)
                        self.delay(1)
                    except Exception as e:
                        self.logger.warning(f"滚动到达人 {creator_name} 位置失败: {e}")
                    
                    # NOTE: legacy comment removed for English-only repo.
                    creator_data = self.process_single_creator(self.main_page, creator_name)
                    if creator_data:
                        self.save_single_creator(creator_data)
                        self.processed_creators.add(creator_name)
                        
                        # NOTE: legacy comment removed for English-only repo.
                        self.new_processed_count += 1  # NOTE: legacy comment removed for English-only repo.
                        self.logger.info(f"✓ 新增进度：{self.new_processed_count}/{self.target_new_count}（{creator_name} 已保存）")
                        if self._should_restart_after_new():
                            self.logger.info("达到单批新增阈值，结束本批以便重启浏览器")
                            break
                        
                        # NOTE: legacy comment removed for English-only repo.
                        if self.new_processed_count % 10 == 0:
                            self.cleanup_stale_pages()
                            self.logger.info(f"已新增处理 {self.new_processed_count} 个达人，活跃页面数: {len(self.active_pages)}")
                    else:
                        self.logger.warning(f"✗ 处理达人 {creator_name} 失败，跳过")
                        if self._record_skip(creator_name, "处理失败"):
                            break

                self.logger.info("=" * 50)
                self.logger.info(f"爬取完成！")
                self.logger.info(f"  - 本次新增处理: {self.new_processed_count} 个达人")
                self.logger.info(f"  - 目标新增: {self.target_new_count} 个")
                skipped_processed = len([c for c in all_loaded_creators[:scanned_creators] if c in self.existing_data])
                self.logger.info(f"  - 跳过已处理: {skipped_processed} 个")
                if self.skip_restart_threshold > 0:
                    self.logger.info(f"  - 累计跳过: {self.skipped_total}/{self.skip_restart_threshold} 个达人")
                if self.restart_requested:
                    self.logger.info(f"  - 触发新会话请求：{self.restart_reason}")
                self.logger.info(f"  - 实际扫描: {scanned_creators}/{total_creators} 个")
                self.logger.info("=" * 50)
                return True
                
            except CrawlerCancelledError:
                self.logger.info("收到取消信号，任务提前结束")
                return False
            except Exception as e:
                self.logger.error(f"程序运行出错: {e}")
                return False
            finally:
                # NOTE: legacy comment removed for English-only repo.
                self.logger.info("Playwright上下文即将关闭，执行安全清理...")
                self.safe_cleanup()
                # NOTE: legacy comment removed for English-only repo.
                self.playwright_context_active = False
                # NOTE: legacy comment removed for English-only repo.
                self.is_cleaned_up = True
                if self._timeout_timer:
                    self._timeout_timer.cancel()
                    self._timeout_timer = None
                self.run_started_at = None

if __name__ == '__main__':
    search_strategy = None
    cli_max = None
    
    # NOTE: legacy comment removed for English-only repo.
    script_dir = Path(__file__).parent
    default_file = script_dir / "dify_out.txt"
    
    # NOTE: legacy comment removed for English-only repo.
    if len(sys.argv) >= 2:
        arg1 = sys.argv[1]
        
        if arg1.isdigit():
            # NOTE: legacy comment removed for English-only repo.
            cli_max = int(arg1)
            if default_file.exists():
                search_strategy = CreatorFullCrawler.load_search_strategy_from_file(default_file)
                if search_strategy:
                    logging.info("从默认文件 dify_out.txt 加载搜索策略成功")
        else:
            # NOTE: legacy comment removed for English-only repo.
            strategy_file = Path(arg1)
            if strategy_file.exists():
                search_strategy = CreatorFullCrawler.load_search_strategy_from_file(strategy_file)
                if search_strategy:
                    logging.info(f"从文件 {strategy_file} 加载搜索策略成功")
            else:
                logging.warning(f"策略文件 {strategy_file} 不存在，尝试使用默认文件")
                if default_file.exists():
                    search_strategy = CreatorFullCrawler.load_search_strategy_from_file(default_file)
                    if search_strategy:
                        logging.info("从默认文件 dify_out.txt 加载搜索策略成功")
        
        # NOTE: legacy comment removed for English-only repo.
        if len(sys.argv) >= 3:
            try:
                cli_max = int(sys.argv[2])
            except ValueError:
                logging.warning(f"忽略无效的命令行上限参数：{sys.argv[2]}")
    else:
        # NOTE: legacy comment removed for English-only repo.
        if default_file.exists():
            search_strategy = CreatorFullCrawler.load_search_strategy_from_file(default_file)
            if search_strategy:
                logging.info("从默认文件 dify_out.txt 加载搜索策略成功")
    
    # NOTE: legacy comment removed for English-only repo.
    crawler = CreatorFullCrawler(search_strategy=search_strategy)
    
    try:
        success = crawler.run(max_creators=cli_max)
        if success:
            logging.info("爬虫执行成功")
        else:
            logging.error("爬虫执行失败")
    except KeyboardInterrupt:
        logging.info("用户中断程序")
    except Exception as e:
        logging.error(f"程序异常: {e}")
    finally:
        if not crawler.is_cleaned_up:
            logging.info("执行最终清理")
            crawler.emergency_cleanup()
        logging.info("程序结束")
