"""
TikTok Chat Send and Product Card
TikTok发送聊天信息与商品卡
nohup python -u product_card/chat_card.py --show-browser --skip-verify --manual-login-timeout 900 > logs/chat_card_manual1106.log 2>&1 &
nohup python -u product_card/chat_card.py --skip-verify > logs/chat_card_auto1105.log 2>&1 &
nohup python -u product_card/chat_card.py --show-browser --skip-verify > logs/chat_card_manual1108.log 2>&1 &
nohup python -u product_card/chat_card.py \
  --region FR \
  --account-name FR1 \
  --show-browser \
  --skip-verify \
  > logs/chat_card_FR1.log 2>&1 &
nohup python -u product_card/chat_card.py \
  --region MX \
  --show-browser \
  --skip-verify \
  > logs/chat_card_1110.log 2>&1 &
"""

import argparse
import json
import logging
import time
import sys
import os
from urllib.parse import quote
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from filelock import FileLock

# Ensure imports work when executed as a script
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models.email_code import GmailVerificationCode
from utils.credentials import get_default_account_from_env, MissingDefaultAccountError

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CARD_JSON = BASE_DIR / "data" / "card" / "card_send_list.json"
DEFAULT_ACCOUNTS_JSON = BASE_DIR / "config" / "accounts.json"
MARKET_MAPPING = {
    "MX": "19",
    "FR": "17",
}
EU_DOMAIN_REGIONS = {"FR"}


def _load_accounts_from_config(config_path: Path) -> List[Dict[str, Any]]:
    """读取账号配置文件。"""
    config_path = config_path.expanduser().resolve()
    if not config_path.exists():
        logger.error("账号配置文件不存在: %s", config_path)
        return []
    try:
        with config_path.open("r", encoding="utf-8") as fp:
            data = json.load(fp)
    except Exception as exc:
        logger.error("读取账号配置文件失败: %s", exc)
        return []

    accounts = data.get("accounts", [])
    for idx, account in enumerate(accounts):
        account.setdefault("id", idx)
    return accounts


def select_account_for_region(
    region: str,
    account_name: Optional[str],
    config_path: Path,
) -> Optional[Dict[str, Any]]:
    """从配置文件选择指定区域的账号。"""
    accounts = _load_accounts_from_config(config_path)
    if not accounts:
        return None

    region_upper = (region or "").upper()
    candidates = [
        acct
        for acct in accounts
        if acct.get("enabled", True) and acct.get("region", "").upper() == region_upper
    ]

    if account_name:
        for acct in candidates:
            if acct.get("name") == account_name:
                logger.info(
                    "已选择账号 %s (%s) 用于区域 %s",
                    acct.get("name"),
                    acct.get("login_email"),
                    region_upper or "默认",
                )
                return acct
        logger.warning(
            "未在区域 %s 的账号中找到名称为 %s 的配置，改为使用该区域的第一个可用账号",
            region_upper,
            account_name,
        )

    if not candidates:
        logger.error("配置中没有区域 %s 的可用账号", region_upper or "(未知)")
        return None

    acct = candidates[0]
    logger.info(
        "自动选择账号 %s (%s) 用于区域 %s",
        acct.get("name"),
        acct.get("login_email"),
        region_upper or "默认",
    )
    return acct

class BrowserClosedError(Exception):
    """Raised when the current Playwright page/context/browser has been closed."""

class CardSender:
    """
    TikTok发送聊天信息与商品卡
    """
    
    def __init__(
        self,
        account_info=None,
        *,
        region: Optional[str] = None,
        headless: bool = True,
        manual_login_timeout: Optional[int] = None,
        verify_delivery: bool = True,
        card_data_path: Optional[Path] = None,
        card_data: Optional[List[Dict[str, Any]]] = None,
    ):
        # Accept caller-provided credentials or fall back to environment variables
        if account_info:
            self.login_email = account_info.get('login_email')
            self.login_password = account_info.get('login_password')
            self.gmail_username = account_info.get('gmail_username')
            self.gmail_app_password = account_info.get('gmail_app_password')
            self.account_name = account_info.get('name', 'Unknown')
            self.account_id = account_info.get('id', -1)
            logger.info(f"Using account: {self.account_name} ({self.login_email})")
        else:
            try:
                fallback = get_default_account_from_env()
            except MissingDefaultAccountError as exc:
                raise MissingDefaultAccountError(
                    "Provide account_info or configure DEFAULT_* environment variables."
                ) from exc

            self.login_email = fallback['login_email']
            self.login_password = fallback['login_password']
            self.gmail_username = fallback['gmail_username']
            self.gmail_app_password = fallback['gmail_app_password']
            self.account_name = fallback.get('name', 'Default Account')
            self.account_id = fallback.get('id', -1)
            logger.warning("Account info not provided; using fallback credentials loaded from the environment.")

        desired_region = (region or "").upper()
        account_region = (
            (account_info.get("region") or "").upper()
            if account_info and account_info.get("region")
            else ""
        )
        if account_region:
            if desired_region and desired_region != account_region:
                logger.warning(
                    "传入 region=%s 但账号所属区域为 %s，已按账号区域执行。",
                    desired_region,
                    account_region,
                )
            self.region = account_region
        elif desired_region:
            self.region = desired_region
        else:
            self.region = "MX"

        self.market_id = MARKET_MAPPING.get(self.region, MARKET_MAPPING.get("MX", "19"))
        self.base_domain = self._resolve_base_domain()
        logger.info(
            "当前运行区域: %s | 域名: %s | Market ID: %s",
            self.region,
            self.base_domain,
            self.market_id,
        )

        self.headless = headless
        if account_info and account_info.get('headless') is not None:
            raw_headless = account_info.get('headless')
            if isinstance(raw_headless, str):
                self.headless = raw_headless.strip().lower() not in ("false", "0", "no", "off")
            else:
                self.headless = bool(raw_headless)
        manual_timeout = manual_login_timeout
        self.verify_delivery = verify_delivery
        self.card_data_override = card_data
        self.card_data_path = (
            Path(card_data_path).expanduser().resolve()
            if card_data_path
            else DEFAULT_CARD_JSON
        )

        if account_info and account_info.get('verify_delivery') is not None:
            raw_verify = account_info.get('verify_delivery')
            if isinstance(raw_verify, str):
                self.verify_delivery = raw_verify.strip().lower() not in ("false", "0", "no", "off")
            else:
                self.verify_delivery = bool(raw_verify)

        self.manual_login_mode = (not self.headless) and (manual_timeout is not None)
        if self.manual_login_mode:
            if manual_timeout and manual_timeout > 0:
                self.manual_login_timeout = manual_timeout
            else:
                self.manual_login_timeout = 600
            logger.info("浏览器模式: 可视模式（需手动登录）")
            logger.info(f"手动登录超时时间: {self.manual_login_timeout} 秒")
        else:
            self.manual_login_timeout = manual_timeout or 0
            mode_desc = "Headless 模式" if self.headless else "可视模式（自动登录）"
            logger.info(f"浏览器模式: {mode_desc}")
        logger.info("发送后验证聊天记录: %s", "启用" if self.verify_delivery else "跳过")

        # Gmail验证码配置
        self.gmail_verifier = GmailVerificationCode(
            username=self.gmail_username,
            app_password=self.gmail_app_password
        )
        
        self.browser = None
        self.context = None
        self.page = None
        self._playwright = None
        self.max_browser_restarts = 10
        self.closed_error_threshold = 5
        self.restart_attempts = 0
        self.consecutive_browser_closed_errors = 0
        self._restart_anchor_index: Optional[int] = None
        self.sent_history: Dict[str, Dict[str, str]] = {}
        self.screenshot_dir = BASE_DIR / 'data' / 'screenshot' / 'product_card'
        self.screenshot_dir.mkdir(parents=True, exist_ok=True)
        self.auto_restart_interval = int(os.getenv("CHAT_CARD_RESTART_INTERVAL", "60"))
        if self.auto_restart_interval <= 0:
            self.auto_restart_interval = 0
        self.creators_since_restart = 0
    
    def delay(self, seconds: float):
        """延迟"""
        time.sleep(seconds)

    def _resolve_base_domain(self) -> str:
        if self.region in EU_DOMAIN_REGIONS:
            return "partner.eu.tiktokshop.com"
        return "partner.tiktokshop.com"

    def _build_login_url(self) -> str:
        redirect = quote(f"//{self.base_domain}/home", safe="")
        return (
            "https://partner-sso.tiktok.com/account/login"
            f"?from=ttspc_logout&redirectURL={redirect}"
            "&lang=en&local_id=localID_Portal_88574979_1758691471679"
            "&userID=51267627&is_through_login=1"
        )

    def _build_chat_url(self, creator_id: str) -> str:
        return (
            f"https://{self.base_domain}/partner/im"
            f"?creator_id={creator_id}&market={self.market_id}&enter_from=find_creator_detail"
        )
    
    def load_card_send_data(self) -> List[Dict[str, Any]]:
        """从JSON文件加载发送卡片的数据（支持单个或列表）"""
        if self.card_data_override is not None:
            logger.info("使用传入的卡片配置数据，共 %s 条", len(self.card_data_override))
            return self.card_data_override

        json_path = Path(self.card_data_path)
        logger.info(f"读取配置文件: {json_path}")

        try:
            with json_path.open('r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, dict):
                data = [data]

            logger.info(f"成功加载 {len(data)} 个达人配置")
            return data
        except Exception as e:
            logger.error(f"读取JSON文件失败: {e}")
            return []

    def _sanitize_filename(self, value: str) -> str:
        """移除文件名中不安全的字符。"""
        return "".join(ch if ch.isalnum() or ch in ('-', '_') else "_" for ch in value)

    def _save_screenshot(self, page, prefix: str) -> Optional[str]:
        """保存当前页面截图。"""
        try:
            safe_prefix = self._sanitize_filename(prefix)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path = self.screenshot_dir / f"{safe_prefix}_{timestamp}.png"
            page.screenshot(path=str(file_path), full_page=True)
            logger.info(f"已保存截图: {file_path}")
            return str(file_path)
        except Exception as exc:
            logger.warning(f"保存截图失败: {exc}")
            return None

    def _load_sent_history(self) -> None:
        """预加载历史发送记录，按 creator_id -> product_id 映射."""
        excel_path = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / 'data' / 'card' / 'product_send.xlsx'
        if not excel_path.exists():
            self.sent_history = {}
            return
        try:
            df = pd.read_excel(excel_path)
            required = {'creator_id', 'product_id'}
            if not required.issubset(df.columns):
                self.sent_history = {}
                return
            history: Dict[str, Dict[str, str]] = {}
            for _, row in df.iterrows():
                c_id = str(row.get('creator_id')).strip()
                p_id = str(row.get('product_id')).strip()
                if not c_id or not p_id:
                    continue
                raw_time = row.get('send_card_time')
                send_time = '' if pd.isna(raw_time) else str(raw_time).strip()
                history.setdefault(c_id, {})[p_id] = send_time
            self.sent_history = history
            logger.info(f"已加载历史发送记录映射，共 {len(self.sent_history)} 个达人")
        except Exception as exc:
            logger.warning(f"加载历史发送记录失败: {exc}")
            self.sent_history = {}

    def _should_skip_creator(self, card_data: Dict[str, Any]) -> bool:
        """根据历史记录判断是否跳过当前达人."""
        creator_id = str(card_data.get('creator_id') or '').strip()
        if not creator_id:
            return False
        product_ids = card_data.get('product_ids', [])
        if not product_ids:
            return False
        history = self.sent_history.get(creator_id, {})
        if not history:
            return False
        # 仅当所有产品都有发送记录并且记录的发送状态非空时跳过
        all_completed = True
        for pid in product_ids:
            pid_str = str(pid).strip()
            send_time = history.get(pid_str, '')
            if not send_time:
                all_completed = False
                break
        if all_completed:
            logger.info(
                "达人 %s 的所有产品卡在历史记录中已发送，跳过本次处理",
                card_data.get('creator_name') or creator_id,
            )
        return all_completed

    @staticmethod
    def _is_browser_closed_error(error: Exception) -> bool:
        message = str(error)
        return "Target page, context or browser has been closed" in message or "Browser has been closed" in message

    def _perform_login(self) -> bool:
        """执行登录流程（带重试）。"""
        max_login_attempts = 1 if not self.headless else 3
        for attempt in range(1, max_login_attempts + 1):
            logger.info(f"=== 登录流程 {attempt}/{max_login_attempts} ===")
            login_success = self.login(self.page)
            if not login_success:
                logger.error("登录失败")
            else:
                if not self.headless:
                    logger.info("检测到手动登录成功")
                    return True
                if self.check_welcome_page(self.page):
                    logger.info("登录成功")
                    return True
            if attempt < max_login_attempts:
                logger.info("准备重试登录...")
        logger.error("登录多次失败，终止任务")
        return False

    def _setup_browser(self, playwright, is_restart: bool = False) -> bool:
        """启动浏览器并完成登录。"""
        try:
            self.browser = playwright.chromium.launch(headless=self.headless, timeout=60000)
            self.context = self.browser.new_context(viewport={'width': 1920, 'height': 1080})
            self.context.set_default_timeout(60000)
            self.page = self.context.new_page()
        except Exception as e:
            logger.error(f"{'重启' if is_restart else '初始化'}浏览器失败: {e}")
            return False

        if not self._perform_login():
            self._teardown_browser()
            return False
        return True

    def _teardown_browser(self):
        """关闭当前浏览器资源。"""
        try:
            if self.page:
                self.page.close()
        except Exception:
            pass
        finally:
            self.page = None

        try:
            if self.context:
                self.context.close()
        except Exception:
            pass
        finally:
            self.context = None

        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass
        finally:
            self.browser = None

    def _restart_browser_session(self) -> bool:
        """重启浏览器，会重新登录。"""
        if self.restart_attempts >= self.max_browser_restarts:
            logger.error("达到最大浏览器重启次数限制，终止任务")
            return False

        self.restart_attempts += 1
        logger.warning(f"连续检测到浏览器关闭异常，准备第 {self.restart_attempts}/{self.max_browser_restarts} 次重启浏览器...")

        self._teardown_browser()

        if not self._setup_browser(self._playwright, is_restart=True):
            logger.error("浏览器重启失败")
            return False

        self.consecutive_browser_closed_errors = 0
        logger.info("浏览器重启成功，重新开始任务")
        return True

    def login(self, page) -> bool:
        """登录TikTok商家平台。Headless=False时等待人工登录。"""
        login_url = self._build_login_url()

        if self.manual_login_mode:
            logger.info("Headless 关闭，将打开浏览器供人工登录。")
            try:
                page.goto(login_url, wait_until="networkidle", timeout=60000)
            except Exception as e:
                logger.error(f"打开登录页失败: {e}")
                return False

            logger.info("请在弹出的浏览器窗口中手动完成登录。")
            logger.info("登录成功后，应跳转到 TikTok Shop Partner Center 欢迎页。")
            timeout_ms = max(5, self.manual_login_timeout) * 1000
            try:
                if self.check_welcome_page(page, timeout_ms=timeout_ms, delay_seconds=0):
                    logger.info("检测到欢迎页面，手动登录成功")
                    return True
                logger.error("等待人工登录超时或未检测到欢迎页面")
                return False
            except Exception as exc:
                logger.error(f"等待人工登录时出错: {exc}")
                return False

        if not self.headless and not self.manual_login_mode:
            logger.info("Headless 关闭但未启用手动登录：将显示浏览器窗口并自动执行登录流程。")

        max_retries = 5
        for i in range(max_retries):
            logger.info(f"开始登录流程（第 {i + 1} 次尝试）...")
            try:
                page.goto(login_url, wait_until="networkidle", timeout=60000)

                logger.info("进入邮件登录模式...")
                email_login_btn = page.get_by_text("Log in with code").first
                email_login_btn.click()

                logger.info(f"输入邮箱: {self.login_email}")
                page.fill('#email input', self.login_email)

                try:
                    send_code_btn = page.locator('div[starling-key="profile_edit_userinfo_send_code"]').first
                    send_code_btn.wait_for(state="visible", timeout=5000)
                    send_code_btn.click()
                    logger.info("已点击 Send code 按钮")
                except Exception as e:
                    logger.error(f"点击 Send code 失败: {e}")

                logger.info("正在从Gmail获取验证码...")
                verification_code = self.gmail_verifier.get_verification_code()
                if not verification_code:
                    logger.error("验证码获取失败")
                    continue

                logger.info(f"成功获取验证码: {verification_code}")
                page.fill('#emailCode_input', verification_code)

                login_btn = page.locator('button[starling-key="account_login_btn_loginform_login_text"]').first
                login_btn.click()

                if self.base_domain in page.url:
                    logger.info("登录成功，等待欢迎页面加载...")
                    if self.check_welcome_page(page, timeout_ms=20000, delay_seconds=0):
                        logger.info("检测到欢迎页面，自动登录成功")
                        return True
                    logger.warning("登录后未检测到欢迎页面，将重试登录流程")
                    time.sleep(5)
                    continue
                else:
                    logger.error("登录失败，未跳转到正确页面")
                    time.sleep(3)
                    continue

            except Exception as e:
                logger.error(f"登录过程出错：{e}")
                if i < max_retries - 1:
                    logger.warning(f"第 {i + 1} 次尝试失败，等待 15 秒后重试...")
                    time.sleep(15)

        logger.error(f"登录失败，已达到最大重试次数 {max_retries}。程序终止。")
        return False
    
    def check_welcome_page(self, page, timeout_ms: int = 20000, delay_seconds: float = 3) -> bool:
        """检查是否到达欢迎页面"""
        logger.info("检查欢迎页面...")
        if delay_seconds > 0:
            time.sleep(delay_seconds)

        selectors = [
            'text=Welcome to TikTok Shop Partner Center',
            'text=欢迎来到 TikTok Shop 合作伙伴中心',
            'text=Account GMV trend',
            'text=View your data and facilitate seller authorizations',
            'text=Hi'
            
        ]

        for idx, selector in enumerate(selectors):
            current_timeout = timeout_ms if idx == 0 else min(15000, timeout_ms)
            try:
                page.wait_for_selector(selector, timeout=current_timeout)
                logger.info("找到欢迎页面提示文字")
                return True
            except Exception:
                continue

        try:
            page.wait_for_url(
                f"**{self.base_domain}/home**",
                timeout=min(5000, timeout_ms),
            )
            logger.info("检测到 TikTok Shop Partner Center 主页 URL")
            return True
        except Exception:
            pass

        if self.base_domain in (page.url or ""):
            logger.info("当前页面已跳转至 TikTok Shop Partner Center")
            return True

        logger.warning("未找到欢迎页面提示")
        return False

    def navigate_to_campaign_page(self, page, creator_id: str) -> bool:
        """跳转到活动页面（直接等待固定时间，不再检查文案）"""
        logger.info("跳转到活动页面...")

        campaign_url = f"https://{self.base_domain}/affiliate-campaign/campaign?"
        try:
            page.goto(campaign_url, wait_until="networkidle", timeout=30000)
        except Exception as exc:
            logger.error(f"跳转活动页面失败：{exc}")
            if self._is_browser_closed_error(exc):
                raise BrowserClosedError(str(exc))
            return False

        self.delay(8)
        logger.info("活动页面跳转完成，已固定等待 8 秒")
        return True

    def _wait_chat_ready(self, page, timeout_ms=60000) -> bool:
        """等待聊天页面的输入区域或"Send message"字样可见"""
        logger.info(f"等待聊天输入框或 'Send message' 字样（<= {timeout_ms/1000:.0f}s）...")
        start = time.time()
        check_interval = 1.5

        def found_in_page(p) -> bool:
            selectors = [
                'textarea[placeholder*="Send a message"]',
                'textarea[placeholder*="发送消息"]',
                'textarea[placeholder*="message" i]',
                '#im_sdk_chat_input textarea',
                'textarea'
            ]
            for sel in selectors:
                try:
                    el = p.locator(sel).first
                    if el.count() > 0 and el.is_visible():
                        logger.info(f"发现输入框：{sel}")
                        return True
                except Exception:
                    pass

            try:
                btn = p.get_by_text("Send message", exact=False)
                if btn and btn.count() > 0 and btn.first.is_visible():
                    logger.info("发现 'Send message' 文案")
                    return True
            except Exception:
                pass

            return False

        while (time.time() - start) * 1000 < timeout_ms:
            try:
                if found_in_page(page):
                    return True
            except Exception:
                pass

            try:
                for fr in page.frames:
                    try:
                        if found_in_page(fr):
                            logger.info("在 iframe 中找到输入框/文案")
                            return True
                    except Exception:
                        continue
            except Exception:
                pass

            time.sleep(check_interval)

        logger.warning("等待聊天区域超时")
        return False

    def open_chat_and_screenshot(self, page, creator_id: str,
                                max_retries: int = 3) -> bool:
        """跳转到聊天页面，等待加载完成"""
        if not creator_id:
            logger.error("缺少 creator_id")
            return False

        chat_url = self._build_chat_url(creator_id)
        logger.info(f"准备跳转聊天页面: {chat_url}")

        for attempt in range(1, max_retries + 1):
            logger.info(f"加载聊天页面 - 尝试 {attempt}/{max_retries}")
            try:
                page.goto(chat_url, wait_until="networkidle", timeout=60000)
            except Exception as e:
                logger.error(f"goto 失败: {e}")
                if self._is_browser_closed_error(e):
                    raise BrowserClosedError(str(e))
                if attempt < max_retries:
                    time.sleep(2)
                    continue
                return False

            self.delay(2)

            current_url = page.url
            if "login" in current_url.lower() or "signin" in current_url.lower():
                logger.error("被重定向回登录页，需先确保登录状态")
                return False

            if self._wait_chat_ready(page, timeout_ms=60000):
                logger.info(f"聊天页面加载完成")
                return True

            if attempt < max_retries:
                logger.warning("未等到聊天输入框，准备重试...")
                try:
                    page.reload(wait_until="networkidle", timeout=40000)
                except Exception:
                    pass
                self.delay(2)

        logger.error("多次尝试后仍未成功加载聊天页面")
        return False

    def analyze_chat_status(self, page) -> Dict[str, Any]:
        """分析聊天历史，返回商家发送数量与达人是否回复等信息"""
        status = {
            "merchant_messages": 0,
            "creator_replied": False,
            "messages_text": [],
        }

        try:
            page.wait_for_load_state("networkidle")
            self.delay(2)

            chat_data = page.evaluate(
                """
                () => {
                    const selectors = [
                        'div.index-module__messageList--GBz6X',
                        'div.messageList-k_OG24',
                        'div.chatd-scrollView',
                        'div[class*="messageList"]',
                        'div[class*="chatd-scrollView"]'
                    ];

                    let container = null;
                    for (const selector of selectors) {
                        const node = document.querySelector(selector);
                        if (node && node.offsetParent !== null) {
                            container = node;
                            break;
                        }
                    }

                    if (!container) {
                        return [];
                    }

                    const nodes = container.querySelectorAll('div.chatd-message');
                    return Array.from(nodes).map(node => {
                        const contentEl = node.querySelector('pre.index-module__content--QKRoB');
                        const text = contentEl ? contentEl.textContent.trim() : '';
                        const className = node.className || '';
                        return {
                            content: text,
                            isMerchant: className.includes('chatd-message--right'),
                            isCreator: className.includes('chatd-message--left'),
                        };
                    });
                }
                """
            )

            if not isinstance(chat_data, list):
                chat_data = []

            status["messages_text"] = [
                msg.get("content", "") for msg in chat_data if msg.get("content")
            ]
            status["merchant_messages"] = sum(
                1 for msg in chat_data if msg.get("isMerchant")
            )
            status["creator_replied"] = any(
                msg.get("isCreator") for msg in chat_data
            )

            logger.info(
                "聊天历史统计 - 商家消息数: %s, 达人是否回复: %s",
                status["merchant_messages"],
                status["creator_replied"],
            )
        except Exception as exc:
            logger.warning(f"分析聊天历史失败: {exc}")

        return status

    def check_and_send_message(
        self,
        page,
        message: str,
        limit_state: Optional[Dict[str, Any]] = None,
        chat_status: Optional[Dict[str, Any]] = None,
    ) -> bool:
        """检查历史记录并发送消息，尊重未回复达人的 5 条限制"""
        if not message or not message.strip():
            logger.info("没有消息内容，跳过")
            return False

        message = message.strip()

        if limit_state and not limit_state.get("creator_replied", False):
            remaining = limit_state.get("remaining", 0) or 0
            if remaining <= 0:
                logger.info("达人尚未回复，且已达到 5 条消息上限，本次不再发送文本消息")
                return False

        logger.info("检查消息历史并准备发送文本消息...")

        try:
            if chat_status is None:
                chat_status = self.analyze_chat_status(page)

            history_messages = chat_status.get("messages_text", []) if chat_status else []

            if message in history_messages:
                logger.info("消息已存在于聊天历史中，跳过发送")
                return False

            # 填充消息
            selectors = [
                'textarea#imTextarea',
                'textarea[placeholder="Send a message"]',
                'textarea[placeholder*="Send a message"]',
                'textarea[placeholder*="message" i]',
                '#im_sdk_chat_input textarea',
                'div[data-e2e="cda68c25-5112-89c2"] textarea',
                'textarea.index-module__textarea--qYh62',
                'textarea'
            ]

            for sel in selectors:
                try:
                    el = page.locator(sel).first
                    if el.count() > 0:
                        el.wait_for(state="visible", timeout=5000)
                        el.click()
                        self.delay(0.5)

                        el.fill(message)
                        logger.info("消息已填入输入框")

                        self.delay(0.5)
                        el.press("Enter")
                        logger.info("消息已发送")

                        if chat_status is not None:
                            chat_status.setdefault("messages_text", []).append(message)

                        if limit_state and not limit_state.get("creator_replied", False):
                            remaining = limit_state.get("remaining")
                            if isinstance(remaining, int):
                                limit_state["remaining"] = max(0, remaining - 1)
                                logger.info(
                                    "达人未回复，剩余可发送消息数：%s/5",
                                    limit_state["remaining"],
                                )

                        return True
                except Exception as e:
                    logger.debug(f"尝试选择器 {sel} 失败：{e}")

            logger.warning("未能填入消息")
            return False

        except Exception as e:
            logger.error(f"检查/发送消息失败: {e}")
            return False
    
    def search_and_open_campaign(self, page, campaign_id: str) -> bool:
        """搜索并打开指定的活动"""
        logger.info(f"搜索活动ID: {campaign_id}")
        
        try:
            # 定位搜索框并输入
            search_input = page.locator('input[data-tid="m4b_input_search"]').first
            search_input.wait_for(state="visible", timeout=10000)
            search_input.click()
            self.delay(0.5)
            
            search_input.fill(campaign_id)
            logger.info(f"已输入活动ID: {campaign_id}")
            
            # 按回车搜索
            search_input.press("Enter")
            logger.info("已按回车搜索")
            self.delay(3)
            
            # 点击 View details
            view_details_link = page.locator('span.arco-link[data-e2e="7924f760-0b3e-22f5"]:has-text("View details")').first
            view_details_link.wait_for(state="visible", timeout=10000)
            view_details_link.click()
            logger.info("已点击 View details")
            self.delay(3)
            
            # 点击 Approved 标签
            approved_tab = page.locator('div.arco-tabs-header-title[role="tab"]:has-text("Approved")').first
            approved_tab.wait_for(state="visible", timeout=10000)
            approved_tab.click()
            logger.info("已点击 Approved 标签")
            self.delay(3)
            
            return True
            
        except Exception as e:
            logger.error(f"搜索打开活动失败: {e}")
            if self._is_browser_closed_error(e):
                raise BrowserClosedError(str(e))
            return False

    def send_product_cards(
        self,
        page,
        product_ids: List[str],
        rate: int,
        creator_name: str,
        creator_id: str,
        campaign_id: str,
        limit_state: Optional[Dict[str, Any]] = None,
    ) -> bool:
        """发送产品卡片到达人"""
        logger.info(f"准备发送 {len(product_ids)} 个产品卡片，佣金率: {rate}%")

        sent_entries = {
            (creator_key, product_id)
            for creator_key, product_map in self.sent_history.items()
            for product_id, send_time in product_map.items()
            if send_time
        }
        if sent_entries:
            logger.info(f"检测到历史发送记录，共 {len(sent_entries)} 条，已加载")
        else:
            logger.info("未发现已发送记录，全部产品将正常发送")
        
        first_product = True
        success_count = 0

        for idx, product_id in enumerate(product_ids, 1):
            logger.info(f"发送第 {idx}/{len(product_ids)} 个产品: {product_id}")
            product_id = str(product_id).strip()
            creator_key = str(creator_id).strip()
            history_for_creator = self.sent_history.get(creator_key, {})

            merchant_baseline = 0
            if limit_state:
                merchant_baseline = limit_state.get("merchant_messages", 0) or 0
                if not limit_state.get("creator_replied", False):
                    remaining = limit_state.get("remaining", 0) or 0
                    if remaining <= 0:
                        logger.info(
                            "达人尚未回复且已达到 5 条消息上限，剩余产品卡将全部跳过"
                        )
                        break

            if history_for_creator.get(product_id):
                logger.info(f"达人 {creator_name} 的产品 {product_id} 已存在于历史记录中，跳过发送")
                continue

            rate_value = rate
            if isinstance(rate_value, (list, tuple)):
                rate_value = rate_value[0] if rate_value else ""
            rate_str = str(rate_value)

            # 产品发送重试机制
            max_product_retries = 3
            product_sent_successfully = False
            
            for retry_attempt in range(1, max_product_retries + 1):
                if retry_attempt > 1:
                    logger.info(f"重试发送产品 {product_id} (尝试 {retry_attempt}/{max_product_retries})")
                    # 重试前等待一段时间
                    self.delay(3)
                    
                failure_for_creator = False
                try:
                    # 每次重试都需要重新搜索产品（不仅仅是第一个产品）
                    if first_product or retry_attempt > 1:
                        # 确保在产品列表页面
                        try:
                            # 先检查是否还在正确的页面
                            page.wait_for_selector('div.arco-select-view:has-text("Product name")', timeout=5000)
                        except:
                            logger.info("页面可能已经跳转，尝试重新导航到活动页面")
                            if not self.navigate_to_campaign_page(page, creator_id):
                                logger.error(f"无法返回活动页面")
                                failure_for_creator = True
                                continue
                            if not self.search_and_open_campaign(page, campaign_id):
                                logger.error(f"无法重新打开活动详情")
                                failure_for_creator = True
                                continue
                        
                        product_dropdown = page.locator('div.arco-select-view:has-text("Product name")').first
                        product_dropdown.wait_for(state="visible", timeout=10000)
                        product_dropdown.click()
                        self.delay(1)

                        product_id_option = page.locator('li.arco-select-option:has-text("Product ID")').first
                        product_id_option.wait_for(state="visible", timeout=10000)
                        product_id_option.click()
                        self.delay(1)

                        if retry_attempt == 1:
                            first_product = False

                    # 搜索产品
                    search_input = page.locator('input[placeholder="Search Product ID"]').first
                    search_input.wait_for(state="visible", timeout=10000)
                    search_input.clear()  # 清空之前的输入
                    self.delay(0.5)
                    search_input.fill(product_id)
                    search_input.press("Enter")
                    logger.info(f"已搜索产品ID: {product_id}")
                    self.delay(5)  # 增加等待时间，确保搜索结果加载

                    # 点击 Share 按钮
                    share_btn = page.locator('button:has-text("Share")').first
                    share_btn.wait_for(state="visible", timeout=10000) 
                    share_btn.click()
                    self.delay(3)  # 增加等待时间

                    # 输入佣金率
                    rate_input = page.locator('input[id="standardCommission_input"]').first
                    rate_input.wait_for(state="visible", timeout=10000)
                    rate_input.clear()  # 清空之前的输入
                    self.delay(0.5)
                    rate_input.fill(rate_str)
                    logger.info(f"已输入佣金率: {rate_str}%")
                    self.delay(2)  # 增加等待时间

                    # 点击 Send in chat
                    send_in_chat_btn = page.locator('button:has-text("Send in chat")').first
                    send_in_chat_btn.wait_for(state="visible", timeout=10000)

                    try:
                        with page.context.expect_page() as new_page_event:
                            send_in_chat_btn.click()
                            # self._save_screenshot(page, f"{creator_name}_{product_id}_send")
                    except Exception as e:
                        logger.error(f"点击 Send in chat 失败: {e}")
                        raise

                    new_page = new_page_event.value
                    new_page.wait_for_load_state("domcontentloaded")
                    logger.info("新页面已打开")
                    # self._save_screenshot(new_page, f"{creator_name}_{product_id}_share_modal")
                    
                    # 等待 Recent creators 加载
                    try:
                        # 等待Recent creators区域出现并且数字变为100
                        max_wait_time = 30  # 最多等待30秒
                        start_time = time.time()
                        creators_loaded = False
                        
                        while time.time() - start_time < max_wait_time:
                            try:
                                # 查找Recent creators旁边的数字
                                creators_count_element = new_page.locator('div[data-e2e="06d3991c-a09b-cb4e"]').first
                                if creators_count_element.count() > 0:
                                    count_text = creators_count_element.text_content()
                                    if count_text and count_text.strip() == "100":
                                        logger.info("Recent creators已加载完成，显示100个达人")
                                        creators_loaded = True
                                        break
                                    else:
                                        logger.debug(f"当前Recent creators数量: {count_text}")
                            except Exception:
                                pass
                            
                            time.sleep(0.5)
                        
                        if not creators_loaded:
                            raise Exception(f"等待超时: Recent creators未能在{max_wait_time}秒内加载到100")
                        
                        logger.info("确认进入产品分享页面，Recent creators已完全加载")
                        
                    except Exception as e:
                        logger.error(f"未找到产品分享页面标识或加载失败: {e}")
                        new_page.close()
                        failure_for_creator = True
                        
                        # 如果是重试的最后一次，记录失败
                        if retry_attempt == max_product_retries:
                            send_time = ''
                            record_entry = {
                                'creator_name': creator_name,
                                'creator_id': creator_id,
                                'campaign_id': campaign_id,
                                'product_id': product_id,
                                'rate': rate_str,
                                'send_card_time': '',
                            }
                            self._append_record_to_excel(record_entry)
                            self.sent_history.setdefault(creator_key, {})[product_id] = ''
                        continue

                    # 搜索并选择达人
                    search_tab = new_page.locator('div[role="tab"]:has-text("Search for a creator")').first
                    search_tab.wait_for(state="visible", timeout=10000)
                    search_tab.click()
                    self.delay(1)

                    creator_search = new_page.locator('input[placeholder="Search for creator"]').first
                    creator_search.wait_for(state="visible", timeout=10000)
                    creator_search.fill(creator_name)
                    logger.info(f"已输入达人名称: {creator_name}")
                    self.delay(1)
                    
                    # 触发搜索
                    try:
                        search_svg = new_page.locator('svg.alliance-icon.alliance-icon-Search').first
                        search_svg.wait_for(state="visible", timeout=5000)
                        try:
                            search_svg.scroll_into_view_if_needed()
                        except Exception:
                            pass
                        search_svg.click()
                        logger.info("已点击搜索 SVG 图标")
                    except Exception as e:
                        logger.warning(f"直接点击 SVG 失败，尝试点击父级: {e}")
                        try:
                            parent = new_page.locator('svg.alliance-icon.alliance-icon-Search').first.locator('xpath=..')
                            parent.wait_for(state="visible", timeout=3000)
                            parent.click()
                            logger.info("已点击 SVG 的父级元素触发搜索")
                        except Exception as e2:
                            logger.warning(f"点击父级也失败，回退为回车触发搜索: {e2}")
                            try:
                                creator_search.press("Enter")
                                logger.info("已按回车触发搜索（备用）")
                            except Exception:
                                logger.error("搜索触发失败")
                    
                    self.delay(3)  # 增加等待搜索结果的时间

                    # 选择达人
                    try:
                        creator_row = new_page.locator(
                            f'div[data-e2e="6cfbf330-2b43-36cc"]:has(span.text-primary-normal:has-text("{creator_name}"))'
                        ).first
                        creator_row.wait_for(state="visible", timeout=10000)
                        creator_row.click()
                        logger.info(f"已点击达人 {creator_name} 的搜索结果行")
                    except Exception as e:
                        logger.error(f"点击达人结果行失败: {e}")
                        new_page.close()
                        failure_for_creator = True
                        
                        if retry_attempt == max_product_retries:
                            send_time = ''
                            record_entry = {
                                'creator_name': creator_name,
                                'creator_id': creator_id,
                                'campaign_id': campaign_id,
                                'product_id': product_id,
                                'rate': rate_str,
                                'send_card_time': '',
                            }
                            self._append_record_to_excel(record_entry)
                            self.sent_history.setdefault(creator_key, {})[product_id] = ''
                        continue

                    self.delay(2)

                    # 最终发送
                    try:
                        final_share_btn = new_page.locator(
                            'button.arco-btn.arco-btn-primary.arco-btn-size-large.arco-btn-shape-square.m4b-button.m4b-button-size-large'
                        ).first
                        final_share_btn.wait_for(state="visible", timeout=10000)
                        final_share_btn.click()
                        logger.info(f"产品 {product_id} 已发送")
                        # self._save_screenshot(new_page, f"{creator_name}_{product_id}_sent")
                    except Exception as e:
                        logger.error(f"点击 Share 按钮失败: {e}")
                        new_page.close()
                        failure_for_creator = True
                        
                        if retry_attempt == max_product_retries:
                            send_time = ''
                            record_entry = {
                                'creator_name': creator_name,
                                'creator_id': creator_id,
                                'campaign_id': campaign_id,
                                'product_id': product_id,
                                'rate': rate_str,
                                'send_card_time': '',
                            }
                            self._append_record_to_excel(record_entry)
                            self.sent_history.setdefault(creator_key, {})[product_id] = ''
                        continue

                    send_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self.delay(3)  # 增加等待时间确保发送完成

                    new_page.close()
                    logger.info("已关闭新页面，返回继续")
                    self.delay(2)

                    # 验证发送
                    if self.verify_delivery:
                        verify_success = self.verify_product_card_delivery(
                            creator_id=creator_id,
                            creator_name=creator_name,
                            campaign_id=campaign_id,
                            product_id=product_id,
                            baseline_messages=merchant_baseline,
                            limit_state=limit_state,
                        )
                    else:
                        verify_success = True
                        logger.info("已配置跳过聊天记录检查，默认判定产品 %s 发送成功", product_id)
                        if limit_state is not None:
                            current_messages = limit_state.get("merchant_messages", 0) or 0
                            limit_state["merchant_messages"] = current_messages + 1
                            if not limit_state.get("creator_replied", False):
                                remaining = limit_state.get("remaining")
                                if isinstance(remaining, int):
                                    limit_state["remaining"] = max(0, remaining - 1)

                    record_entry = {
                        'creator_name': creator_name,
                        'creator_id': creator_id,
                        'campaign_id': campaign_id,
                        'product_id': product_id,
                        'rate': rate_str,
                        'send_card_time': send_time if verify_success else '',
                    }

                    if verify_success:
                        success_count += 1
                        logger.info(f"产品 {product_id} 验证通过，已记录发送时间 {send_time}")
                        self._append_record_to_excel(record_entry)
                        self.sent_history.setdefault(creator_key, {})[product_id] = send_time
                        product_sent_successfully = True
                        break  # 成功了，跳出重试循环
                    else:
                        logger.warning(f"产品 {product_id} 未检测到新增消息，本次判定失败")
                        if retry_attempt == max_product_retries:
                            self._append_record_to_excel(record_entry)
                            self.sent_history.setdefault(creator_key, {})[product_id] = ''
                        failure_for_creator = True

                except Exception as e:
                    logger.error(f"发送产品 {product_id} 失败 (尝试 {retry_attempt}/{max_product_retries}): {e}")
                    if self._is_browser_closed_error(e):
                        raise BrowserClosedError(str(e))
                        
                    if retry_attempt == max_product_retries:
                        # 最后一次重试失败，记录到Excel
                        record_entry = {
                            'creator_name': creator_name,
                            'creator_id': creator_id,
                            'campaign_id': campaign_id,
                            'product_id': product_id,
                            'rate': rate_str,
                            'send_card_time': '',
                        }
                        self._append_record_to_excel(record_entry)
                        self.sent_history.setdefault(creator_key, {})[product_id] = ''
                    failure_for_creator = True

                # 如果当前重试失败但还有重试机会，等待后继续
                if failure_for_creator and retry_attempt < max_product_retries:
                    logger.info(f"产品 {product_id} 发送失败，等待 5 秒后重试...")
                    self.delay(5)
                    continue
                elif product_sent_successfully:
                    break  # 成功发送，跳出重试循环

            # 处理下一个产品前的准备
            if not product_sent_successfully:
                logger.info(f"产品 {product_id} 在 {max_product_retries} 次尝试后仍然失败，继续下一个产品")
                
            if idx < len(product_ids):
                logger.info("准备继续下一个产品...")
                self.delay(2)

        if success_count == len(product_ids):
            logger.info(f"所有 {success_count} 个产品卡片均验证成功")
        else:
            logger.info(f"本次成功发送 {success_count}/{len(product_ids)} 个产品卡片")
        return success_count == len(product_ids)

    def _append_record_to_excel(self, record: Dict[str, Any]) -> None:
        """将单条发送记录立即追加到 Excel。"""
        excel_path = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))) / 'data' / 'card' / 'product_send.xlsx'
        lock_path = excel_path.with_suffix(excel_path.suffix + '.lock')
        headers = ['creator_name', 'creator_id', 'campaign_id', 'product_id', 'rate', 'send_card_time']

        lock = FileLock(str(lock_path))
        with lock:
            if excel_path.exists():
                wb = load_workbook(excel_path)
                sheet = wb.active
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.title = "product_send"
                sheet.append(headers)

            row = [str(record.get(column, '') or '') for column in headers]
            sheet.append(row)
            wb.save(excel_path)
            wb.close()

    def verify_product_card_delivery(
        self,
        creator_id: str,
        creator_name: str,
        campaign_id: str,
        baseline_messages: int,
        product_id: Optional[str] = None,
        limit_state: Optional[Dict[str, Any]] = None,
    ) -> bool:
        """验证聊天中是否出现新的商家消息（产品卡）。"""
        if not self.open_chat_and_screenshot(self.page, creator_id, max_retries=2):
            logger.warning(f"无法重新打开聊天页面验证 {creator_name} 的产品卡发送情况")
            return False

        chat_status = self.analyze_chat_status(self.page)
        new_count = chat_status.get("merchant_messages", baseline_messages)
        creator_replied = chat_status.get("creator_replied", False)
        success = new_count > baseline_messages

        if limit_state is not None:
            limit_state["merchant_messages"] = new_count
            limit_state["creator_replied"] = creator_replied
            if creator_replied:
                limit_state["remaining"] = None
            else:
                limit_state["remaining"] = max(0, 5 - new_count)

        if success:
            logger.info(f"验证成功：商家消息数从 {baseline_messages} 增加到 {new_count}")
            if limit_state and not limit_state.get("creator_replied"):
                logger.info(
                    "达人未回复，剩余可发送消息数：%s/5",
                    limit_state.get("remaining"),
                )
        else:
            logger.warning(f"未检测到新增商家消息（当前 {new_count}），判定发送失败")

        if product_id:
            suffix = "verify_success" if success else "verify_failure"
            # self._save_screenshot(self.page, f"{creator_name}_{product_id}_{suffix}")
        else:
            suffix = "verify_success" if success else "verify_failure"
            # self._save_screenshot(self.page, f"{creator_name}_{suffix}")

        # 验证后重新回到活动页面
        if not self.navigate_to_campaign_page(self.page, creator_id):
            logger.error(f"产品卡验证后返回活动页失败，任务中断：{creator_name}")
            return False
        if not self.search_and_open_campaign(self.page, campaign_id):
            logger.error(f"产品卡验证后重新打开活动详情失败，任务中断：{campaign_id}")
            return False

        return success

    def _record_creator_processed(self, processed_count: int, total_creators: int) -> bool:
        if self.auto_restart_interval <= 0:
            return True
        self.creators_since_restart += 1
        if (
            self.auto_restart_interval > 0
            and self.creators_since_restart >= self.auto_restart_interval
            and processed_count < total_creators
        ):
            logger.info(
                "已连续处理 %s 个达人，触发自动重启浏览器流程...",
                self.creators_since_restart,
            )
            if not self._restart_browser_session():
                logger.error("自动重启浏览器失败，任务终止")
                return False
            self.creators_since_restart = 0
            self.consecutive_browser_closed_errors = 0
            self._restart_anchor_index = None
        return True

    def run(self) -> bool:
        """运行主流程"""
        logger.info("=" * 50)
        logger.info("TikTok Chat and Card Sender")
        logger.info("=" * 50)

        with sync_playwright() as p:
            try:
                self._playwright = p
                if not self._setup_browser(p):
                    return False

                self.creators_since_restart = 0

                # 读取所有达人配置
                all_creators_data = self.load_card_send_data()
                if not all_creators_data:
                    logger.error("未能加载配置数据")
                    return False

                self._load_sent_history()

                total_creators = len(all_creators_data)
                creator_idx = 0

                while creator_idx < total_creators:
                    card_data = all_creators_data[creator_idx]
                    logger.info("=" * 50)
                    logger.info(f"处理第 {creator_idx + 1}/{total_creators} 个达人")
                    logger.info("=" * 50)
                    
                    creator_id = card_data.get('creator_id')
                    creator_name = card_data.get('creator_name')
                    if self._should_skip_creator(card_data):
                        creator_idx += 1
                        self.consecutive_browser_closed_errors = 0
                        self._restart_anchor_index = None
                        if not self._record_creator_processed(creator_idx, total_creators):
                            return False
                        continue
                    
                    try:
                        if not self.open_chat_and_screenshot(self.page, creator_id, max_retries=3):
                            logger.error(f"达人 {creator_name} 聊天页面加载失败，跳过")
                            creator_idx += 1
                            self.consecutive_browser_closed_errors = 0
                            self._restart_anchor_index = None
                            if not self._record_creator_processed(creator_idx, total_creators):
                                return False
                            continue

                        chat_status = self.analyze_chat_status(self.page)
                        creator_replied = chat_status.get("creator_replied", False)
                        remaining = None if creator_replied else max(
                            0, 5 - chat_status.get("merchant_messages", 0)
                        )
                        limit_state = {
                            "creator_replied": creator_replied,
                            "remaining": remaining,
                            "merchant_messages": chat_status.get("merchant_messages", 0),
                        }

                        if creator_replied:
                            logger.info("达人已回复，后续聊天消息不受 5 条限制")
                        else:
                            logger.info(
                                "达人尚未回复，当前仍可发送消息数量：%s/5",
                                limit_state["remaining"],
                            )
                            if (limit_state["remaining"] or 0) <= 0:
                                logger.info("商家消息已达上限，本次任务跳过达人 %s", creator_name)
                                creator_idx += 1
                                self.consecutive_browser_closed_errors = 0
                                self._restart_anchor_index = None
                                if not self._record_creator_processed(creator_idx, total_creators):
                                    return False
                                continue

                        message = (card_data.get('message') or '').strip()
                        if message:
                            self.check_and_send_message(
                                self.page, message, limit_state, chat_status
                            )
                        else:
                            logger.info("消息为空，本次不发送文本消息")

                        if not self.navigate_to_campaign_page(self.page, creator_id):
                            logger.error(f"达人 {creator_name} 活动页面加载失败，跳过")
                            creator_idx += 1
                            self.consecutive_browser_closed_errors = 0
                            self._restart_anchor_index = None
                            if not self._record_creator_processed(creator_idx, total_creators):
                                return False
                            continue

                        campaign_id = card_data.get('campaign_id')
                        if not campaign_id or not self.search_and_open_campaign(self.page, campaign_id):
                            logger.error(f"达人 {creator_name} 活动详情打开失败，跳过")
                            creator_idx += 1
                            self.consecutive_browser_closed_errors = 0
                            self._restart_anchor_index = None
                            if not self._record_creator_processed(creator_idx, total_creators):
                                return False
                            continue

                        product_ids = card_data.get('product_ids', [])
                        rate = card_data.get('rate', 14)
                        send_success = True
                        if product_ids:
                            send_success = self.send_product_cards(
                                self.page,
                                product_ids,
                                rate,
                                creator_name,
                                creator_id,
                                campaign_id,
                                limit_state=limit_state,
                            )

                        if send_success:
                            logger.info(f"达人 {creator_name} 处理完成")
                        else:
                            logger.warning(f"达人 {creator_name} 的商品卡未全部发送成功，进入下一位达人")
                        creator_idx += 1
                        self.consecutive_browser_closed_errors = 0
                        self._restart_anchor_index = None
                        if not self._record_creator_processed(creator_idx, total_creators):
                            return False

                    except BrowserClosedError as closed_error:
                        logger.error(f"检测到浏览器已关闭，处理达人 {creator_name} 时出错: {closed_error}")
                        self.consecutive_browser_closed_errors += 1
                        if self._restart_anchor_index is None:
                            self._restart_anchor_index = creator_idx

                        if self.consecutive_browser_closed_errors >= self.closed_error_threshold:
                            if not self._restart_browser_session():
                                return False
                            creator_idx = self._restart_anchor_index or creator_idx
                            self._restart_anchor_index = None
                            continue
                        else:
                            logger.warning(
                                "浏览器关闭异常未达重启阈值，将在下一个达人继续尝试（当前连续失败次数：%s/%s）",
                                self.consecutive_browser_closed_errors,
                                self.closed_error_threshold,
                            )
                            # 短暂等待再重试同一达人
                            time.sleep(2)
                            continue

                logger.info("=" * 50)
                logger.info("所有达人处理完成")
                logger.info("=" * 50)
                return True

            except Exception as e:
                logger.error(f"程序运行出错: {e}")
                return False
            finally:
                self._teardown_browser()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Send chat messages and product cards via TikTok Shop Partner Center.")
    parser.add_argument(
        '--headless',
        dest='headless',
        action='store_true',
        default=True,
        help='使用无头模式运行浏览器（默认启用）',
    )
    parser.add_argument(
        '--show-browser',
        dest='headless',
        action='store_false',
        help='显示浏览器窗口；若未配合 --manual-login-timeout 指定值，则仍自动登录',
    )
    parser.add_argument(
        '--manual-login-timeout',
        type=int,
        default=None,
        help='在手动登录模式下等待欢迎页面的最长时间（秒）；仅在同时使用 --show-browser 且设置该参数时进入手动登录模式',
    )
    parser.add_argument(
        '--skip-verify',
        dest='skip_verify',
        action='store_true',
        help='发送后跳过聊天消息验证（默认开启验证）',
    )
    parser.add_argument(
        '--card-json',
        type=str,
        default=None,
        help='自定义 card_send_list JSON 路径（默认 data/card/card_send_list.json）',
    )
    parser.add_argument(
        '--region',
        type=str,
        default="MX",
        help='账号/达人所属地区代码，例如 MX 或 FR（默认 MX）',
    )
    parser.add_argument(
        '--account-name',
        type=str,
        default=None,
        help='config/accounts.json 中的账号 name；为空则自动选择对应地区的首个可用账号',
    )
    parser.add_argument(
        '--accounts-config',
        type=str,
        default=str(DEFAULT_ACCOUNTS_JSON),
        help='账号配置文件路径（默认 config/accounts.json）',
    )

    args = parser.parse_args()

    accounts_config_path = Path(args.accounts_config).expanduser()
    account_info = select_account_for_region(args.region, args.account_name, accounts_config_path)
    if not account_info:
        logger.error("无法找到区域 %s 的可用账号，请检查配置文件 %s", args.region.upper(), accounts_config_path)
        sys.exit(1)

    sender = CardSender(
        account_info=account_info,
        region=args.region,
        headless=args.headless,
        manual_login_timeout=args.manual_login_timeout,
        verify_delivery=not args.skip_verify,
        card_data_path=args.card_json,
    )
    sender.run()
