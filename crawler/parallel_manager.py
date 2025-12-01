"""
轻量化的爬虫任务管理器
负责接收 API 请求并调度单个爬虫线程执行。
"""
from __future__ import annotations

import copy
import json
import logging
import time
from concurrent.futures import Future, ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from threading import Event, Lock, Thread
from typing import Any, Dict, List, Optional

from zoneinfo import ZoneInfo
from openpyxl import load_workbook

from database.outreach_tasks_repo import (
    persist_task_snapshot,
    list_pending_task_snapshots,
    cancel_incomplete_tasks_on_startup,
)
from models.account_pool import AccountPool, get_account_pool
from schemas.crawler import (
    CrawlerSummaryResponse,
    CrawlerTaskCreateRequest,
    CrawlerTaskStatus,
    CrawlerTaskUpdateRequest,
)
from .task_worker import CrawlerTaskResult, CrawlerTaskWorker
from .creator_full_crawler import update_task_info_row

logger = logging.getLogger(__name__)


def _now() -> datetime:
    return datetime.now(tz=ZoneInfo("UTC"))


def _deep_update(dest: Dict, updates: Dict) -> Dict:
    for key, value in updates.items():
        if isinstance(value, dict) and isinstance(dest.get(key), dict):
            _deep_update(dest[key], value)
        else:
            dest[key] = value
    return dest


def _format_duration(seconds: int) -> str:
    total = max(0, int(seconds))
    hours = total // 3600
    minutes = (total % 3600) // 60
    secs = total % 60
    return f"{hours:02d}h{minutes:02d}min{secs:02d}s"


@dataclass
class TaskRecord:
    """内部任务记录，保存运行时状态。"""

    task_id: str
    payload: CrawlerTaskCreateRequest
    dify_payload: Dict
    task_dir: Path
    user: str
    task_name: Optional[str]
    campaign_id: Optional[str]
    campaign_name: Optional[str]
    product_name: Optional[str]
    product_id: Optional[str]
    max_creators: int
    target_new: int
    run_at_time: Optional[datetime]
    run_at_time_utc: Optional[datetime]
    run_end_time: Optional[datetime]
    run_end_time_utc: Optional[datetime]
    submitted_at: datetime = field(default_factory=_now)
    status: str = "pending"
    message: Optional[str] = None
    started_at: Optional[datetime] = None
    finished_at: Optional[datetime] = None
    account: Optional[Dict] = None
    new_creators: Optional[int] = None
    total_creators: Optional[int] = None
    output_files: List[str] = field(default_factory=list)
    log_path: Optional[str] = None
    connect_creator: Optional[str] = None
    future: Optional[Future] = field(default=None, repr=False, compare=False)
    cancel_requested: bool = False
    cancel_event: Event = field(default_factory=Event, repr=False, compare=False)
    wait_log_emitted: bool = False
    force_terminated: bool = False

    def to_status(self) -> CrawlerTaskStatus:
        """转换为API层可见的状态模型。"""
        account_email = None
        if self.account:
            account_email = self.account.get("login_email")

        run_time_text: Optional[str] = None
        if self.started_at:
            end_time = self.finished_at or _now()
            try:
                delta = end_time - self.started_at
                run_time_text = _format_duration(int(delta.total_seconds()))
            except Exception:  # pragma: no cover - 防御性
                run_time_text = None

        return CrawlerTaskStatus(
            task_id=self.task_id,
            task_type="Connect",
            status=self.status,
            message=self.message,
            submitted_at=self.submitted_at,
            started_at=self.started_at,
            finished_at=self.finished_at,
            user=self.user,
            task_name=self.task_name,
            campaign_id=self.campaign_id,
            campaign_name=self.campaign_name,
            region=(self.payload.region or None),
            brand_name=self.payload.brand.name if self.payload.brand else None,
            account_email=account_email,
            new_creators=self.new_creators,
            total_creators=self.total_creators,
            task_dir=str(self.task_dir),
            log_path=self.log_path,
            product_name=self.product_name,
            product_id=self.product_id,
            connect_creator=self.connect_creator,
            run_time=run_time_text,
            output_files=self.output_files,
            max_creators=self.max_creators,
            target_new_creators=self.target_new,
            run_at_time=self.run_at_time,
            run_end_time=self.run_end_time,
            payload=self.dify_payload,
        )

    def _format_list(self, values: Optional[List[Any]]) -> str:
        if not values:
            return ""
        return ", ".join(str(item).strip() for item in values if str(item).strip())

    def to_snapshot(self) -> Dict[str, Any]:
        brand = self.payload.brand
        strategy = self.payload.search_strategy
        email_first = self.payload.email_first
        email_later = self.payload.email_later

        run_time_text = None
        if self.started_at:
            end_time = self.finished_at or _now()
            try:
                run_time_text = _format_duration(int((end_time - self.started_at).total_seconds()))
            except Exception:  # pragma: no cover - 防御性
                run_time_text = None

        account_email = None
        if self.account:
            account_email = self.account.get("login_email")

        snapshot = {
            "task_id": self.task_id,
            "task_name": self.task_name,
            "campaign_id": self.campaign_id,
            "campaign_name": self.campaign_name,
            "product_id": self.product_id,
            "product_name": self.product_name,
            "region": self.payload.region,
            "brand": brand.name if brand else None,
            "only_first": (str(brand.only_first) if brand and brand.only_first is not None else None),
            "task_type": "Connect",
            "status": self.status,
            "message": self.message,
            "created_by": self.user,
            "account_email": account_email,
            "search_keywords": strategy.search_keywords or "",
            "product_category": strategy.product_category or "",
            "fans_age_range": self._format_list(strategy.fans_age_range),
            "fans_gender": strategy.fans_gender or "",
            "min_fans": strategy.min_fans,
            "content_type": self._format_list(strategy.content_type),
            "gmv": self._format_list(strategy.gmv),
            "sales": self._format_list(strategy.sales),
            "min_GMV": strategy.min_GMV,
            "max_GMV": strategy.max_GMV,
            "avg_views": strategy.avg_views,
            "min_engagement_rate": strategy.min_engagement_rate,
            "email_first_subject": email_first.subject if email_first else "",
            "email_first_body": email_first.email_body if email_first else "",
            "email_later_subject": email_later.subject if email_later else "",
            "email_later_body": email_later.email_body if email_later else "",
            "target_new_creators": self.target_new,
            "max_creators": self.max_creators,
            "run_at_time": self.run_at_time,
            "run_end_time": self.run_end_time,
            "run_time": run_time_text,
            "task_directory": str(self.task_dir),
            "log_path": self.log_path,
            "connect_creator": self.connect_creator,
            "new_creators": self.new_creators,
            "total_creators": self.total_creators,
            "submitted_at": self.submitted_at,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "payload": copy.deepcopy(self.dify_payload),
            "output_files": list(self.output_files or []),
        }
        return snapshot


class ParallelTaskManager:
    """
    任务管理器：使用线程池异步执行每一次爬虫任务。
    与早期的多进程并行不同，这里强调“按需触发”的执行模式。
    """

    def __init__(
        self,
        max_workers: int = 1,
        task_root: str = "task",
        task_data_dir: str = "data/tasks",
        account_pool_config: str = "config/accounts.json",
        task_timeout_minutes: int = 120,
    ) -> None:
        self.max_workers = max_workers
        self.task_root = Path(task_root)
        self.task_data_dir = Path(task_data_dir)
        self.task_data_dir.mkdir(parents=True, exist_ok=True)
        self.task_timeout = timedelta(minutes=task_timeout_minutes)

        self._executor = ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="crawler-worker")
        self._lock = Lock()
        self._tasks: Dict[str, TaskRecord] = {}
        self._account_pool: AccountPool = get_account_pool(account_pool_config)
        self._known_task_ids: set[str] = set()
        self._sequence_counter: int = 1
        self._load_existing_task_ids()
        cancelled = cancel_incomplete_tasks_on_startup()
        if cancelled:
            logger.info("服务启动：自动将 %s 个未完成任务标记为取消", cancelled)
        self._restore_pending_tasks()

        logger.info(
            "ParallelTaskManager 初始化完成 (max_workers=%s, task_data_dir=%s)",
            max_workers,
            self.task_data_dir,
        )

    def _load_existing_task_ids(self) -> None:
        info_path = Path("data/task_info.xlsx")
        if not info_path.exists():
            return
        try:
            wb = load_workbook(info_path, read_only=True)
            ws = wb.active
            header = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
            if "task_id" in header:
                idx = header.index("task_id")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    task_id = row[idx]
                    if task_id:
                        self._known_task_ids.add(str(task_id))
                numeric = [int(tid) for tid in self._known_task_ids if str(tid).isdigit()]
                if numeric:
                    self._sequence_counter = max(numeric) + 1
            wb.close()
        except Exception as exc:  # pragma: no cover - 容错
            logger.warning("读取历史任务信息失败: %s", exc)

    def _slugify(self, value: Optional[str]) -> str:
        if not value:
            return ""
        cleaned = "".join(ch if ch.isalnum() else "_" for ch in value).strip("_")
        return cleaned or "task"

    def _ensure_unique_task_id(self, provided: Optional[str], fallback_name: Optional[str]) -> str:
        with self._lock:
            if provided:
                normalized = provided.strip()
                if not normalized:
                    raise ValueError("task_id 不能为空字符串")
                if normalized in self._known_task_ids:
                    raise ValueError(f"任务ID {normalized} 已存在")
                self._known_task_ids.add(normalized)
                return normalized

            while True:
                candidate = f"{self._sequence_counter:08d}"
                self._sequence_counter += 1
                if candidate not in self._known_task_ids:
                    self._known_task_ids.add(candidate)
                    return candidate

    def _build_task_directory(self, payload: CrawlerTaskCreateRequest, task_id: str) -> Path:
        brand_name = payload.brand.name if payload.brand else "task"
        brand_slug = self._slugify(brand_name)
        if payload.task_name:
            base_name = self._slugify(payload.task_name)
        else:
            base_name = brand_slug
        dir_name = "_".join(part for part in (base_name, task_id) if part)
        return self.task_data_dir / brand_slug / dir_name

    def _persist_record(self, record: TaskRecord) -> None:
        try:
            persist_task_snapshot(record.to_snapshot())
        except Exception as exc:  # pragma: no cover - 容错
            logger.warning("持久化任务 %s 失败: %s", record.task_id, exc)

    def _build_payload_from_snapshot(self, snapshot: Dict[str, Any]) -> Optional[CrawlerTaskCreateRequest]:
        def _parse_list(value):
            if value in (None, "", []):
                return []
            if isinstance(value, list):
                return [str(item) for item in value]
            if isinstance(value, str):
                try:
                    loaded = json.loads(value)
                    if isinstance(loaded, list):
                        return [str(item) for item in loaded]
                except Exception:  # pragma: no cover
                    pass
                return [item.strip() for item in value.split(",") if item.strip()]
            return [str(value)]

        working_snapshot = dict(snapshot)
        run_at_time = working_snapshot.get("run_at_time")
        if isinstance(run_at_time, datetime) and run_at_time.tzinfo is None:
            working_snapshot["run_at_time"] = run_at_time.replace(tzinfo=ZoneInfo("Asia/Shanghai"))
        run_end_time = working_snapshot.get("run_end_time")
        if isinstance(run_end_time, datetime) and run_end_time.tzinfo is None:
            working_snapshot["run_end_time"] = run_end_time.replace(tzinfo=ZoneInfo("Asia/Shanghai"))

        base_payload = snapshot.get("payload") or {}
        fallback_brand = {
            "name": working_snapshot.get("brand") or "",
            "only_first": working_snapshot.get("only_first") or 0,
            "key_word": base_payload.get("brand", {}).get("key_word", ""),
        }
        fallback_strategy = {
            "search_keywords": working_snapshot.get("search_keywords") or "",
            "product_category": working_snapshot.get("product_category") or "",
            "fans_age_range": _parse_list(working_snapshot.get("fans_age_range")),
            "fans_gender": working_snapshot.get("fans_gender") or "",
            "min_fans": working_snapshot.get("min_fans"),
            "content_type": _parse_list(working_snapshot.get("content_type")),
            "gmv": _parse_list(working_snapshot.get("gmv")),
            "sales": _parse_list(working_snapshot.get("sales")),
            "min_GMV": working_snapshot.get("min_GMV"),
            "max_GMV": working_snapshot.get("max_GMV"),
            "avg_views": working_snapshot.get("avg_views"),
            "min_engagement_rate": working_snapshot.get("min_engagement_rate") or 0,
        }
        fallback_email_first = {
            "subject": working_snapshot.get("email_first_subject") or "",
            "email_body": working_snapshot.get("email_first_body") or "",
        }
        fallback_email_later = {
            "subject": working_snapshot.get("email_later_subject") or fallback_email_first["subject"],
            "email_body": working_snapshot.get("email_later_body") or fallback_email_first["email_body"],
        }

        merged_payload = {
            "task_id": working_snapshot.get("task_id"),
            "task_name": working_snapshot.get("task_name"),
            "region": working_snapshot.get("region"),
            "product_name": working_snapshot.get("product_name"),
            "product_id": working_snapshot.get("product_id"),
            "campaign_id": working_snapshot.get("campaign_id"),
            "campaign_name": working_snapshot.get("campaign_name"),
            "run_at_time": working_snapshot.get("run_at_time"),
            "run_end_time": working_snapshot.get("run_end_time"),
            "brand": fallback_brand,
            "search_strategy": fallback_strategy,
            "email_first": fallback_email_first,
            "email_later": fallback_email_later,
            "max_creators": working_snapshot.get("max_creators") or base_payload.get("max_creators") or 500,
            "target_new_creators": working_snapshot.get("target_new_creators") or base_payload.get("target_new_creators") or 50,
        }

        if base_payload:
            merged_payload = _deep_update(base_payload, merged_payload)

        try:
            return CrawlerTaskCreateRequest(**merged_payload)
        except Exception as exc:  # pragma: no cover
            logger.warning("恢复任务 %s 时构造 payload 失败: %s", snapshot.get("task_id"), exc)
            return None

    def _restore_pending_tasks(self) -> None:
        snapshots = list_pending_task_snapshots()
        if not snapshots:
            return

        for snapshot in snapshots:
            task_id = snapshot.get("task_id")
            if not task_id:
                continue
            if task_id in self._tasks:
                continue

            payload = self._build_payload_from_snapshot(snapshot)
            if not payload:
                continue

            dify_payload = payload.model_dump(mode="json", exclude_none=True)
            dify_payload["task_id"] = task_id
            if payload.run_at_time:
                dify_payload["run_at_time"] = payload.run_at_time.isoformat()
            if payload.run_end_time:
                dify_payload["run_end_time"] = payload.run_end_time.isoformat()

            task_dir = snapshot.get("task_directory") or str(self._build_task_directory(payload, task_id))
            task_dir_path = Path(task_dir)
            task_dir_path.mkdir(parents=True, exist_ok=True)

            record = TaskRecord(
                task_id=task_id,
                payload=payload,
                dify_payload=dify_payload,
                task_dir=task_dir_path,
                user=snapshot.get("created_by") or "system",
                task_name=payload.task_name,
                campaign_id=payload.campaign_id,
                campaign_name=payload.campaign_name,
                product_name=payload.product_name,
                product_id=payload.product_id,
                max_creators=payload.max_creators,
                target_new=payload.target_new_creators,
                run_at_time=payload.run_at_time,
                run_at_time_utc=payload.run_at_time.astimezone(ZoneInfo("UTC")) if payload.run_at_time else None,
                run_end_time=payload.run_end_time,
                run_end_time_utc=payload.run_end_time.astimezone(ZoneInfo("UTC")) if payload.run_end_time else None,
                submitted_at=snapshot.get("submitted_at") or _now(),
                status="pending",
                message="任务恢复后等待执行",
                started_at=None,
                finished_at=None,
                account=None,
                new_creators=snapshot.get("new_creators"),
                total_creators=snapshot.get("total_creators"),
                output_files=snapshot.get("output_files") or [],
                log_path=snapshot.get("log_path"),
                connect_creator=snapshot.get("connect_creator"),
            )

            self._known_task_ids.add(task_id)
            with self._lock:
                if task_id in self._tasks:
                    continue
                self._tasks[task_id] = record
                record.future = self._executor.submit(self._run_task, task_id)

    # ------------------------------------------------------------------
    # 对外接口
    # ------------------------------------------------------------------
    def submit_task(self, payload: CrawlerTaskCreateRequest, created_by: str) -> str:
        """创建任务并提交到执行器。"""

        task_id = self._ensure_unique_task_id(payload.task_id, payload.task_name)
        run_at_time = payload.run_at_time
        if run_at_time:
            now_shanghai = datetime.now(tz=ZoneInfo("Asia/Shanghai"))
            if run_at_time <= now_shanghai:
                run_at_time = None
                run_at_time_utc = None
            else:
                run_at_time_utc = run_at_time.astimezone(ZoneInfo("UTC"))
        else:
            run_at_time_utc = None

        task_dict = payload.model_dump(mode="json", exclude_none=True)
        task_dict["task_id"] = task_id
        if run_at_time:
            task_dict["run_at_time"] = run_at_time.isoformat()
        else:
            task_dict.pop("run_at_time", None)
        run_end_time = payload.run_end_time
        run_end_time_utc = run_end_time.astimezone(ZoneInfo("UTC")) if run_end_time else None
        if run_end_time:
            task_dict["run_end_time"] = run_end_time.isoformat()

        task_dir = self._build_task_directory(payload, task_id)
        task_dir.mkdir(parents=True, exist_ok=True)

        record = TaskRecord(
            task_id=task_id,
            payload=payload,
            dify_payload=task_dict,
            task_dir=task_dir,
            user=created_by,
            task_name=payload.task_name,
            campaign_id=payload.campaign_id,
            campaign_name=payload.campaign_name,
            product_name=payload.product_name,
            product_id=payload.product_id,
            max_creators=payload.max_creators,
            target_new=payload.target_new_creators,
            run_at_time=run_at_time,
            run_at_time_utc=run_at_time_utc,
            run_end_time=run_end_time,
            run_end_time_utc=run_end_time_utc,
            status="pending",
            message="等待执行",
            wait_log_emitted=False,
        )

        with self._lock:
            self._tasks[task_id] = record
            record.future = self._executor.submit(self._run_task, task_id)

        logger.info(
            "已提交任务 %s (brand=%s, region=%s, user=%s)",
            task_id,
            payload.brand.name if payload.brand else "N/A",
            payload.region,
            created_by,
        )
        self._persist_record(record)
        return task_id

    def update_task(self, task_id: str, update_request: CrawlerTaskUpdateRequest) -> None:
        update_dict = update_request.model_dump(mode="json", exclude_unset=True)

        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                raise ValueError("任务不存在")
            if record.status != "pending":
                raise ValueError("任务已开始执行，无法修改")

            if "task_id" in update_dict and update_dict["task_id"] not in (None, task_id):
                raise ValueError("不允许修改 task_id")

            base_data = record.payload.model_dump(mode="json", exclude_none=False)

        merged_data = _deep_update(base_data, {k: v for k, v in update_dict.items() if k != "task_id"})
        merged_data["task_id"] = task_id

        new_payload = CrawlerTaskCreateRequest(**merged_data)
        new_dict = new_payload.model_dump(mode="json", exclude_none=True)
        new_dict["task_id"] = task_id
        new_run_at_time = new_payload.run_at_time
        if new_run_at_time:
            new_dict["run_at_time"] = new_run_at_time.isoformat()
        else:
            new_dict.pop("run_at_time", None)
        new_run_at_time_utc = new_run_at_time.astimezone(ZoneInfo("UTC")) if new_run_at_time else None
        new_run_end_time = new_payload.run_end_time
        if new_run_end_time:
            new_dict["run_end_time"] = new_run_end_time.isoformat()
        else:
            new_dict.pop("run_end_time", None)
        new_run_end_time_utc = new_run_end_time.astimezone(ZoneInfo("UTC")) if new_run_end_time else None

        new_dir = self._build_task_directory(new_payload, task_id)

        with self._lock:
            record = self._tasks.get(task_id)
            if not record or record.status != "pending":
                raise ValueError("任务已开始执行，无法修改")

            old_dir = record.task_dir
            if new_dir != old_dir:
                new_dir.parent.mkdir(parents=True, exist_ok=True)
                if new_dir.exists():
                    raise ValueError("目标任务目录已存在，无法重命名")
                if old_dir.exists():
                    old_dir.rename(new_dir)
                else:
                    new_dir.mkdir(parents=True, exist_ok=True)
            record.task_dir = new_dir
            record.payload = new_payload
            record.dify_payload = new_dict
            record.task_name = new_payload.task_name
            record.campaign_id = new_payload.campaign_id
            record.campaign_name = new_payload.campaign_name
            record.product_name = new_payload.product_name
            record.product_id = new_payload.product_id
            record.max_creators = new_payload.max_creators
            record.target_new = new_payload.target_new_creators
            record.run_at_time = new_run_at_time
            record.run_at_time_utc = new_run_at_time_utc
            record.run_end_time = new_run_end_time
            record.run_end_time_utc = new_run_end_time_utc
            record.message = "任务信息已更新，等待执行"
            record.wait_log_emitted = False

        logger.info("任务 %s 已更新，新的执行时间: %s", task_id, new_run_at_time)
        self._persist_record(record)

    def update_task_name(self, task_id: str, task_name: str) -> None:
        new_name = (task_name or "").strip()
        if not new_name:
            raise ValueError("任务名称不能为空")

        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                raise ValueError("任务不存在")
            if record.status != "pending":
                raise ValueError("任务已开始执行，无法修改")

            record.payload = record.payload.model_copy(update={"task_name": new_name})
            record.task_name = new_name
            record.dify_payload["task_name"] = new_name
            record.message = "任务名称已更新"
        self._persist_record(record)

    def run_task_now(self, task_id: str) -> None:
        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                raise ValueError("任务不存在")
            if record.status != "pending":
                raise ValueError("任务已开始执行，无法立即执行")
            now_utc = _now()
            record.run_at_time_utc = now_utc
            record.run_at_time = now_utc.astimezone(ZoneInfo("Asia/Shanghai"))
            record.status = "to-be-run"
            record.message = "任务已设为立即执行，等待启动"
            record.wait_log_emitted = False
        logger.info("任务 %s 已被设置为立即执行", task_id)
        self._persist_record(record)

    def force_cancel_task(self, task_id: str) -> bool:
        """强制终止任务（无论是否已开始执行）。"""
        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                return False

            record.cancel_requested = True
            record.force_terminated = True
            if record.status != "cancelled":
                record.status = "to-be-cancel"
            record.message = "强制终止请求已发送"
            record.cancel_event.set()

            if record.status == "pending" and record.future:
                if record.future.cancel():
                    record.status = "cancelled"
                    record.finished_at = _now()
        if not record:
            return False

        self._persist_record(record)
        return True


    def list_tasks(
        self,
        *,
        brand_name: Optional[str] = None,
        region: Optional[str] = None,
        status: Optional[str] = None,
        task_name: Optional[str] = None,
        run_at_time: Optional[datetime] = None,
        run_end_time: Optional[datetime] = None,
        page: int = 1,
        page_size: int = 20,
        sort: Optional[str] = None,
    ) -> tuple[List[CrawlerTaskStatus], int]:
        """按条件过滤并分页返回任务列表。"""
        with self._lock:
            records = list(self._tasks.values())

        brand_lower = brand_name.lower() if brand_name else None
        region_lower = region.lower() if region else None
        status_lower = status.lower() if status else None
        task_name_lower = task_name.lower() if task_name else None

        def _normalize_to_tz(dt: Optional[datetime]) -> Optional[datetime]:
            if dt is None:
                return None
            try:
                return dt.astimezone(ZoneInfo("Asia/Shanghai"))
            except Exception:  # pragma: no cover
                return dt

        run_at_time_local = _normalize_to_tz(run_at_time)
        run_end_time_local = _normalize_to_tz(run_end_time)

        filtered: List[TaskRecord] = []
        for record in records:
            payload_brand = (record.payload.brand.name if record.payload.brand else "") or ""
            if brand_lower and brand_lower not in payload_brand.lower():
                continue
            if region_lower and region_lower != (record.payload.region or "").lower():
                continue
            if status_lower and status_lower != record.status.lower():
                continue
            if task_name_lower:
                candidate_name = (record.task_name or "") or ""
                if task_name_lower not in candidate_name.lower():
                    continue

            if run_at_time_local:
                candidate = record.run_at_time or record.started_at
                if not candidate:
                    continue
                candidate_local = _normalize_to_tz(candidate)
                if not candidate_local or candidate_local < run_at_time_local:
                    continue

            if run_end_time_local:
                candidate_end = record.run_end_time or record.finished_at
                if not candidate_end:
                    continue
                candidate_end_local = _normalize_to_tz(candidate_end)
                if not candidate_end_local or candidate_end_local > run_end_time_local:
                    continue

            filtered.append(record)

        total = len(filtered)

        def _timestamp(dt: Optional[datetime], *, default_high: bool) -> float:
            if dt is None:
                return float("inf") if default_high else float("-inf")
            try:
                return dt.timestamp()
            except Exception:  # pragma: no cover
                return float("inf") if default_high else float("-inf")

        def _start_dt(rec: TaskRecord) -> Optional[datetime]:
            return rec.run_at_time or rec.started_at or rec.submitted_at

        def _end_dt(rec: TaskRecord) -> Optional[datetime]:
            return rec.run_end_time or rec.finished_at

        def _run_seconds(rec: TaskRecord) -> int:
            if rec.started_at:
                end_time = rec.finished_at or _now()
                try:
                    return max(0, int((end_time - rec.started_at).total_seconds()))
                except Exception:  # pragma: no cover
                    return 0
            return 0

        if sort == "startAsc":
            filtered.sort(key=lambda rec: _timestamp(_start_dt(rec), default_high=True))
        elif sort == "startDesc":
            filtered.sort(key=lambda rec: _timestamp(_start_dt(rec), default_high=False), reverse=True)
        elif sort == "endAsc":
            filtered.sort(key=lambda rec: _timestamp(_end_dt(rec), default_high=True))
        elif sort == "endDesc":
            filtered.sort(key=lambda rec: _timestamp(_end_dt(rec), default_high=False), reverse=True)
        elif sort == "timeAsc":
            filtered.sort(key=lambda rec: _run_seconds(rec))
        elif sort == "timeDesc":
            filtered.sort(key=lambda rec: _run_seconds(rec), reverse=True)
        else:
            filtered.sort(key=lambda rec: rec.submitted_at, reverse=True)

        safe_page = max(1, page)
        safe_page_size = max(1, min(200, page_size))
        start_index = (safe_page - 1) * safe_page_size
        end_index = start_index + safe_page_size
        page_records = filtered[start_index:end_index]

        return ([record.to_status() for record in page_records], total)

    def get_task(self, task_id: str) -> Optional[CrawlerTaskStatus]:
        """获取单个任务状态。"""
        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                return None
            return record.to_status()

    def cancel_task(self, task_id: str) -> bool:
        """尝试取消尚未开始执行的任务。"""
        with self._lock:
            record = self._tasks.get(task_id)
            if not record:
                return False

            if record.status == "cancelled":
                return False

            persist_needed = False
            result = False

            if record.status in {"pending", "to-be-run"}:
                record.cancel_requested = True
                if record.future and record.future.cancel():
                    record.status = "cancelled"
                    record.finished_at = _now()
                    record.message = "任务已取消"
                    logger.info("任务 %s 已成功取消", task_id)
                    persist_needed = True
                    result = True
                else:
                    record.cancel_requested = False

            if record.status in {"running", "pending", "to-be-run"} and not persist_needed:
                record.cancel_requested = True
                record.status = "to-be-cancel"
                record.message = "取消请求已发送，等待任务退出"
                record.cancel_event.set()
                logger.info("任务 %s 正在取消中", task_id)
                persist_needed = True
                result = True

        if persist_needed:
            self._persist_record(record)
        return result

    def get_summary(self) -> CrawlerSummaryResponse:
        """整体统计信息。"""
        with self._lock:
            statuses = [record.status for record in self._tasks.values()]

        def count(status: str) -> int:
            return statuses.count(status)

        return CrawlerSummaryResponse(
            total=len(statuses),
            pending=count("pending"),
            to_be_run=count("to-be-run"),
            running=count("running"),
            to_be_cancel=count("to-be-cancel"),
            completed=count("completed"),
            failed=count("failed"),
            cancelled=count("cancelled"),
            in_queue=count("pending") + count("to-be-run"),
        )

    # 内部实现
    def _run_task(self, task_id: str) -> None:
        """在线程池中执行的实际任务入口。"""
        with self._lock:
            record = self._tasks.get(task_id)
            if record and record.status == "pending":
                if record.run_at_time_utc is None or record.run_at_time_utc <= _now():
                    record.status = "to-be-run"
                    if not record.message or record.message == "等待执行":
                        record.message = "等待分配执行资源"
        if not record:
            logger.error("任务 %s 未找到", task_id)
            return
        self._persist_record(record)

        def _handle_progress(progress: Any) -> None:
            if progress is None:
                return

            latest_creator: Optional[str] = None
            new_creators_count: Optional[int] = None

            if isinstance(progress, dict):
                latest_creator = progress.get("latest_creator")
                value = progress.get("new_creators")
                if value is not None:
                    try:
                        new_creators_count = int(value)
                    except (TypeError, ValueError):  # pragma: no cover - 容错
                        new_creators_count = None
            else:
                latest_creator = str(progress)

            if not latest_creator and new_creators_count is None:
                return

            with self._lock:
                current = self._tasks.get(task_id)
                if not current:
                    return
                if latest_creator:
                    current.connect_creator = latest_creator
                if new_creators_count is not None:
                    current.new_creators = new_creators_count

                    # 在运行时同步写入 task_info.xlsx
                    try:
                        update_task_info_row(task_id, {"new_creators": new_creators_count})
                    except Exception as exc:  # pragma: no cover - 容错
                        logger.warning("更新任务 %s new_creators 失败: %s", task_id, exc)

        while True:
            scheduled = record.run_at_time_utc
            if scheduled is None:
                break
            now = _now()
            if record.cancel_event.is_set():
                with self._lock:
                    record.status = "cancelled"
                    record.finished_at = now
                    record.message = "任务在执行前被取消"
                self._persist_record(record)
                return
            delta = (scheduled - now).total_seconds()
            if delta <= 0:
                break
            wait_seconds = min(5, max(1, delta))
            if not record.wait_log_emitted:
                logger.info(
                    "任务 %s 计划于 %s 执行（北京时间）",
                    task_id,
                    record.run_at_time.astimezone(ZoneInfo("Asia/Shanghai")) if record.run_at_time else scheduled,
                )
                record.wait_log_emitted = True
            else:
                logger.debug(
                    "任务 %s 等待执行，剩余 %.0f 秒",
                    task_id,
                    delta,
                )
            time.sleep(wait_seconds)

        if record.cancel_requested:
            with self._lock:
                record.status = "cancelled"
                record.finished_at = _now()
                record.message = "任务在执行前被取消"
            self._persist_record(record)
            return

        with self._lock:
            record.status = "running"
            record.started_at = _now()
            record.message = "开始执行"
            record.new_creators = 0
        self._persist_record(record)

        watcher_thread: Optional[Thread] = None
        if record.run_end_time_utc:
            def _watch_run_end_time() -> None:
                while True:
                    if record.cancel_event.is_set():
                        return
                    now = _now()
                    remaining = (record.run_end_time_utc - now).total_seconds()
                    if remaining <= 0:
                        logger.info("任务 %s 达到 run_end_time，触发自动终止", task_id)
                        with self._lock:
                            current = self._tasks.get(task_id)
                            if current:
                                current.cancel_requested = True
                                if not current.cancel_event.is_set():
                                    current.cancel_event.set()
                                if not current.message or "run_end_time" not in (current.message or ""):
                                    current.message = "达到 run_end_time，任务自动结束"
                        return
                    time.sleep(min(30, max(1, remaining / 5)))

            watcher_thread = Thread(target=_watch_run_end_time, name=f"run-end-{task_id}", daemon=True)
            watcher_thread.start()

        account = None
        try:
            region = (record.payload.region or "").upper()
            if region:
                account = self._account_pool.acquire_account_by_region(task_id, region)
            else:
                account = self._account_pool.acquire_account(task_id)

            if not account:
                raise RuntimeError(f"无法为任务 {task_id} 分配可用账号")

            with self._lock:
                record.account = {
                    "id": account.get("id"),
                    "name": account.get("name"),
                    "login_email": account.get("login_email"),
                    "region": account.get("region"),
                }

            task_config = copy.deepcopy(record.dify_payload)
            worker = CrawlerTaskWorker(
                task_id=task_id,
                task_dir=record.task_dir,
                task_config=task_config,
                max_creators=record.max_creators,
                target_new=record.target_new,
                account_info=account,
                cancel_event=record.cancel_event,
                progress_callback=_handle_progress,
            )
            result: CrawlerTaskResult = worker.run()

            with self._lock:
                record.new_creators = result.new_creators
                record.total_creators = result.total_creators
                record.output_files = result.output_files
                record.log_path = worker.log_path
                record.finished_at = _now()
                record.message = result.message
                if result.latest_creator:
                    record.connect_creator = result.latest_creator
                if record.force_terminated or result.cancelled:
                    record.status = "cancelled"
                else:
                    record.status = "completed" if result.success else "failed"
                record.cancel_requested = False
                record.force_terminated = False

        except Exception as exc:  # pylint: disable=broad-except
            logger.exception("任务 %s 执行失败: %s", task_id, exc)
            with self._lock:
                record.status = "failed"
                record.message = str(exc)
                record.finished_at = _now()
                if record.log_path is None:
                    log_path = record.task_dir / "crawler.log"
                    if log_path.exists():
                        record.log_path = str(log_path)

        finally:
            if watcher_thread and watcher_thread.is_alive():
                record.cancel_event.set()
                watcher_thread.join(timeout=0.5)

            if account:
                self._account_pool.release_account(account.get("id"), task_id)

            # 兜底：记录完成时间、日志路径
            with self._lock:
                if record.finished_at is None:
                    record.finished_at = _now()
                if record.log_path is None:
                    log_path = record.task_dir / "crawler.log"
                    if log_path.exists():
                        record.log_path = str(log_path)

                # 超时提示
                if (
                    record.status == "running"
                    and record.started_at
                    and _now() - record.started_at > self.task_timeout
                ):
                    record.status = "failed"
                    record.message = f"任务超时（> {self.task_timeout.total_seconds() / 60:.0f} 分钟）"
                    record.finished_at = _now()
                if (
                    record.status == "cancelled"
                    and record.run_end_time
                    and record.cancel_event.is_set()
                ):
                    record.message = record.message or "达到 run_end_time，任务自动结束"
                if record.status == "cancelled" and record.force_terminated:
                    record.message = record.message or "任务已强制终止"
                    record.force_terminated = False

                connect_snapshot = record.connect_creator
                started_at = record.started_at
                finished_at = record.finished_at
                run_end_time_value = record.run_end_time

            if started_at:
                try:
                    duration = max(0, int(((finished_at or _now()) - started_at).total_seconds()))
                    update_task_info_row(task_id, {"run_time": _format_duration(duration)})
                except Exception as exc:  # pragma: no cover
                    logger.warning("更新任务 %s 运行时间失败: %s", task_id, exc)
            if run_end_time_value:
                try:
                    update_task_info_row(task_id, {"run_end_time": run_end_time_value})
                except Exception as exc:  # pragma: no cover
                    logger.warning("更新任务 %s run_end_time 失败: %s", task_id, exc)
            if connect_snapshot:
                try:
                    update_task_info_row(task_id, {"connect_creator": connect_snapshot})
                except Exception as exc:  # pragma: no cover
                    logger.warning("更新任务 %s connect_creator 失败: %s", task_id, exc)

            self._persist_record(record)
