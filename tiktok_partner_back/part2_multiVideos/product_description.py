# -*- coding: utf-8 -*-
r"""
product_description.py（Excel 输出版，带自动换行）
从 TikTok Shop 商品页抓取 "About this product" 四栏：
- origin（Country/País de origen/产地）
- imported（Imported/Productos importados/是否进口）
- warranty（Warranty/Tipo de garantía/保修）
- description（从 Product description 开始的正文，保留换行与列表）
并抓取标题 title。输出到 --outdir/tiktok_shop_products.xlsx

改进：
- 输出 Excel（.xlsx）
- 自动换行（wrap_text）
- description 等长文本会按 max_len 自动分行，避免一行太长

用法示例：
python product_description.py ^
  --shop_url "https://shop.tiktok.com/view/product/1732196378595133383?region=MX&locale=en" ^
  --outdir "D:\Tiktok\workflow_output\description" ^
  --user-data-dir "C:\Users\YOU\AppData\Local\Microsoft\Edge\User Data" ^
  --headless
"""
import re, time, os, argparse
import pandas as pd
from typing import Optional, List, Dict
from bs4 import BeautifulSoup, Tag
from playwright.sync_api import sync_playwright

ABOUT_HEADERS = [
    "about this product", "sobre este producto", "acerca de este producto",
    "sobre o produto", "关于此商品", "关于这个商品", "thông tin sản phẩm", "à propos de ce produit"
]
LABEL_ORIGIN = [
    "país de origen", "country of origin", "país de origem",
    "产地", "xuất xứ"
]
LABEL_IMPORTED = [
    "productos importados", "imported products", "produtos importados",
    "是否进口", "hàng nhập khẩu"
]
LABEL_WARRANTY = [
    "tipo de garantía", "warranty type", "tipo de garantia",
    "保修", "bảo hành"
]
DESC_HEADERS = [
    "product description", "descripción del producto", "descrição do produto",
    "商品描述", "产品描述", "mô tả sản phẩm", "detalles del producto" ,  "Description du produit"
]
SHOW_MORE_LABELS = [
    "show more", "ver más", "ver mas", "ver mais", "显示更多", "展开",
    "xem thêm", "mở rộng"
]

def norm(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()

def to_lines(s: str) -> str:
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def html_to_text_preserve(html: str) -> str:
    soup = BeautifulSoup(html or "", "html.parser")
    for br in soup.find_all("br"):
        br.replace_with("\n")
    for p in soup.find_all("p"):
        if p.text and not p.text.endswith("\n"):
            p.append("\n")
    for li in soup.find_all("li"):
        li.insert(0, " - ")
        if not li.text.endswith("\n"):
            li.append("\n")
    return to_lines(soup.get_text())

def contains_any(text: str, keys: List[str]) -> bool:
    low = (text or "").lower()
    return any(k in low for k in keys)

def click_show_more_if_any(page):
    for label in SHOW_MORE_LABELS:
        try:
            page.get_by_role("button", name=re.compile(label, re.I)).click(timeout=1200)
            time.sleep(0.3)
        except Exception:
            pass
        try:
            page.locator(
                "xpath=//*[contains(translate(normalize-space(.),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),"
                f" '{label}')]"
            ).first.click(timeout=800)
            time.sleep(0.3)
        except Exception:
            pass

def find_about_container_soup(page) -> Optional[BeautifulSoup]:
    node = page.locator(
        "xpath=//*[self::h2 or self::h3 or self::div or self::span]"
    ).filter(
        has_text=re.compile("|".join([re.escape(h) for h in ABOUT_HEADERS]), re.I)
    ).first
    if node.count() == 0:
        return None
    container = node.locator("xpath=..")
    try:
        html = container.inner_html(timeout=2500)
    except Exception:
        html = page.locator("body").inner_html(timeout=2500)
    return BeautifulSoup(html, "html.parser")

def extract_label_value_from_lines(lines: List[str], label_keys: List[str]) -> str:
    for i, line in enumerate(lines):
        raw = line.strip()
        low = raw.lower()
        if any(k in low for k in label_keys):
            if ":" in raw or "：" in raw or " - " in raw:
                val = re.split(r"[:：-]", raw, maxsplit=1)[-1]
                val = norm(val)
                if val and not contains_any(val.lower(), label_keys):
                    return val
            if i + 1 < len(lines):
                nxt = norm(lines[i + 1])
                if nxt and not contains_any(nxt.lower(), LABEL_ORIGIN + LABEL_IMPORTED + LABEL_WARRANTY + DESC_HEADERS):
                    return nxt
    return ""

def extract_description_from_about_soup(soup: BeautifulSoup) -> str:
    target: Optional[Tag] = None
    for node in soup.find_all(string=True):
        t = str(node).strip()
        if not t:
            continue
        if contains_any(t.lower(), DESC_HEADERS):
            target = node.parent if isinstance(node, Tag) else getattr(node, "parent", None)
            if target:
                break
    if not target:
        return ""

    buf: List[str] = []
    container = target.parent
    if not container:
        return ""

    found_desc_header = False
    for element in container.children:
        if isinstance(element, str):
            continue
        if not found_desc_header:
            if element == target or (hasattr(element, 'find') and element.find(string=re.compile('|'.join(DESC_HEADERS), re.I))):
                found_desc_header = True
            continue
        if isinstance(element, Tag):
            text = element.get_text(" ", strip=True)
            if text:
                if contains_any(text.lower(), LABEL_ORIGIN + LABEL_IMPORTED + LABEL_WARRANTY + ABOUT_HEADERS):
                    if len(text) > 20:
                        break
                content = html_to_text_preserve(str(element))
                if content.strip():
                    buf.append(content.strip())

    text = "\n".join(buf).strip()
    if not text:
        full_text = soup.get_text("\n")
        lines = full_text.split("\n")
        desc_start = -1
        for i, line in enumerate(lines):
            if contains_any(line.lower(), DESC_HEADERS):
                desc_start = i
                break
        if desc_start >= 0:
            desc_lines = []
            for i in range(desc_start + 1, len(lines)):
                line = lines[i].strip()
                if not line:
                    continue
                if contains_any(line.lower(), LABEL_ORIGIN + LABEL_IMPORTED + LABEL_WARRANTY + ABOUT_HEADERS):
                    if len(line) <= 30:
                        break
                desc_lines.append(line)
            text = "\n".join(desc_lines)
    return to_lines(text)

def scrape_about_four(page) -> Dict[str, str]:
    data = {"origin": "", "imported": "", "warranty": "", "description": ""}
    soup = find_about_container_soup(page)
    if not soup:
        return data
    plain = soup.get_text("\n")
    lines = [l.strip() for l in plain.split("\n") if l.strip()]
    data["origin"] = extract_label_value_from_lines(lines, LABEL_ORIGIN)
    data["imported"] = extract_label_value_from_lines(lines, LABEL_IMPORTED)
    data["warranty"] = extract_label_value_from_lines(lines, LABEL_WARRANTY)
    data["description"] = extract_description_from_about_soup(soup)
    return {k: norm(v) if k != "description" else to_lines(v) for k, v in data.items()}

def scrape_one(url: str, user_data_dir: Optional[str], headless: bool, timeout_ms: int) -> Dict[str, str]:
    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(user_data_dir, headless=headless) if user_data_dir else None
        browser = ctx or p.chromium.launch(headless=headless)
        try:
            page = browser.new_page() if ctx is None else (browser.pages[0] if browser.pages else browser.new_page())
            page.set_default_timeout(timeout_ms)
            page.goto(url, wait_until="domcontentloaded")

            for _ in range(4):
                page.evaluate("() => window.scrollBy(0, document.body.scrollHeight * 0.35)")
                time.sleep(0.7)

            click_show_more_if_any(page)

            title = None
            for sel in ["meta[property='og:title']", "meta[name='twitter:title']", "meta[name='title']"]:
                try:
                    el = page.locator(sel).first
                    if el.count():
                        title = el.evaluate("e => e.content")
                        if title:
                            break
                except Exception:
                    pass
            if not title:
                for sel in ['[data-e2e="product-title"]', '[data-e2e="pdp-title"]', "h1", ".product-title",
                            ".pdp-title", "header h1"]:
                    try:
                        el = page.locator(sel).first
                        if el.count():
                            title = el.inner_text(timeout=2000)
                            if title:
                                break
                    except Exception:
                        pass
            title = norm(title or "")

            about = scrape_about_four(page)

            return {
                "url": url,
                "title": title,
                "origin": about["origin"],
                "imported": about["imported"],
                "warranty": about["warranty"],
                "description": about["description"],
            }
        finally:
            try:
                browser.close()
            except Exception:
                pass

def main(args=None):
    parser = argparse.ArgumentParser(description="抓取 TikTok Shop 商品页关于信息到 Excel")
    parser.add_argument("--shop_url", action="append", required=True, help="可多次传入商品链接")
    parser.add_argument("--outdir", required=True, help="输出目录（将写入 tiktok_shop_products.xlsx）")
    parser.add_argument("--user-data-dir", help="Edge/Chrome 用户数据目录（可复用登录态）")
    parser.add_argument("--headless", action="store_true", help="无头模式")
    parser.add_argument("--timeout-ms", type=int, default=20000, help="页面超时（毫秒）")
    if args is None:
        args = parser.parse_args()

    os.makedirs(args.outdir, exist_ok=True)
    out_xlsx = os.path.join(args.outdir, "tiktok_shop_products.xlsx")

    rows = []
    for u in args.shop_url:
        try:
            rows.append(scrape_one(u, args.user_data_dir, args.headless, args.timeout_ms))
        except Exception as e:
            rows.append({
                "url": u, "title": "", "origin": "", "imported": "", "warranty": "",
                "description": "", "error": str(e)
            })

    df = pd.DataFrame(rows)

    # 保存为 Excel
    df.to_excel(out_xlsx, index=False, engine="openpyxl")

    # === 自动换行 & 分行格式化 ===
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(out_xlsx)
    ws = wb.active

    max_len = 80  # 每行最大字符数

    for row in ws.iter_rows(min_row=2):  # 跳过表头
        for cell in row:
            if isinstance(cell.value, str):
                # 自动换行 + 顶端对齐
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # 长文本强制分行
                if len(cell.value) > max_len and "\n" not in cell.value:
                    text = "\n".join(
                        [cell.value[i:i+max_len] for i in range(0, len(cell.value), max_len)]
                    )
                    cell.value = text

    wb.save(out_xlsx)
    print(f"✅ Done with formatting. Saved to: {out_xlsx}")

    try:
        print(df[["url", "title", "origin", "imported", "warranty"]])
    except Exception:
        pass

if __name__ == "__main__":
    main()
